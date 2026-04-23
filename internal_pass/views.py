from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Q, Count
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import os
from django.conf import settings as django_settings
from reportlab.lib.pagesizes import A5, A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from .models import InternalGatePass, GatePassApproval, get_workflow_stages, BYPASS_ROLES
from accounts.notification_service import send_workflow_notification

from .forms import InternalGatePassForm, ApprovalForm, ReturnForm
from accounts.models import Employee
from datetime import date


# ─── Helpers ────────────────────────────────────────────────────────────────

def _base_url(request=None):
    # Always prefer SITE_BASE_URL setting for email links (works from any IP)
    site_url = getattr(settings, 'SITE_BASE_URL', None)
    if site_url:
        return site_url.rstrip('/')
    # Only fall back to request if SITE_BASE_URL is not configured
    return f"{request.scheme}://{request.get_host()}" if request else 'http://localhost:8000'


def _find_approver_for_role(role, department=None):
    """Find an employee with the given role. For dept_hod, match department."""
    qs = Employee.objects.filter(
        Q(role=role, is_active=True) |
        Q(additional_roles__contains=f'|{role}|', is_active=True)
    )
    if role == 'department_hod' and department:
        qs = qs.filter(
            Q(department=department) |
            Q(additional_departments__contains=f'|{department}|')
        )
    return qs.first()


def _create_approval_stages(gate_pass):
    """Create GatePassApproval records for all stages of the workflow."""
    stages = get_workflow_stages(gate_pass.employee.role)
    for i, (label, role) in enumerate(stages):
        GatePassApproval.objects.get_or_create(
            gate_pass=gate_pass,
            stage=i,
            defaults={'stage_label': label, 'approver_role': role}
        )


def _build_email_header(title, subtitle, pass_number, icon='&#128196;'):
    """Shared extraordinary email header HTML."""
    return f"""
  <tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:4px;"></td></tr>
  <tr>
    <td style="background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 50%,#1e40af 100%);padding:36px 40px;text-align:center;">
      <div style="font-size:11px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;">Unity Cement &mdash; ERP Department</div>
      <div style="width:64px;height:64px;border-radius:50%;background:rgba(255,255,255,0.12);border:2px solid rgba(255,255,255,0.25);display:inline-flex;align-items:center;justify-content:center;margin-bottom:16px;font-size:28px;">{icon}</div>
      <div style="font-size:26px;font-weight:800;color:white;margin-bottom:6px;letter-spacing:-0.5px;">{title}</div>
      <div style="font-size:14px;color:rgba(255,255,255,0.7);margin-bottom:16px;">{subtitle}</div>
      <div style="display:inline-block;background:rgba(240,165,0,0.2);border:1.5px solid rgba(240,165,0,0.5);border-radius:50px;padding:6px 20px;font-size:14px;color:#fbbf24;font-weight:700;letter-spacing:1px;">{pass_number}</div>
    </td>
  </tr>"""


def _build_email_footer():
    return """
  <tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:3px;"></td></tr>
  <tr>
    <td style="background:#f8fafc;padding:20px 40px;text-align:center;">
      <p style="font-size:11px;color:#94a3b8;margin:0;">This is an automated email from <strong>ERP Department &mdash; Unity Cement (PMC Cement Ltd.)</strong></p>
      <p style="font-size:11px;color:#94a3b8;margin:6px 0 0;">Do not reply to this email &bull; For support contact IT/ERP Department</p>
    </td>
  </tr>"""


def _build_workflow_progress(all_stages, current_stage_pk):
    """Build horizontal workflow progress bar HTML."""
    cells = []
    for i, s in enumerate(all_stages):
        is_current = s.pk == current_stage_pk
        if s.status == 'approved':
            bg, icon, fg, lbl = '#d1fae5', '&#10003;', '#065f46', 'APPROVED'
        elif s.status == 'rejected':
            bg, icon, fg, lbl = '#fee2e2', '&#10007;', '#991b1b', 'REJECTED'
        elif is_current:
            bg, icon, fg, lbl = '#fef3c7', '&#9654;', '#92400e', 'AWAITING'
        else:
            bg, icon, fg, lbl = '#f1f5f9', '&#9679;', '#94a3b8', 'PENDING'
        border = 'border:2px solid #2563eb;box-shadow:0 0 0 3px rgba(37,99,235,0.15);' if is_current else ''
        short = s.stage_label.split(' ')[0]
        cell = (
            '<td style="text-align:center;padding:0 6px;">'
            '<div style="display:inline-flex;flex-direction:column;align-items:center;gap:5px;">'
            '<div style="width:38px;height:38px;border-radius:50%;background:' + bg + ';color:' + fg + ';'
            'display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;' + border + '">' + icon + '</div>'
            '<div style="font-size:10px;font-weight:600;color:' + fg + ';white-space:nowrap;">' + short + '</div>'
            '<div style="font-size:9px;color:#94a3b8;">' + lbl + '</div>'
            '</div></td>'
        )
        cells.append(cell)
        if i < len(all_stages) - 1:
            conn_color = '#10b981' if s.status == 'approved' else '#e2e8f0'
            cells.append('<td style="width:20px;"><div style="height:2px;background:' + conn_color + ';margin:0 2px;margin-bottom:28px;"></div></td>')
    return ''.join(cells)


def _build_details_table(rows):
    """Build a styled details table from list of (label, value) tuples."""
    html = '<div style="border-radius:12px;overflow:hidden;border:1px solid #e2e8f0;">'
    html += '<div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);padding:12px 16px;">'
    html += '<span style="font-size:11px;font-weight:700;color:white;letter-spacing:1.5px;text-transform:uppercase;">&#128203; Details</span></div>'
    html += '<table width="100%" cellpadding="0" cellspacing="0">'
    for i, (lbl, val) in enumerate(rows):
        bg = 'white' if i % 2 == 0 else '#f8fafc'
        html += (
            '<tr style="background:' + bg + ';">'
            '<td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;width:38%;border-bottom:1px solid #f1f5f9;">' + str(lbl) + '</td>'
            '<td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">' + str(val) + '</td>'
            '</tr>'
        )
    html += '</table></div>'
    return html


def _build_action_buttons(approve_url, reject_url):
    return """
      <div style="background:#f8fafc;border-radius:12px;padding:24px;border:1px solid #e2e8f0;">
        <p style="font-size:13px;color:#475569;font-weight:600;margin:0 0 16px;text-align:center;">&#9889; Take Action Now</p>
        <table cellpadding="0" cellspacing="0" width="100%%">
          <tr>
            <td width="47%%" align="center">
              <a href="%s" style="display:block;background:linear-gradient(135deg,#059669,#10b981);color:white;text-decoration:none;padding:16px 24px;border-radius:12px;font-size:16px;font-weight:800;text-align:center;box-shadow:0 4px 14px rgba(16,185,129,0.4);">&#10003;&nbsp;&nbsp;APPROVE</a>
            </td>
            <td width="6%%"></td>
            <td width="47%%" align="center">
              <a href="%s" style="display:block;background:linear-gradient(135deg,#dc2626,#ef4444);color:white;text-decoration:none;padding:16px 24px;border-radius:12px;font-size:16px;font-weight:800;text-align:center;box-shadow:0 4px 14px rgba(239,68,68,0.4);">&#10007;&nbsp;&nbsp;REJECT</a>
            </td>
          </tr>
        </table>
        <p style="font-size:11px;color:#94a3b8;text-align:center;margin:14px 0 0;">Clicking a button opens a secure confirmation page where you can add remarks.</p>
      </div>""" % (approve_url, reject_url)



def _wrap_email(body_html):
    return (
        "<!DOCTYPE html><html><head><meta charset=\"UTF-8\">"
        + "<body style=\"margin:0;padding:0;background:#f0f4f8;font-family:Inter,Arial,sans-serif;\">"
        + "<table width=\"100%\" cellpadding=\"0\" cellspacing=\"0\" style=\"background:#f0f4f8;padding:32px 0;\">"
        + "<tr><td align=\"center\">"
        + "<table width=\"620\" cellpadding=\"0\" cellspacing=\"0\" style=\"background:white;border-radius:20px;overflow:hidden;\">"
        + body_html
        + "</table></td></tr></table></body></html>"
    )

def _send_stage_mail(gate_pass, stage_obj, request=None):
    """Send extraordinary HTML approval request email with Approve/Reject buttons."""
    try:
        base = _base_url(request)
        approver = _find_approver_for_role(
            stage_obj.approver_role,
            department=gate_pass.employee.department
        )
        if not approver or not approver.email:
            return

        approve_url = base + '/internal-pass/stage-action/' + stage_obj.token + '/approve/'
        reject_url  = base + '/internal-pass/stage-action/' + stage_obj.token + '/reject/'
        total_stages = gate_pass.approvals.count()
        subject = '[Action Required] IGP Approval — Stage ' + str(stage_obj.stage + 1) + '/' + str(total_stages) + ': ' + gate_pass.pass_number

        all_stages = list(gate_pass.approvals.all())
        progress_html = _build_workflow_progress(all_stages, stage_obj.pk)

        on_leave = '' if gate_pass.expected_return_time else ' <span style="background:#fef3c7;color:#92400e;padding:2px 8px;border-radius:4px;font-size:11px;">On Leave</span>'
        ret_time = str(gate_pass.expected_return_time) if gate_pass.expected_return_time else 'N/A' + on_leave

        details = _build_details_table([
            ('Pass Number',    '<strong style="color:#1e3a5f;font-size:14px;">' + gate_pass.pass_number + '</strong>'),
            ('Employee',       gate_pass.employee.employee_name + ' (' + gate_pass.employee.employee_code + ')'),
            ('Department',     gate_pass.employee.department),
            ('Designation',    gate_pass.employee.designation),
            ('Purpose',        gate_pass.get_purpose_display()),
            ('Destination',    gate_pass.destination),
            ('Date',           str(gate_pass.out_date)),
            ('Out Time',       str(gate_pass.out_time)),
            ('Expected Return', ret_time),
            ('Transport',      gate_pass.get_transport_mode_display()),
            ('Purpose Details', gate_pass.purpose_detail),
        ])

        body = (
            _build_email_header(
                'Action Required',
                'Internal Gate Pass &mdash; Stage ' + str(stage_obj.stage + 1) + ' of ' + str(total_stages) + ': <strong style="color:white;">' + stage_obj.stage_label + '</strong>',
                gate_pass.pass_number
            ) +
            '<tr><td style="padding:24px 40px;background:#f8fafc;border-bottom:1px solid #e2e8f0;">'
            '<div style="font-size:11px;font-weight:700;color:#94a3b8;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px;">Approval Progress</div>'
            '<table cellpadding="0" cellspacing="0" style="width:100%;"><tr style="vertical-align:middle;">' + progress_html + '</tr></table>'
            '</td></tr>'
            '<tr><td style="padding:28px 40px 0;">'
            '<p style="font-size:16px;color:#1e293b;margin:0 0 8px;">Dear <strong style="color:#1e3a5f;">' + approver.employee_name + '</strong>,</p>'
            '<p style="font-size:14px;color:#64748b;margin:0;line-height:1.6;">An Internal Gate Pass requires your approval at <strong>Stage ' + str(stage_obj.stage + 1) + '</strong>. Please review the details carefully and take action.</p>'
            '</td></tr>'
            '<tr><td style="padding:20px 40px;">' + details + '</td></tr>'
            '<tr><td style="padding:8px 40px 32px;">' + _build_action_buttons(approve_url, reject_url) + '</td></tr>' +
            _build_email_footer()
        )

        html_body = _wrap_email(body)
        plain = ("Dear " + approver.employee_name + ",\n\nIGP " + gate_pass.pass_number +
                 " requires your approval at Stage " + str(stage_obj.stage + 1) + ": " + stage_obj.stage_label +
                 ".\n\nEmployee: " + gate_pass.employee.employee_name +
                 "\nDepartment: " + gate_pass.employee.department +
                 "\nPurpose: " + gate_pass.get_purpose_display() +
                 "\nDestination: " + gate_pass.destination +
                 "\nDate: " + str(gate_pass.out_date) +
                 "\n\nAPPROVE: " + approve_url +
                 "\nREJECT : " + reject_url +
                 "\n\nRegards,\nERP Department - Unity Cement")

        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [approver.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def _send_status_mail(gate_pass, action):
    """Notify the employee about final approval/rejection with extraordinary HTML."""
    try:
        if not gate_pass.employee.email:
            return
        subject = 'IGP ' + action.upper() + ' — ' + gate_pass.pass_number

        if action == 'approved':
            hdr_bg = 'linear-gradient(135deg,#064e3b,#059669)'
            s_bg, s_fg = '#d1fae5', '#065f46'
            s_icon, s_label = '&#10003;', 'FULLY APPROVED'
            msg_text = 'Congratulations! Your Internal Gate Pass has been <strong>fully approved</strong> through all workflow stages. You may proceed.'
        else:
            hdr_bg = 'linear-gradient(135deg,#7f1d1d,#dc2626)'
            s_bg, s_fg = '#fee2e2', '#991b1b'
            s_icon, s_label = '&#10007;', 'REJECTED'
            msg_text = 'Your Internal Gate Pass has been <strong>rejected</strong>. Please contact your approver or raise a new request.'

        trail_rows = ''
        for i, s in enumerate(gate_pass.approvals.all()):
            bg = 'white' if i % 2 == 0 else '#f8fafc'
            if s.status == 'approved':
                badge_bg, badge_fg = '#d1fae5', '#065f46'
            elif s.status == 'rejected':
                badge_bg, badge_fg = '#fee2e2', '#991b1b'
            else:
                badge_bg, badge_fg = '#fef3c7', '#92400e'
            approver_name = s.approver.employee_name if s.approver else 'N/A'
            trail_rows += (
                '<tr style="background:' + bg + ';">'
                '<td style="padding:8px 16px;font-size:12px;font-weight:600;color:#475569;width:5%;">' + str(s.stage + 1) + '</td>'
                '<td style="padding:8px 16px;font-size:12px;color:#1e293b;">' + s.stage_label + '</td>'
                '<td style="padding:8px 16px;font-size:12px;color:#475569;">' + approver_name + '</td>'
                '<td style="padding:8px 16px;"><span style="font-size:11px;font-weight:700;padding:3px 10px;border-radius:20px;background:' + badge_bg + ';color:' + badge_fg + ';">' + s.status.upper() + '</span></td>'
                '<td style="padding:8px 16px;font-size:11px;color:#64748b;">' + (s.remarks or '—') + '</td>'
                '</tr>'
            )

        body = (
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:4px;"></td></tr>'
            '<tr><td style="background:' + hdr_bg + ';padding:36px 40px;text-align:center;">'
            '<div style="font-size:11px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;">Unity Cement &mdash; ERP Department</div>'
            '<div style="width:72px;height:72px;border-radius:50%;background:rgba(255,255,255,0.15);border:2px solid rgba(255,255,255,0.3);display:inline-flex;align-items:center;justify-content:center;margin-bottom:16px;font-size:32px;">' + s_icon + '</div>'
            '<div style="font-size:26px;font-weight:800;color:white;margin-bottom:8px;">Gate Pass ' + action.title() + '</div>'
            '<div style="display:inline-block;background:' + s_bg + ';color:' + s_fg + ';border-radius:50px;padding:6px 20px;font-size:13px;font-weight:700;letter-spacing:1px;">' + s_label + '</div>'
            '<div style="margin-top:14px;display:inline-block;background:rgba(255,255,255,0.12);border:1.5px solid rgba(255,255,255,0.25);border-radius:50px;padding:5px 18px;font-size:13px;color:rgba(255,255,255,0.9);font-weight:600;">' + gate_pass.pass_number + '</div>'
            '</td></tr>'
            '<tr><td style="padding:28px 40px 0;">'
            '<p style="font-size:16px;color:#1e293b;margin:0 0 8px;">Dear <strong style="color:#1e3a5f;">' + gate_pass.employee.employee_name + '</strong>,</p>'
            '<p style="font-size:14px;color:#64748b;margin:0;line-height:1.6;">' + msg_text + '</p>'
            '</td></tr>'
            '<tr><td style="padding:20px 40px;">' +
            _build_details_table([
                ('Pass Number', '<strong>' + gate_pass.pass_number + '</strong>'),
                ('Purpose',     gate_pass.get_purpose_display()),
                ('Destination', gate_pass.destination),
                ('Date',        str(gate_pass.out_date)),
                ('Out Time',    str(gate_pass.out_time)),
            ]) +
            '</td></tr>'
            '<tr><td style="padding:0 40px 28px;">'
            '<div style="border-radius:12px;overflow:hidden;border:1px solid #e2e8f0;">'
            '<div style="background:#f1f5f9;padding:12px 16px;border-bottom:1px solid #e2e8f0;">'
            '<span style="font-size:11px;font-weight:700;color:#475569;letter-spacing:1.5px;text-transform:uppercase;">&#9989; Approval Trail</span></div>'
            '<table width="100%%" cellpadding="0" cellspacing="0">'
            '<tr style="background:#f8fafc;"><th style="padding:8px 16px;font-size:11px;color:#94a3b8;text-align:left;font-weight:600;">#</th>'
            '<th style="padding:8px 16px;font-size:11px;color:#94a3b8;text-align:left;font-weight:600;">Stage</th>'
            '<th style="padding:8px 16px;font-size:11px;color:#94a3b8;text-align:left;font-weight:600;">Approver</th>'
            '<th style="padding:8px 16px;font-size:11px;color:#94a3b8;text-align:left;font-weight:600;">Status</th>'
            '<th style="padding:8px 16px;font-size:11px;color:#94a3b8;text-align:left;font-weight:600;">Remarks</th></tr>'
            + trail_rows +
            '</table></div></td></tr>' +
            _build_email_footer()
        )

        html_body = _wrap_email(body)
        plain = ("Dear " + gate_pass.employee.employee_name + ",\n\nYour IGP (" + gate_pass.pass_number + ") has been " + action + ".\n\nPurpose: " + gate_pass.get_purpose_display() +
                 "\nDestination: " + gate_pass.destination +
                 "\nDate: " + str(gate_pass.out_date) +
                 "\n\nRegards,\nERP Department - Unity Cement")

        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [gate_pass.employee.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def _advance_workflow(gate_pass, request=None):
    """Move to next pending stage or mark fully approved."""
    stages = gate_pass.approvals.all()
    for stage in stages:
        if stage.status == 'pending':
            gate_pass.current_stage = stage.stage
            gate_pass.status = 'in_progress'
            gate_pass.save()
            _send_stage_mail(gate_pass, stage, request=request)
            return
    # All stages approved
    gate_pass.status = 'approved'
    gate_pass.approved_at = timezone.now()
    gate_pass.save()
    _send_status_mail(gate_pass, 'approved')


# ─── Views ──────────────────────────────────────────────────────────────────

def _notify_approvers_and_admins(notification_type, title, description, related_id, related_module, recipient_employee=None):
    """Create workflow notifications based on system settings."""
    send_workflow_notification(
        module_key='igp',
        notification_type=notification_type,
        title=title,
        description=description,
        related_id=related_id,
        related_module=related_module,
        requester=recipient_employee,
    )

@login_required
def pass_list(request):
    if not request.user.can_access_internal_pass and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')

    user = request.user
    if user.is_superuser or user.has_role('administrator'):
        passes = InternalGatePass.objects.all()
    elif user.has_role('security'):
        # Security sees passes where their stage is current
        passes = InternalGatePass.objects.filter(
            status__in=['in_progress', 'approved', 'returned', 'rejected']
        ) | InternalGatePass.objects.filter(
            approvals__approver_role='security',
            approvals__status='pending',
            status='in_progress'
        )
        passes = passes.distinct()
    elif user.has_role('hr'):
        passes = InternalGatePass.objects.filter(
            Q(approvals__approver_role='hr', approvals__status='pending', status='in_progress') |
            Q(approvals__approver_role='hr', approvals__status__in=['approved', 'rejected'])
        ).distinct()
    elif user.has_role('management'):
        passes = InternalGatePass.objects.filter(
            Q(approvals__approver_role='management', approvals__status='pending', status='in_progress') |
            Q(approvals__approver_role='management', approvals__status__in=['approved', 'rejected'])
        ).distinct()
    elif user.has_role('department_hod'):
        dept_q = Q(employee__department__in=user.get_all_departments())
        extra_dept_q = Q()
        for department in user.get_all_departments():
            extra_dept_q |= Q(employee__additional_departments__contains=f'|{department}|')
        passes = InternalGatePass.objects.filter(
            dept_q | extra_dept_q |
            Q(approvals__approver_role='department_hod', approvals__status='pending', status='in_progress')
        ).distinct()
    elif user.has_role('president_plant_head'):
        passes = InternalGatePass.objects.filter(
            Q(approvals__approver_role='president_plant_head', approvals__status='pending', status='in_progress') |
            Q(approvals__approver_role='president_plant_head', approvals__status__in=['approved', 'rejected'])
        ).distinct()
    else:
        passes = InternalGatePass.objects.filter(employee=user)

    # Department filter (admin/management/president see all)
    FULL_ROLES = ('administrator', 'management', 'president_plant_head')
    dept_filter = request.GET.get('dept', '')
    if dept_filter and (user.is_superuser or any(user.has_role(role) for role in FULL_ROLES)):
        passes = passes.filter(employee__department=dept_filter)
    elif not user.is_superuser and not any(user.has_role(role) for role in FULL_ROLES) and not user.has_role('hr') and not user.has_role('security') and not user.has_role('department_hod'):
        # Regular employees only see their own department
        dept_q = Q(employee__department__in=user.get_all_departments())
        extra_dept_q = Q()
        for department in user.get_all_departments():
            extra_dept_q |= Q(employee__additional_departments__contains=f'|{department}|')
        passes = passes.filter(dept_q | extra_dept_q)

    status_filter = request.GET.get('status', '')
    search = request.GET.get('q', '').strip()
    if status_filter:
        passes = passes.filter(status=status_filter)
    if search:
        passes = passes.filter(
            Q(pass_number__icontains=search) |
            Q(employee__employee_name__icontains=search) |
            Q(employee__employee_code__icontains=search) |
            Q(employee__department__icontains=search) |
            Q(destination__icontains=search)
        )

    actionable_pks = set()
    for gp in passes:
        if _can_user_act(user, gp):
            actionable_pks.add(gp.pk)

    from accounts.models import DEPARTMENT_CHOICES
    from django.core.paginator import Paginator
    paginator = Paginator(passes, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'internal_pass/pass_list.html', {
        'passes': page_obj, 'page_obj': page_obj,
        'status_filter': status_filter,
        'search': search,
        'actionable_pks': actionable_pks,
        'dept_filter': dept_filter,
        'department_choices': DEPARTMENT_CHOICES,
        'show_dept_filter': user.is_superuser or any(user.has_role(role) for role in FULL_ROLES),
    })


def _can_user_act(user, gate_pass):
    """Check if this user is the expected approver for the current pending stage."""
    if user.is_superuser or user.has_role('administrator'):
        return gate_pass.status in ('pending', 'in_progress')
    pending_stage = gate_pass.approvals.filter(status='pending').first()
    if not pending_stage:
        return False
    if not user.has_role(pending_stage.approver_role):
        return False
    if pending_stage.approver_role == 'department_hod':
        return user.has_department(gate_pass.employee.department)
    return True


@login_required
def pass_create(request):
    if not request.user.can_access_internal_pass and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')
    form = InternalGatePassForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        gp = form.save(commit=False)
        gp.employee = request.user
        gp.status = 'pending'
        gp.save()
        _create_approval_stages(gp)
        if gp.approvals.exists():
            _advance_workflow(gp, request=request)
        else:
            gp.status = 'approved'
            gp.approved_at = timezone.now()
            gp.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'igp_create', 'IGP', f'IGP created: {gp.pass_number} by {request.user.username}')
        messages.success(request, f'Gate Pass {gp.pass_number} submitted successfully.')
        return redirect('internal_pass:pass_list')
    return render(request, 'internal_pass/pass_form.html', {'form': form, 'title': 'New Internal Gate Pass'})


@login_required
def pass_detail(request, pk):
    gp = get_object_or_404(InternalGatePass, pk=pk)
    stages = gp.approvals.all()
    can_act = _can_user_act(request.user, gp)
    return render(request, 'internal_pass/pass_detail.html', {
        'gp': gp, 'stages': stages, 'can_act': can_act
    })


@login_required
def pass_approve(request, pk):
    gp = get_object_or_404(InternalGatePass, pk=pk)
    if not _can_user_act(request.user, gp):
        messages.error(request, 'You are not authorized to act on this gate pass at this stage.')
        return redirect('internal_pass:pass_list')

    pending_stage = gp.approvals.filter(status='pending').first()
    form = ApprovalForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        action = form.cleaned_data['action']
        remarks = form.cleaned_data['remarks']

        pending_stage.status = action
        pending_stage.approver = request.user
        pending_stage.remarks = remarks
        pending_stage.acted_at = timezone.now()
        pending_stage.save()

        if action == 'rejected':
            gp.status = 'rejected'
            gp.approval_remarks = remarks
            gp.approver = request.user
            gp.approved_at = timezone.now()
            gp.save()
            _send_status_mail(gp, 'rejected')
            from accounts.models import AuditLog
            AuditLog.log(request, 'igp_reject', 'IGP', f'IGP rejected: {gp.pass_number} at stage {pending_stage.stage+1} by {request.user.username}')
            messages.error(request, f'Gate Pass {gp.pass_number} rejected.')
            
            # Create notification for employee and admins
            _notify_approvers_and_admins(
                notification_type='igp_rejected',
                title=f'IGP {gp.pass_number} Rejected',
                description=f'Your Internal Gate Pass has been rejected at stage {pending_stage.stage + 1}',
                related_id=str(gp.id),
                related_module='IGP',
                recipient_employee=gp.employee
            )
        else:
            from accounts.models import AuditLog
            AuditLog.log(request, 'igp_approve', 'IGP', f'IGP stage {pending_stage.stage+1} approved: {gp.pass_number} by {request.user.username}')
            messages.success(request, f'Stage {pending_stage.stage + 1} approved. Moving to next stage.')
            _advance_workflow(gp, request=request)
            
            # Check if this was the final stage
            if gp.status == 'approved':
                _notify_approvers_and_admins(
                    notification_type='igp_approved',
                    title=f'IGP {gp.pass_number} Approved',
                    description='Your Internal Gate Pass has been fully approved',
                    related_id=str(gp.id),
                    related_module='IGP',
                    recipient_employee=gp.employee
                )

        return redirect('internal_pass:pass_list')

    return render(request, 'internal_pass/pass_approve.html', {
        'gp': gp, 'form': form, 'stage': pending_stage,
        'stages': gp.approvals.all()
    })


def stage_action(request, token, action):
    """Token-based approval from email link — shows both Approve & Reject buttons."""
    stage_obj = get_object_or_404(GatePassApproval, token=token)
    gp = stage_obj.gate_pass

    if stage_obj.status != 'pending':
        return render(request, 'internal_pass/email_action_done.html', {
            'message': f'This stage has already been {stage_obj.status}.', 'gp': gp,
            'stages': gp.approvals.all()
        })
    if gp.status == 'rejected':
        return render(request, 'internal_pass/email_action_done.html', {
            'message': 'This gate pass has already been rejected.', 'gp': gp,
            'stages': gp.approvals.all()
        })

    if request.method == 'POST':
        # action_choice comes from the button clicked (approve/reject)
        chosen = request.POST.get('action_choice', action)
        if chosen not in ('approve', 'reject'):
            chosen = action
        remarks = request.POST.get('remarks', '')
        stage_obj.status = 'approved' if chosen == 'approve' else 'rejected'
        stage_obj.remarks = remarks
        stage_obj.acted_at = timezone.now()
        stage_obj.save()

        if chosen == 'reject':
            gp.status = 'rejected'
            gp.approval_remarks = remarks
            gp.approved_at = timezone.now()
            gp.save()
            _send_status_mail(gp, 'rejected')
            
            # Create notification for employee and admins
            _notify_approvers_and_admins(
                notification_type='igp_rejected',
                title=f'IGP {gp.pass_number} Rejected',
                description=f'Your Internal Gate Pass has been rejected at stage {stage_obj.stage + 1}',
                related_id=str(gp.id),
                related_module='IGP',
                recipient_employee=gp.employee
            )
        else:
            _advance_workflow(gp, request=request)
            
            # Check if this was the final stage (all approvals done)
            if gp.status == 'approved':
                _notify_approvers_and_admins(
                    notification_type='igp_approved',
                    title=f'IGP {gp.pass_number} Approved',
                    description='Your Internal Gate Pass has been fully approved',
                    related_id=str(gp.id),
                    related_module='IGP',
                    recipient_employee=gp.employee
                )

        return render(request, 'internal_pass/email_action_done.html', {
            'message': f'Stage {stage_obj.stage + 1} ({stage_obj.stage_label}) has been {stage_obj.status} successfully.',
            'gp': gp, 'stages': gp.approvals.all()
        })

    # GET — show the action page with both buttons
    return render(request, 'internal_pass/email_action.html', {
        'gp': gp, 'action': action, 'stage': stage_obj,
        'stages': gp.approvals.all()
    })


@login_required
def mark_returned(request, pk):
    gp = get_object_or_404(InternalGatePass, pk=pk)
    form = ReturnForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        gp.actual_return_time = form.cleaned_data['actual_return_time']
        gp.status = 'returned'
        gp.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'igp_return', 'IGP', f'IGP marked returned: {gp.pass_number} by {request.user.username}')
        messages.success(request, 'Marked as returned.')
        return redirect('internal_pass:pass_list')
    return render(request, 'internal_pass/mark_returned.html', {'gp': gp, 'form': form})


@login_required
def print_pass(request, pk):
    gp = get_object_or_404(InternalGatePass, pk=pk)
    buf = BytesIO()

    from reportlab.lib.pagesizes import A4
    # Half A4 portrait = 210mm wide x 148.5mm tall
    PAGE_W = 210 * mm
    PAGE_H = 148.5 * mm

    doc = SimpleDocTemplate(buf, pagesize=(PAGE_W, PAGE_H),
                            topMargin=8*mm, bottomMargin=6*mm,
                            leftMargin=10*mm, rightMargin=10*mm)

    navy   = colors.HexColor('#1e3a5f')
    light  = colors.HexColor('#e8f0fe')
    alt    = colors.HexColor('#f8f9fa')
    border = colors.HexColor('#dee2e6')

    def sty(name, **kw):
        kw.setdefault('fontName', 'Helvetica')
        kw.setdefault('fontSize', 8)
        return ParagraphStyle(name, **kw)

    elements = []

    # ── HEADER: logo | title | pass no ──
    logo_path = os.path.join(django_settings.BASE_DIR, 'static', 'images', 'Unity_Logo_Horizontal.png')
    logo_cell = Image(logo_path, width=36*mm, height=10*mm) if os.path.exists(logo_path) \
                else Paragraph('UNITY CEMENT', sty('lc', fontName='Helvetica-Bold', fontSize=10, textColor=navy))

    title_para = Paragraph(
        '<b><font size=12 color="#1e3a5f">INTERNAL GATE PASS</font></b><br/>'
        '<font size=7 color="#718096">Unity Cement — PMC Cement Limited</font>',
        sty('tp', alignment=TA_CENTER, leading=14)
    )
    passno_para = Paragraph(
        f'<font size=7 color="#718096">Pass No.</font><br/>'
        f'<b><font size=10 color="#1e3a5f">{gp.pass_number}</font></b>',
        sty('pp', alignment=2, leading=13)
    )
    hdr = Table([[logo_cell, title_para, passno_para]], colWidths=[38*mm, 100*mm, 52*mm])
    hdr.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(hdr)
    elements.append(HRFlowable(width='100%', thickness=1.5, color=navy, spaceAfter=3*mm))

    # ── PASS DETAILS TABLE (2 columns) ──
    lbl = sty('lbl', fontName='Helvetica-Bold', fontSize=7.5, textColor=navy)
    val = sty('val', fontSize=8)

    def row(l1, v1, l2='', v2=''):
        return [
            Paragraph(l1, lbl), Paragraph(str(v1) if v1 else '-', val),
            Paragraph(l2, lbl), Paragraph(str(v2) if v2 else '-', val),
        ]

    # Status badge color
    status_colors = {
        'approved': ('#d1e7dd', '#0f5132'),
        'rejected': ('#f8d7da', '#842029'),
        'pending':  ('#fff3cd', '#856404'),
        'in_progress': ('#cfe2ff', '#084298'),
        'returned': ('#cff4fc', '#055160'),
    }
    sc_bg, sc_fg = status_colors.get(gp.status, ('#e2e3e5', '#383d41'))
    status_para = Paragraph(
        f'<b><font color="{sc_fg}">{gp.get_status_display()}</font></b>',
        sty('sp', fontSize=8)
    )

    details = [
        row('Pass Number', gp.pass_number,       'Date',          str(gp.out_date)),
        row('Employee',    gp.employee.employee_name, 'Emp Code',  gp.employee.employee_code),
        row('Department',  gp.employee.department,   'Designation', gp.employee.designation),
        row('Purpose',     gp.get_purpose_display(),  'Transport',  gp.get_transport_mode_display()),
        row('Destination', gp.destination,        'Vehicle No',    gp.vehicle_number or '-'),
        row('Out Time',    str(gp.out_time),       'Exp. Return',   str(gp.expected_return_time)),
        row('Actual Return', str(gp.actual_return_time or '-'), 'Status', gp.get_status_display()),
    ]

    dt = Table(details, colWidths=[24*mm, 58*mm, 24*mm, 58*mm])
    dt.setStyle(TableStyle([
        ('FONTSIZE',    (0,0), (-1,-1), 7.5),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, alt]),
        ('GRID',        (0,0), (-1,-1), 0.4, border),
        ('PADDING',     (0,0), (-1,-1), 3),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('LINEABOVE',   (0,0), (-1,0),  0.8, navy),
        ('LINEBELOW',   (0,-1),(-1,-1), 0.8, navy),
    ]))
    elements.append(dt)
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(f'<b>Purpose Details:</b> {gp.purpose_detail}',
                               sty('pd', fontSize=7.5, textColor=colors.HexColor('#2d3748'))))

    # ── APPROVAL TRAIL ──
    elements.append(Spacer(1, 2*mm))
    elements.append(HRFlowable(width='100%', thickness=0.8, color=border, spaceAfter=1*mm))
    elements.append(Paragraph('Approval Trail',
                               sty('at', fontName='Helvetica-Bold', fontSize=8, textColor=navy)))
    elements.append(Spacer(1, 1*mm))

    stage_data = [['Stage', 'Description', 'Approver', 'Status', 'Remarks']]
    for s in gp.approvals.all():
        stage_data.append([
            str(s.stage + 1), s.stage_label,
            s.approver.employee_name if s.approver else '-',
            s.status.upper(), s.remarks or '-'
        ])
    if len(stage_data) > 1:
        st = Table(stage_data, colWidths=[10*mm, 52*mm, 36*mm, 18*mm, 44*mm])
        st.setStyle(TableStyle([
            ('FONTNAME',    (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,-1), 7),
            ('BACKGROUND',  (0,0), (-1,0),  navy),
            ('TEXTCOLOR',   (0,0), (-1,0),  colors.white),
            ('GRID',        (0,0), (-1,-1), 0.3, border),
            ('PADDING',     (0,0), (-1,-1), 3),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt]),
            ('ALIGN',       (0,0), (0,-1),  'CENTER'),
            ('ALIGN',       (3,0), (3,-1),  'CENTER'),
        ]))
        elements.append(st)

    # ── SIGNATURES ──
    elements.append(Spacer(1, 3*mm))
    sig_s = sty('sig', alignment=TA_CENTER, fontSize=7, textColor=colors.HexColor('#555'))
    sig_n = sty('sn',  alignment=TA_CENTER, fontSize=7.5, fontName='Helvetica-Bold', textColor=navy)
    sig_data = [
        [Paragraph('_________________', sig_s), Paragraph('_________________', sig_s), Paragraph('_________________', sig_s)],
        [Paragraph(gp.employee.employee_name, sig_n),
         Paragraph(gp.approver.employee_name if gp.approver else 'Approver', sig_n),
         Paragraph('Security Officer', sig_n)],
        [Paragraph('Employee Signature', sig_s), Paragraph('Approver Signature', sig_s), Paragraph('Security Signature', sig_s)],
    ]
    sig_t = Table(sig_data, colWidths=[62*mm, 62*mm, 62*mm])
    sig_t.setStyle(TableStyle([
        ('ALIGN',   (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',  (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 1),
        ('BOTTOMPADDING', (0,0), (-1,-1), 1),
    ]))
    elements.append(sig_t)
    elements.append(HRFlowable(width='100%', thickness=0.5, color=border, spaceBefore=2*mm))
    elements.append(Paragraph('This is a computer generated document. &nbsp;|&nbsp; Unity Cement — PMC Cement Limited',
                               sty('ft', alignment=TA_CENTER, fontSize=6.5, textColor=colors.grey)))

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="IGP_{gp.pass_number}.pdf"'
    return response


@login_required
def daily_report(request):
    report_date = request.GET.get('date', str(date.today()))
    dept = request.GET.get('dept', '')
    view_mode = request.GET.get('view', 'detail')  # summary | detail
    passes = InternalGatePass.objects.filter(out_date=report_date)
    if dept: passes = passes.filter(employee__department=dept)
    from accounts.models import DEPARTMENT_CHOICES
    return render(request, 'internal_pass/report.html', {
        'passes': passes, 'report_date': report_date, 'report_type': 'Daily',
        'dept_filter': dept, 'department_choices': DEPARTMENT_CHOICES,
        'view_mode': view_mode,
        'summary': _igp_summary(passes),
    })


@login_required
def monthly_report(request):
    month = request.GET.get('month', date.today().strftime('%Y-%m'))
    year, mon = month.split('-')
    dept = request.GET.get('dept', '')
    view_mode = request.GET.get('view', 'detail')  # summary | detail
    passes = InternalGatePass.objects.filter(out_date__year=year, out_date__month=mon)
    if dept: passes = passes.filter(employee__department=dept)
    from accounts.models import DEPARTMENT_CHOICES
    return render(request, 'internal_pass/report.html', {
        'passes': passes, 'report_date': month, 'report_type': 'Monthly',
        'dept_filter': dept, 'department_choices': DEPARTMENT_CHOICES,
        'view_mode': view_mode,
        'summary': _igp_summary(passes),
    })


def _igp_summary(qs):
    """Build summary aggregates for internal gate pass reports."""
    status_map = dict(InternalGatePass._meta.get_field('status').choices)
    purpose_map = dict(InternalGatePass._meta.get_field('purpose').choices)

    total = qs.count()

    by_status = []
    for row in qs.values('status').annotate(count=Count('id')).order_by('-count'):
        key = row['status'] or ''
        by_status.append({'key': key, 'label': status_map.get(key, key or '—'), 'count': row['count']})

    by_purpose = []
    for row in qs.values('purpose').annotate(count=Count('id')).order_by('-count'):
        key = row['purpose'] or ''
        by_purpose.append({'key': key, 'label': purpose_map.get(key, key or '—'), 'count': row['count']})

    by_department = []
    for row in qs.values('employee__department').annotate(count=Count('id')).order_by('-count'):
        key = row['employee__department'] or ''
        by_department.append({'label': key or '—', 'count': row['count']})

    return {'total': total, 'by_status': by_status, 'by_purpose': by_purpose, 'by_department': by_department}


@login_required
def export_report(request):
    report_type = request.GET.get('type', 'daily')
    report_date = request.GET.get('date', str(date.today()))
    if report_type == 'monthly':
        year, mon = report_date.split('-')
        passes = InternalGatePass.objects.filter(out_date__year=year, out_date__month=mon)
    else:
        passes = InternalGatePass.objects.filter(out_date=report_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Internal Gate Pass Report'
    headers = ['Pass No', 'Employee', 'Emp Code', 'Department', 'Purpose', 'Destination',
               'Date', 'Out Time', 'Exp. Return', 'Actual Return', 'Transport',
               'Approved By', 'Approval Date & Time', 'Status']
    hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20
    for gp in passes:
        approved_at_str = gp.approved_at.strftime('%d-%m-%Y %H:%M') if gp.approved_at else ''
        ws.append([
            gp.pass_number, gp.employee.employee_name, gp.employee.employee_code,
            gp.employee.department, gp.get_purpose_display(), gp.destination,
            str(gp.out_date), str(gp.out_time), str(gp.expected_return_time or ''),
            str(gp.actual_return_time or ''), gp.get_transport_mode_display(),
            gp.approver.employee_name if gp.approver else '',
            approved_at_str, gp.get_status_display(),
        ])
    from accounts.report_utils import add_excel_logo_and_note
    add_excel_logo_and_note(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="IGP_Report_{report_date}.xlsx"'
    return response


@login_required
def export_report_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    report_type = request.GET.get('type', 'daily')
    report_date = request.GET.get('date', str(date.today()))
    if report_type == 'monthly':
        year, mon = report_date.split('-')
        passes = InternalGatePass.objects.filter(out_date__year=year, out_date__month=mon)
    else:
        passes = InternalGatePass.objects.filter(out_date=report_date)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=12*mm, bottomMargin=10*mm,
                            leftMargin=10*mm, rightMargin=10*mm)

    navy  = colors.HexColor('#1e3a5f')
    light = colors.HexColor('#e8f0fe')
    alt   = colors.HexColor('#f8f9fa')

    title_style = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=14,
                                 alignment=TA_CENTER, textColor=navy, spaceAfter=2)
    sub_style   = ParagraphStyle('s', fontName='Helvetica', fontSize=9,
                                 alignment=TA_CENTER, textColor=colors.grey, spaceAfter=6)

    elements = []
    from accounts.report_utils import build_pdf_header_table
    elements.append(build_pdf_header_table(
        'INTERNAL GATE PASS REPORT',
        f'{report_type} Report \u2014 {report_date}  |  Total Records: {passes.count()}  |  Generated: {timezone.now().strftime("%d %b %Y %H:%M")}',
        277  # A4 landscape width in mm
    ))
    elements.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=4*mm))

    headers = ['#', 'Pass No', 'Employee', 'Dept', 'Purpose', 'Destination',
               'Date', 'Out', 'Exp.Return', 'Act.Return', 'Transport', 'Status']
    col_w = [8*mm, 22*mm, 32*mm, 28*mm, 24*mm, 30*mm,
             20*mm, 16*mm, 20*mm, 20*mm, 22*mm, 18*mm]

    data = [headers]
    for i, gp in enumerate(passes, 1):
        data.append([
            str(i), gp.pass_number, gp.employee.employee_name,
            gp.employee.department, gp.get_purpose_display(), gp.destination,
            str(gp.out_date), str(gp.out_time), str(gp.expected_return_time),
            str(gp.actual_return_time or '-'), gp.get_transport_mode_display(),
            gp.get_status_display(),
        ])

    if len(data) == 1:
        data.append(['', 'No records found for this period.', '', '', '', '', '', '', '', '', '', ''])

    table = Table(data, colWidths=col_w, repeatRows=1)
    row_bg = [colors.white, alt]
    table.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  8),
        ('BACKGROUND',  (0, 0), (-1, 0),  navy),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 1), (-1, -1), 7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), row_bg),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('LINEABOVE',   (0, 0), (-1, 0),  1.5, navy),
        ('LINEBELOW',   (0, 0), (-1, 0),  1.5, navy),
        ('ALIGN',       (0, 0), (0, -1),  'CENTER'),
        ('ALIGN',       (6, 0), (11, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',     (0, 0), (-1, -1), 3),
        ('WORDWRAP',    (0, 0), (-1, -1), True),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph('This is a computer generated report. \u2014 Unity Cement ERP System', sub_style))

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="IGP_Report_{report_date}.pdf"'
    return response


@login_required
def bulk_action(request):
    if request.method != 'POST':
        return redirect('internal_pass:pass_list')
    action = request.POST.get('action')
    ids    = request.POST.getlist('selected_ids')
    if not ids:
        messages.warning(request, 'No passes selected.')
        return redirect('internal_pass:pass_list')
    passes = InternalGatePass.objects.filter(pk__in=ids)
    if action == 'delete':
        if not request.user.is_superuser:
            messages.error(request, 'Access Denied.')
            return redirect('internal_pass:pass_list')
        count = passes.count()
        passes.delete()
        from accounts.models import AuditLog
        AuditLog.log(request, 'igp_bulk_delete', 'IGP', f'{count} IGPs deleted in bulk by {request.user.username}')
        messages.success(request, f'{count} pass(es) deleted.')
    elif action == 'cancel':
        updated = passes.exclude(status__in=['returned', 'rejected']).update(status='rejected')
        from accounts.models import AuditLog
        AuditLog.log(request, 'igp_bulk_cancel', 'IGP', f'{updated} IGPs cancelled in bulk by {request.user.username}')
        messages.success(request, f'{updated} pass(es) cancelled.')
    elif action == 'duplicate':
        import uuid as _uuid
        for gp in passes:
            gp.pk              = None
            gp.pass_number     = ''
            gp.status          = 'pending'
            gp.approver        = None
            gp.approved_at     = None
            gp.approval_remarks = ''
            gp.approval_token  = _uuid.uuid4()
            gp.actual_return_time = None
            gp.save()
            for approval in GatePassApproval.objects.filter(gate_pass__id=gp.pk):
                approval.pk = None
                approval.gate_pass = gp
                approval.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'igp_bulk_duplicate', 'IGP', f'{len(ids)} IGPs duplicated in bulk by {request.user.username}')
        messages.success(request, f'{len(ids)} pass(es) duplicated.')
    else:
        messages.error(request, 'Unknown action.')
    return redirect('internal_pass:pass_list')

