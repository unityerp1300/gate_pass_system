from django.db import models
from django.db.models import Count, Sum
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import os, base64
from django.conf import settings as django_settings
from django.core.files.base import ContentFile
from reportlab.lib.pagesizes import A5, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, BaseDocTemplate, Frame, PageTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.utils import ImageReader
from .models import VisitorGatePass
from .forms import VisitorGatePassForm, VisitorApprovalForm, CheckoutForm, PhotoCaptureForm
from accounts.notification_service import send_workflow_notification
from datetime import date


def send_visitor_approval_request(vgp, request=None):
    """Send extraordinary HTML approval email to contact person (Stage 1 of 2)."""
    try:
        host = vgp.person_to_meet
        if not host or not host.email:
            return
        base = getattr(settings, 'SITE_BASE_URL', None)
        if not base:
            base = f"{request.scheme}://{request.get_host()}" if request else 'http://localhost:8000'
        base = base.rstrip('/')
        approve_url = f"{base}/visitor-pass/token/{vgp.approval_token}/approve/"
        reject_url  = f"{base}/visitor-pass/token/{vgp.approval_token}/reject/"
        subject = f'[Action Required] Visitor Approval — {vgp.pass_number}'

        # Workflow progress: Stage 1 = Contact Person (current), Stage 2 = Security
        progress = (
            '<td style="text-align:center;padding:0 6px;">'
            '<div style="display:inline-flex;flex-direction:column;align-items:center;gap:5px;">'
            '<div style="width:38px;height:38px;border-radius:50%;background:#fef3c7;color:#92400e;'
            'display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;'
            'border:2px solid #2563eb;box-shadow:0 0 0 3px rgba(37,99,235,0.15);">&#9654;</div>'
            '<div style="font-size:10px;font-weight:600;color:#92400e;">Contact</div>'
            '<div style="font-size:9px;color:#94a3b8;">AWAITING</div>'
            '</div></td>'
            '<td style="width:20px;"><div style="height:2px;background:#e2e8f0;margin:0 2px;margin-bottom:28px;"></div></td>'
            '<td style="text-align:center;padding:0 6px;">'
            '<div style="display:inline-flex;flex-direction:column;align-items:center;gap:5px;">'
            '<div style="width:38px;height:38px;border-radius:50%;background:#f1f5f9;color:#94a3b8;'
            'display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;">&#9679;</div>'
            '<div style="font-size:10px;font-weight:600;color:#94a3b8;">Security</div>'
            '<div style="font-size:9px;color:#94a3b8;">PENDING</div>'
            '</div></td>'
        )

        details_rows = [
            ('Pass Number',    f'<strong style="color:#1e3a5f;font-size:14px;">{vgp.pass_number}</strong>'),
            ('Visitor Name',   vgp.visitor_name),
            ('Organization',   vgp.visitor_company or 'N/A'),
            ('Phone',          vgp.visitor_phone),
            ('Purpose',        vgp.get_visit_purpose_display()),
            ('Visit Details',  vgp.visit_detail),
            ('Date',           str(vgp.visit_date)),
            ('In Time',        str(vgp.in_time)),
            ('No. of Visitors', str(vgp.no_of_visitors)),
            ('Vehicle No.',    vgp.vehicle_number or 'N/A'),
        ]
        details_html = '<div style="border-radius:12px;overflow:hidden;border:1px solid #e2e8f0;">'
        details_html += '<div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);padding:12px 16px;"><span style="font-size:11px;font-weight:700;color:white;letter-spacing:1.5px;text-transform:uppercase;">&#128203; Visitor Details</span></div>'
        details_html += '<table width="100%" cellpadding="0" cellspacing="0">'
        for i, (lbl, val) in enumerate(details_rows):
            bg = 'white' if i % 2 == 0 else '#f8fafc'
            details_html += f'<tr style="background:{bg};"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;width:38%;border-bottom:1px solid #f1f5f9;">{lbl}</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{val}</td></tr>'
        details_html += '</table></div>'

        action_btns = (
            '<div style="background:#f8fafc;border-radius:12px;padding:24px;border:1px solid #e2e8f0;">'
            '<p style="font-size:13px;color:#475569;font-weight:600;margin:0 0 16px;text-align:center;">&#9889; Take Action Now</p>'
            '<table cellpadding="0" cellspacing="0" width="100%"><tr>'
            f'<td width="47%" align="center"><a href="{approve_url}" style="display:block;background:linear-gradient(135deg,#059669,#10b981);color:white;text-decoration:none;padding:16px 24px;border-radius:12px;font-size:16px;font-weight:800;text-align:center;box-shadow:0 4px 14px rgba(16,185,129,0.4);">&#10003;&nbsp;&nbsp;APPROVE</a></td>'
            '<td width="6%"></td>'
            f'<td width="47%" align="center"><a href="{reject_url}" style="display:block;background:linear-gradient(135deg,#dc2626,#ef4444);color:white;text-decoration:none;padding:16px 24px;border-radius:12px;font-size:16px;font-weight:800;text-align:center;box-shadow:0 4px 14px rgba(239,68,68,0.4);">&#10007;&nbsp;&nbsp;REJECT</a></td>'
            '</tr></table>'
            '<p style="font-size:11px;color:#94a3b8;text-align:center;margin:14px 0 0;">Clicking a button opens a secure confirmation page where you can add remarks.</p>'
            '</div>'
        )

        html_body = (
            '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"></head>'
            '<body style="margin:0;padding:0;background:#f0f4f8;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Inter,Arial,sans-serif;">'
            '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:32px 0;"><tr><td align="center">'
            '<table width="620" cellpadding="0" cellspacing="0" style="background:white;border-radius:20px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,0.12);">'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:4px;"></td></tr>'
            '<tr><td style="background:linear-gradient(135deg,#0f172a 0%,#1e3a5f 50%,#1e40af 100%);padding:36px 40px;text-align:center;">'
            '<div style="font-size:11px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;">Unity Cement &mdash; ERP Department</div>'
            '<div style="width:64px;height:64px;border-radius:50%;background:rgba(255,255,255,0.12);border:2px solid rgba(255,255,255,0.25);display:inline-flex;align-items:center;justify-content:center;margin-bottom:16px;font-size:28px;">&#128101;</div>'
            '<div style="font-size:26px;font-weight:800;color:white;margin-bottom:6px;">Visitor Approval Required</div>'
            '<div style="font-size:14px;color:rgba(255,255,255,0.7);margin-bottom:16px;">Stage 1 of 2: <strong style="color:white;">Contact Person Approval</strong></div>'
            f'<div style="display:inline-block;background:rgba(240,165,0,0.2);border:1.5px solid rgba(240,165,0,0.5);border-radius:50px;padding:6px 20px;font-size:14px;color:#fbbf24;font-weight:700;">{vgp.pass_number}</div>'
            '</td></tr>'
            '<tr><td style="padding:24px 40px;background:#f8fafc;border-bottom:1px solid #e2e8f0;">'
            '<div style="font-size:11px;font-weight:700;color:#94a3b8;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px;">Approval Progress</div>'
            '<table cellpadding="0" cellspacing="0" style="width:100%;"><tr style="vertical-align:middle;">' + progress + '</tr></table>'
            '</td></tr>'
            f'<tr><td style="padding:28px 40px 0;"><p style="font-size:16px;color:#1e293b;margin:0 0 8px;">Dear <strong style="color:#1e3a5f;">{host.employee_name}</strong>,</p>'
            '<p style="font-size:14px;color:#64748b;margin:0;line-height:1.6;">A visitor has arrived and is requesting to meet you. Please review the details and take action.</p></td></tr>'
            '<tr><td style="padding:20px 40px;">' + details_html + '</td></tr>'
            '<tr><td style="padding:8px 40px 32px;">' + action_btns + '</td></tr>'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:3px;"></td></tr>'
            '<tr><td style="background:#f8fafc;padding:20px 40px;text-align:center;">'
            '<p style="font-size:11px;color:#94a3b8;margin:0;">This is an automated email from <strong>ERP Department &mdash; Unity Cement (PMC Cement Ltd.)</strong></p>'
            '<p style="font-size:11px;color:#94a3b8;margin:6px 0 0;">Do not reply to this email &bull; For support contact IT/ERP Department</p>'
            '</td></tr>'
            '</table></td></tr></table></body></html>'
        )

        plain_body = f"""Dear {host.employee_name},

Visitor {vgp.visitor_name} is requesting to meet you.

Pass: {vgp.pass_number}
Organization: {vgp.visitor_company or 'N/A'}
Purpose: {vgp.get_visit_purpose_display()}
Date: {vgp.visit_date}

APPROVE: {approve_url}
REJECT : {reject_url}

Regards,
ERP Department — Unity Cement"""

        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain_body, settings.DEFAULT_FROM_EMAIL, [host.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def _send_vgp_security_mail(vgp, request=None):
    """Stage 2: Send approval request to Security after contact person approves."""
    try:
        from accounts.models import Employee as Emp
        security = Emp.objects.filter(is_active=True).filter(
            models.Q(role='security') | models.Q(additional_roles__contains='|security|')
        ).first()
        if not security or not security.email:
            return
        base = getattr(settings, 'SITE_BASE_URL', None)
        if not base:
            base = 'http://localhost:8000'
        base = base.rstrip('/')
        # Security uses the same token action but we mark it as security stage
        approve_url = f"{base}/visitor-pass/token/{vgp.approval_token}/security-approve/"
        reject_url  = f"{base}/visitor-pass/token/{vgp.approval_token}/security-reject/"
        subject = f'[Action Required] VGP Security Clearance — {vgp.pass_number}'

        progress = (
            '<td style="text-align:center;padding:0 6px;">'
            '<div style="display:inline-flex;flex-direction:column;align-items:center;gap:5px;">'
            '<div style="width:38px;height:38px;border-radius:50%;background:#d1fae5;color:#065f46;'
            'display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;">&#10003;</div>'
            '<div style="font-size:10px;font-weight:600;color:#065f46;">Contact</div>'
            '<div style="font-size:9px;color:#94a3b8;">APPROVED</div>'
            '</div></td>'
            '<td style="width:20px;"><div style="height:2px;background:#10b981;margin:0 2px;margin-bottom:28px;"></div></td>'
            '<td style="text-align:center;padding:0 6px;">'
            '<div style="display:inline-flex;flex-direction:column;align-items:center;gap:5px;">'
            '<div style="width:38px;height:38px;border-radius:50%;background:#fef3c7;color:#92400e;'
            'display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;'
            'border:2px solid #2563eb;box-shadow:0 0 0 3px rgba(37,99,235,0.15);">&#9654;</div>'
            '<div style="font-size:10px;font-weight:600;color:#92400e;">Security</div>'
            '<div style="font-size:9px;color:#94a3b8;">AWAITING</div>'
            '</div></td>'
        )

        html_body = (
            '<!DOCTYPE html><html><head><meta charset="UTF-8"></head>'
            '<body style="margin:0;padding:0;background:#f0f4f8;font-family:Inter,Arial,sans-serif;">'
            '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:32px 0;"><tr><td align="center">'
            '<table width="620" cellpadding="0" cellspacing="0" style="background:white;border-radius:20px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,0.12);">'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:4px;"></td></tr>'
            '<tr><td style="background:linear-gradient(135deg,#0f172a,#1e3a5f,#1e40af);padding:36px 40px;text-align:center;">'
            '<div style="font-size:11px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;">Unity Cement &mdash; ERP Department</div>'
            '<div style="width:64px;height:64px;border-radius:50%;background:rgba(255,255,255,0.12);border:2px solid rgba(255,255,255,0.25);display:inline-flex;align-items:center;justify-content:center;margin-bottom:16px;font-size:28px;">&#128737;</div>'
            '<div style="font-size:26px;font-weight:800;color:white;margin-bottom:6px;">Security Clearance Required</div>'
            '<div style="font-size:14px;color:rgba(255,255,255,0.7);margin-bottom:16px;">Stage 2 of 2: <strong style="color:white;">Security Approval</strong></div>'
            f'<div style="display:inline-block;background:rgba(240,165,0,0.2);border:1.5px solid rgba(240,165,0,0.5);border-radius:50px;padding:6px 20px;font-size:14px;color:#fbbf24;font-weight:700;">{vgp.pass_number}</div>'
            '</td></tr>'
            '<tr><td style="padding:24px 40px;background:#f8fafc;border-bottom:1px solid #e2e8f0;">'
            '<div style="font-size:11px;font-weight:700;color:#94a3b8;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px;">Approval Progress</div>'
            '<table cellpadding="0" cellspacing="0" style="width:100%;"><tr style="vertical-align:middle;">' + progress + '</tr></table>'
            '</td></tr>'
            f'<tr><td style="padding:28px 40px 0;"><p style="font-size:16px;color:#1e293b;margin:0 0 8px;">Dear <strong style="color:#1e3a5f;">{security.employee_name}</strong>,</p>'
            f'<p style="font-size:14px;color:#64748b;margin:0;line-height:1.6;">Visitor <strong>{vgp.visitor_name}</strong> has been approved by the contact person. Please provide security clearance.</p></td></tr>'
            '<tr><td style="padding:20px 40px;">'
            '<div style="border-radius:12px;overflow:hidden;border:1px solid #e2e8f0;">'
            '<div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);padding:12px 16px;"><span style="font-size:11px;font-weight:700;color:white;letter-spacing:1.5px;text-transform:uppercase;">&#128203; Visitor Details</span></div>'
            '<table width="100%" cellpadding="0" cellspacing="0">'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;width:38%;border-bottom:1px solid #f1f5f9;">Visitor</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{vgp.visitor_name}</td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Contact Person</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{vgp.person_to_meet.employee_name}</td></tr>'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Purpose</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{vgp.get_visit_purpose_display()}</td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;">Date</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;">{vgp.visit_date}</td></tr>'
            '</table></div></td></tr>'
            '<tr><td style="padding:8px 40px 32px;">'
            '<div style="background:#f8fafc;border-radius:12px;padding:24px;border:1px solid #e2e8f0;">'
            '<p style="font-size:13px;color:#475569;font-weight:600;margin:0 0 16px;text-align:center;">&#9889; Take Action Now</p>'
            '<table cellpadding="0" cellspacing="0" width="100%"><tr>'
            f'<td width="47%" align="center"><a href="{approve_url}" style="display:block;background:linear-gradient(135deg,#059669,#10b981);color:white;text-decoration:none;padding:16px 24px;border-radius:12px;font-size:16px;font-weight:800;text-align:center;box-shadow:0 4px 14px rgba(16,185,129,0.4);">&#10003;&nbsp;&nbsp;APPROVE</a></td>'
            '<td width="6%"></td>'
            f'<td width="47%" align="center"><a href="{reject_url}" style="display:block;background:linear-gradient(135deg,#dc2626,#ef4444);color:white;text-decoration:none;padding:16px 24px;border-radius:12px;font-size:16px;font-weight:800;text-align:center;box-shadow:0 4px 14px rgba(239,68,68,0.4);">&#10007;&nbsp;&nbsp;REJECT</a></td>'
            '</tr></table>'
            '<p style="font-size:11px;color:#94a3b8;text-align:center;margin:14px 0 0;">Clicking a button opens a secure confirmation page.</p>'
            '</div></td></tr>'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:3px;"></td></tr>'
            '<tr><td style="background:#f8fafc;padding:20px 40px;text-align:center;">'
            '<p style="font-size:11px;color:#94a3b8;margin:0;">This is an automated email from <strong>ERP Department &mdash; Unity Cement (PMC Cement Ltd.)</strong></p>'
            '</td></tr>'
            '</table></td></tr></table></body></html>'
        )
        plain = f"""Dear {security.employee_name},

VGP {vgp.pass_number} requires security clearance.
Visitor: {vgp.visitor_name}
Contact Person: {vgp.person_to_meet.employee_name} (APPROVED)

APPROVE: {approve_url}
REJECT : {reject_url}

Regards,
ERP Department — Unity Cement"""
        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [security.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def send_visitor_status_mail(vgp):
    try:
        if not vgp.visitor_email:
            return
        subject = 'Your Visit Request - %s - %s' % (vgp.get_status_display(), vgp.pass_number)
        body = """Dear %s,\n\nYour visit request (%s) has been %s.\n\nPerson to Meet: %s\nDate: %s\nRemarks: %s\n\nRegards,\nERP Department""" % (
            vgp.visitor_name, vgp.pass_number, vgp.get_status_display(),
            vgp.person_to_meet.employee_name, vgp.visit_date, vgp.approval_remarks or 'N/A')
        send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [vgp.visitor_email], fail_silently=True)
    except Exception:
        pass


@login_required
def pass_list(request):
    if not request.user.can_access_visitor_pass and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')
    user = request.user
    FULL_ROLES = ('administrator', 'management', 'president_plant_head')
    if user.is_superuser or user.role in FULL_ROLES or user.can_approve_visitor_pass:
        passes = VisitorGatePass.objects.all()
    else:
        passes = VisitorGatePass.objects.filter(
            person_to_meet__department=user.department
        ) | VisitorGatePass.objects.filter(person_to_meet=user)
        passes = passes.distinct()
    status_filter = request.GET.get('status', '')
    dept_filter   = request.GET.get('dept', '')
    search        = request.GET.get('q', '')
    if status_filter: passes = passes.filter(status=status_filter)
    if dept_filter and (user.is_superuser or user.role in FULL_ROLES):
        passes = passes.filter(person_to_meet__department=dept_filter)
    if search:
        passes = passes.filter(
            models.Q(visitor_name__icontains=search) |
            models.Q(visitor_phone__icontains=search) |
            models.Q(visitor_company__icontains=search) |
            models.Q(pass_number__icontains=search)
        )
    from accounts.models import DEPARTMENT_CHOICES
    from django.core.paginator import Paginator
    paginator = Paginator(passes, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    return render(request, 'visitor_pass/pass_list.html', {
        'passes': page_obj, 'page_obj': page_obj,
        'status_filter': status_filter, 'search': search,
        'dept_filter': dept_filter, 'department_choices': DEPARTMENT_CHOICES,
        'show_dept_filter': user.is_superuser or user.role in FULL_ROLES,
    })


@login_required
def pass_create(request):
    if not request.user.can_access_visitor_pass and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')
    form = VisitorGatePassForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        vgp = form.save(commit=False)
        vgp.created_by = request.user
        vgp.save()
        send_visitor_approval_request(vgp, request)
        from accounts.models import AuditLog
        AuditLog.log(request, 'vgp_create', 'VGP', f'VGP created: {vgp.pass_number} for {vgp.visitor_name} by {request.user.username}')
        messages.success(request, 'Visitor Gate Pass %s created. Please capture visitor photo.' % vgp.pass_number)
        return redirect('visitor_pass:capture_photo', pk=vgp.pk)
    workflow_steps = [
        'Requester creates Visitor Gate Pass',
        'System sends approval notification to Contact Person',
        'Contact Person approves or rejects the request',
        'Decision notification reflects to requester, approver workflow users, and admin',
        'If approved, visitor can enter and later be checked out by security/authorized user',
    ]
    return render(request, 'visitor_pass/pass_form.html', {
        'form': form,
        'title': 'New Visitor Gate Pass',
        'workflow_steps': workflow_steps,
    })


@login_required
def capture_photo(request, pk):
    vgp = get_object_or_404(VisitorGatePass, pk=pk)
    if request.method == 'POST':
        webcam_data = request.POST.get('webcam_data')
        if webcam_data and webcam_data.startswith('data:image'):
            fmt, imgstr = webcam_data.split(';base64,')
            ext = fmt.split('/')[-1]
            vgp.visitor_photo.save('visitor_%s.%s' % (vgp.pass_number, ext),
                                   ContentFile(base64.b64decode(imgstr)), save=True)
            messages.success(request, 'Photo captured successfully.')
            return redirect('visitor_pass:pass_detail', pk=vgp.pk)
        form = PhotoCaptureForm(request.POST, request.FILES, instance=vgp)
        if form.is_valid():
            form.save()
            messages.success(request, 'Photo uploaded successfully.')
            return redirect('visitor_pass:pass_detail', pk=vgp.pk)
    else:
        form = PhotoCaptureForm(instance=vgp)
    return render(request, 'visitor_pass/capture_photo.html', {'vgp': vgp, 'form': form})


@login_required
def pass_detail(request, pk):
    vgp = get_object_or_404(VisitorGatePass, pk=pk)
    return render(request, 'visitor_pass/pass_detail.html', {'vgp': vgp})


@login_required
def pass_approve(request, pk):
    if not request.user.can_approve_visitor_pass and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('visitor_pass:pass_list')
    vgp = get_object_or_404(VisitorGatePass, pk=pk)
    form = VisitorApprovalForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        vgp.status = form.cleaned_data['action']
        vgp.approval_remarks = form.cleaned_data['remarks']
        vgp.approved_by = request.user
        vgp.approved_at = timezone.now()
        vgp.save()
        send_visitor_status_mail(vgp)
        from accounts.models import AuditLog
        action_key = 'vgp_approve' if vgp.status == 'approved' else 'vgp_reject'
        AuditLog.log(request, action_key, 'VGP', f'VGP {vgp.status}: {vgp.pass_number} by {request.user.username}')
        
        notification_type = 'vgp_approved' if vgp.status == 'approved' else 'vgp_rejected'
        title = f'VGP {vgp.pass_number} {vgp.status.title()}'
        description = f'Your Visitor Gate Pass has been {vgp.status}'
        send_workflow_notification(
            module_key='vgp',
            notification_type=notification_type,
            title=title,
            description=description,
            related_id=str(vgp.id),
            related_module='VGP',
            requester=getattr(vgp, 'created_by', None),
            extra_users=[vgp.person_to_meet] if vgp.person_to_meet else None,
        )
        
        messages.success(request, 'Visitor pass %s.' % vgp.status)
        return redirect('visitor_pass:pass_list')
    return render(request, 'visitor_pass/pass_approve.html', {'vgp': vgp, 'form': form})


@login_required
def checkout(request, pk):
    vgp = get_object_or_404(VisitorGatePass, pk=pk)
    form = CheckoutForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        vgp.actual_out_time = form.cleaned_data['actual_out_time']
        vgp.status = 'checked_out'
        vgp.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'vgp_checkout', 'VGP', f'VGP checked out: {vgp.pass_number} by {request.user.username}')
        messages.success(request, 'Visitor checked out.')
        return redirect('visitor_pass:pass_list')
    return render(request, 'visitor_pass/checkout.html', {'vgp': vgp, 'form': form})


@login_required
def print_pass(request, pk):
    vgp = get_object_or_404(VisitorGatePass, pk=pk)
    buf = BytesIO()

    # Match the printed layout in the provided sample: A5 landscape (210 × 148 mm)
    PAGE_W = 210 * mm
    PAGE_H = 148 * mm
    MARGIN = 6 * mm  # small margin; outer page border is drawn inside it

    # Inner content area (inside outer border)
    USABLE_W = PAGE_W - 2 * MARGIN
    USABLE_H = PAGE_H - 2 * MARGIN

    black = colors.black

    def S(name, **kw):
        kw.setdefault('fontName', 'Helvetica')
        kw.setdefault('fontSize', 7)
        return ParagraphStyle(name, **kw)

    def fmt_time(t):
        if not t: return '-'
        h, m = t.hour, t.minute
        return '%d:%02d %s' % (h % 12 or 12, m, 'AM' if h < 12 else 'PM')

    lbl = S('lbl', fontName='Helvetica-Bold', fontSize=9)
    val = S('val', fontSize=9)

    def row(label, value):
        v = value if isinstance(value, Paragraph) else Paragraph(str(value) if value else '-', val)
        return [Paragraph(label + ' :', lbl), v]

    d = vgp.visit_date
    date_str = '%02d.%02d.%d' % (d.day, d.month, d.year)

    # Checkbox for material category
    nr_box = '\u2611' if vgp.material_category == 'non_returnable' else '\u2610'
    r_box  = '\u2611' if vgp.material_category == 'returnable'     else '\u2610'
    mat_cat_para = Paragraph('%s Non - Returnable&nbsp;&nbsp;&nbsp;&nbsp;%s Returnable' % (nr_box, r_box), val)

    # ─────── PHOTO SECTION (define before LEFT_W) ───────
    PHOTO_W = 60 * mm
    PHOTO_H = 60 * mm

    # ─────── LEFT TABLE SECTION ───────
    # Keep photo tight on the right (no extra empty column space).
    LEFT_W = USABLE_W - PHOTO_W
    COL1 = 40 * mm      # Label column
    COL2 = LEFT_W - COL1
    ROW_H = 4.5 * mm    # Row height

    ts = TableStyle([
        ('GRID',          (0,0),(-1,-1), 0.5, black),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0),(-1,-1), 1),
        ('BOTTOMPADDING', (0,0),(-1,-1), 1),
        ('LEFTPADDING',   (0,0),(-1,-1), 2),
        ('RIGHTPADDING',  (0,0),(-1,-1), 2),
    ])

    info_table = Table([
        row('Request ID',         vgp.pass_number),
        row('Date',               date_str),
        row("Visitor's Name",     vgp.visitor_name),
        row('Visitor Mobile No.', vgp.visitor_phone),
        row('Organization',       vgp.visitor_company or '-'),
        row('City',               vgp.visitor_city or '-'),
        row('Material',           vgp.material or '-'),
        [Paragraph('Material Category :', lbl), mat_cat_para],
        row('Purpose',            vgp.get_visit_purpose_display() or '-'),
        row('Access Card',        vgp.access_card_no or '-'),
    ], colWidths=[COL1, COL2])
    info_table.setStyle(ts)

    if vgp.visitor_photo and os.path.exists(vgp.visitor_photo.path):
        try:    photo_cell = Image(vgp.visitor_photo.path, width=PHOTO_W - 2*mm, height=PHOTO_H - 2*mm)
        except: photo_cell = None
    else:
        photo_cell = None

    if photo_cell is None:
        from reportlab.platypus import Flowable
        class SilhouettePlaceholder(Flowable):
            def __init__(self, w, h):
                self.width, self.height = w, h
            def draw(self):
                c = self.canv
                c.setFillColor(colors.HexColor('#3d3d3d'))
                c.rect(0, 0, self.width, self.height, fill=1, stroke=0)
                cx = self.width / 2
                head_r = self.width * 0.22
                head_cy = self.height * 0.62
                c.setFillColor(colors.HexColor('#b0b0b0'))
                c.circle(cx, head_cy, head_r, fill=1, stroke=0)
                body_w = self.width * 0.55
                body_h = self.height * 0.32
                body_cx = cx
                body_cy = self.height * 0.18
                c.ellipse(body_cx - body_w/2, body_cy,
                          body_cx + body_w/2, body_cy + body_h, fill=1, stroke=0)
        photo_cell = SilhouettePlaceholder(PHOTO_W, PHOTO_H)

    photo_box = Table([[photo_cell]], colWidths=[PHOTO_W], rowHeights=[PHOTO_H])
    photo_box.setStyle(TableStyle([
        ('ALIGN',  (0,0),(-1,-1), 'CENTER'),
        ('VALIGN', (0,0),(-1,-1), 'MIDDLE'),
    ]))

    # ─────── TOP SECTION: Header (Company centered, Logo on right) ───────
    LOGO_W = 42 * mm
    LOGO_H = 18 * mm

    logo_path = os.path.join(django_settings.BASE_DIR, 'static', 'images', 'Unity_Logo_Horizontal.png')

    def fit_logo(path, box_w, box_h):
        """Scale image to fit box while preserving aspect ratio."""
        ir = ImageReader(path)
        iw, ih = ir.getSize()
        if not iw or not ih:
            return Image(path, width=box_w, height=box_h)
        scale = min(box_w / iw, box_h / ih)
        return Image(path, width=iw * scale, height=ih * scale)

    if os.path.exists(logo_path):
        logo_cell = fit_logo(logo_path, LOGO_W, LOGO_H)
    else:
        logo_cell = Paragraph('UNITY', S('lc', fontName='Helvetica-Bold', fontSize=12, alignment=TA_CENTER))

    header_style = S('header', fontName='Helvetica-Bold', fontSize=14, alignment=TA_CENTER)
    title_style = S('title', fontName='Helvetica-Bold', fontSize=11, alignment=TA_CENTER)

    mid_w = USABLE_W - (2 * LOGO_W)
    header_text = Table(
        [
            [Paragraph('PMC CEMENT LIMITED', header_style)],
            [Spacer(2, 2.5*mm)],
            [Paragraph('Visitor Pass', title_style)],
        ],
        colWidths=[mid_w],
    )
    header_text.setStyle(TableStyle([
        ('ALIGN', (0,0),(-1,-1), 'CENTER'),
        ('VALIGN', (0,0),(-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0),(-1,-1), 0),
        ('BOTTOMPADDING', (0,0),(-1,-1), 0),
    ]))

    # 3-column header keeps title perfectly centered on the page
    header = Table([[Spacer(1, 1), header_text, logo_cell]],
                   colWidths=[LOGO_W, mid_w, LOGO_W])
    header.setStyle(TableStyle([
        ('VALIGN', (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
        ('ALIGN', (2,0), (2,0), 'RIGHT'),
        ('LEFTPADDING', (0,0),(-1,-1), 0),
        ('RIGHTPADDING', (0,0),(-1,-1), 0),
        ('TOPPADDING', (0,0),(-1,-1), 0),
        ('BOTTOMPADDING', (0,0),(-1,-1), 3),
        ('LINEBELOW', (0,0),(-1,-1), 1.8, black),  # bold separator under header
    ]))

    # ─────── BOTTOM TABLE: Contact + Time ───────
    contact_table = Table([
        row('Contact person', vgp.person_to_meet.employee_name),
        row('In time',        fmt_time(vgp.in_time)),
        row('Out time',       fmt_time(vgp.actual_out_time or vgp.expected_out_time)),
    ], colWidths=[COL1, COL2])
    contact_table.setStyle(ts)

    # ─────── MAIN LAYOUT: Header + Info/Photo + Contact ───────
    body_row = Table([[info_table, photo_box]],
                     colWidths=[LEFT_W, USABLE_W - LEFT_W])
    body_row.setStyle(TableStyle([
        ('VALIGN', (0,0),(-1,-1), 'TOP'),
        ('LEFTPADDING',  (0,0),(-1,-1), 0),
        ('RIGHTPADDING', (0,0),(-1,-1), 0),
        ('TOPPADDING',   (0,0),(-1,-1), 0),
        ('BOTTOMPADDING',(0,0),(-1,-1), 0),
    ]))

    main_pass = Table(
        [
            [header],
            [Spacer(1, 2*mm)],
            [body_row],
            [contact_table],
        ],
        colWidths=[USABLE_W],
    )
    main_pass.setStyle(TableStyle([
        ('LEFTPADDING',   (0,0),(-1,-1), 0),
        ('RIGHTPADDING',  (0,0),(-1,-1), 0),
        ('TOPPADDING',    (0,0),(-1,-1), 0),
        ('BOTTOMPADDING', (0,0),(-1,-1), 0),
    ]))

    # ─────── SIGNATURE SECTION (3 columns) ───────
    sig_style = S('sig', alignment=TA_CENTER, fontSize=9, fontName='Helvetica-Bold')
    sig_name_style = S('signame', alignment=TA_CENTER, fontSize=8, fontName='Helvetica-Bold')
    SIG_W = USABLE_W / 3

    sig_table = Table([
        [Paragraph(vgp.created_by.employee_name if vgp.created_by else 'Security', sig_name_style),
         Paragraph(vgp.visitor_name or '-', sig_name_style),
         Paragraph(vgp.approved_by.employee_name if vgp.approved_by else '-', sig_name_style)],
        [Paragraph('Security Sign', sig_style),
         Paragraph('Visitor Sign', sig_style),
         Paragraph('Employee Sign', sig_style)],
    ], colWidths=[SIG_W, SIG_W, SIG_W])
    sig_table.setStyle(TableStyle([
        ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
        ('TOPPADDING',    (0,0),(-1,-1), 2*mm),
        ('BOTTOMPADDING', (0,0),(-1,-1), 0),
    ]))

    # ─────── FOOTER ───────
    footer_style = S('footer', alignment=TA_CENTER, fontSize=10, fontName='Helvetica-BoldOblique')
    footer = Paragraph('"We hope you had a pleasant experience. We look forward to welcoming you again."', footer_style)

    # ─────── PAGE LAYOUT ───────
    def draw_border(canvas, doc):
        canvas.saveState()
        # Outer page border (like the sample)
        canvas.setLineWidth(1.6)
        canvas.rect(MARGIN, MARGIN, USABLE_W, USABLE_H, fill=0, stroke=1)
        canvas.restoreState()

    frame = Frame(MARGIN, MARGIN, USABLE_W, USABLE_H,
                  leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    pt = PageTemplate(id='main', frames=[frame], onPage=draw_border)
    doc = BaseDocTemplate(buf, pagesize=landscape(A5),
                          leftMargin=MARGIN, rightMargin=MARGIN, 
                          topMargin=MARGIN, bottomMargin=MARGIN)
    doc.addPageTemplates([pt])

    # Build document (main section + footer + signatures)
    doc.build([
        main_pass,
        Spacer(1, 6*mm),
        footer,
        Spacer(1, 4*mm),
        sig_table,
    ])

    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="VGP_%s.pdf"' % vgp.pass_number
    return response


@login_required
def print_preview(request, pk):
    vgp = get_object_or_404(VisitorGatePass, pk=pk)
    
    d = vgp.visit_date
    date_str = '%02d.%02d.%d' % (d.day, d.month, d.year)
    
    # Checkbox for material category
    nr_checked = vgp.material_category == 'non_returnable'
    r_checked = vgp.material_category == 'returnable'
    
    def fmt_time(t):
        if not t: return '-'
        h, m = t.hour, t.minute
        return '%d:%02d %s' % (h % 12 or 12, m, 'AM' if h < 12 else 'PM')
    
    info_fields = [
        ('Request ID', vgp.pass_number),
        ('Date', date_str),
        ("Visitor's Name", vgp.visitor_name),
        ('Visitor Mobile No.', vgp.visitor_phone),
        ('Organization', vgp.visitor_company or '-'),
        ('City', vgp.visitor_city or '-'),
        ('Material', vgp.material or '-'),
        ('Purpose', vgp.get_visit_purpose_display() or '-'),
        ('Access Card No.', vgp.access_card_no or '-'),
    ]
    
    contact_fields = [
        ('Contact person', vgp.person_to_meet.employee_name if vgp.person_to_meet else '-'),
        ('In time', fmt_time(vgp.in_time)),
        ('Out time', fmt_time(vgp.actual_out_time or vgp.expected_out_time)),
    ]
    
    context = {
        'vgp': vgp,
        'date_str': date_str,
        'nr_checked': nr_checked,
        'r_checked': r_checked,
        'info_fields': info_fields,
        'contact_fields': contact_fields,
    }
    
    return render(request, 'visitor_pass/print_preview.html', context)


@login_required
def bulk_action(request):
    if request.method != 'POST':
        return redirect('visitor_pass:pass_list')
    action = request.POST.get('action')
    ids    = request.POST.getlist('selected_ids')
    if not ids:
        messages.warning(request, 'No passes selected.')
        return redirect('visitor_pass:pass_list')
    passes = VisitorGatePass.objects.filter(pk__in=ids)
    if action == 'delete':
        if not request.user.is_superuser and not request.user.can_approve_visitor_pass:
            messages.error(request, 'Access Denied.')
            return redirect('visitor_pass:pass_list')
        count = passes.count()
        passes.delete()
        messages.success(request, f'{count} pass(es) deleted.')
    elif action == 'cancel':
        updated = passes.exclude(status__in=['checked_out']).update(status='rejected')
        messages.success(request, f'{updated} pass(es) cancelled.')
    elif action == 'duplicate':
        import uuid as _uuid
        for vgp in passes:
            vgp.pk          = None
            vgp.pass_number = ''
            vgp.status      = 'pending'
            vgp.approved_by = None
            vgp.approved_at = None
            vgp.approval_remarks = ''
            vgp.actual_out_time  = None
            vgp.visitor_photo    = None
            vgp.approval_token   = _uuid.uuid4()
            vgp.created_by       = request.user
            vgp.save()
        messages.success(request, f'{len(ids)} pass(es) duplicated.')
    else:
        messages.error(request, 'Unknown action.')
    return redirect('visitor_pass:pass_list')


def token_action(request, token, action):
    vgp = get_object_or_404(VisitorGatePass, approval_token=token)
    is_security = 'security' in action
    clean_action = action.replace('security-', '')

    # Block if already fully processed
    if vgp.status in ('approved', 'rejected', 'checked_out'):
        return render(request, 'visitor_pass/token_done.html', {
            'message': f'This pass has already been {vgp.get_status_display()}. No further action is required.',
            'vgp': vgp
        })
    if vgp.status not in ('pending', 'contact_approved'):
        return render(request, 'visitor_pass/token_done.html', {
            'message': f'This pass has already been {vgp.status}.', 'vgp': vgp
        })
    if request.method == 'POST':
        chosen  = request.POST.get('action_choice', clean_action)
        remarks = request.POST.get('remarks', '')
        if chosen == 'approve':
            if is_security:
                # Final approval by security
                vgp.status = 'approved'
                vgp.approved_by = request.user if request.user.is_authenticated else None
                vgp.approval_remarks = remarks
                vgp.approved_at = timezone.now()
                vgp.save()
                send_visitor_status_mail(vgp)
            else:
                # Contact person approved — now send to security
                vgp.status = 'contact_approved'
                vgp.approval_remarks = remarks
                vgp.save()
                _send_vgp_security_mail(vgp, request)
        else:
            vgp.status = 'rejected'
            vgp.approval_remarks = remarks
            vgp.approved_at = timezone.now()
            vgp.save()
            send_visitor_status_mail(vgp)

        send_workflow_notification(
            module_key='vgp',
            notification_type='vgp_approved' if vgp.status == 'approved' else 'vgp_rejected',
            title=f'VGP {vgp.pass_number} {vgp.status.title()}',
            description=f'Visitor Gate Pass has been {vgp.status}',
            related_id=str(vgp.id),
            related_module='VGP',
            requester=getattr(vgp, 'created_by', None),
            extra_users=[vgp.person_to_meet] if vgp.person_to_meet else None,
        )
        stage_msg = 'Contact person approved. Security clearance email sent.' if vgp.status == 'contact_approved' else f'Visitor pass {vgp.pass_number} has been {vgp.status}.'
        return render(request, 'visitor_pass/token_done.html', {
            'message': stage_msg, 'vgp': vgp
        })
    return render(request, 'visitor_pass/token_action.html', {
        'vgp': vgp, 'action': clean_action, 'is_security': is_security
    })


@login_required
def daily_report(request):
    report_date = request.GET.get('date', str(date.today()))
    dept = request.GET.get('dept', '')
    view_mode = request.GET.get('view', 'detail')  # summary | detail
    passes = VisitorGatePass.objects.filter(visit_date=report_date)
    if dept: passes = passes.filter(person_to_meet__department=dept)
    from accounts.models import DEPARTMENT_CHOICES
    return render(request, 'visitor_pass/report.html', {
        'passes': passes, 'report_date': report_date, 'report_type': 'Daily',
        'dept_filter': dept, 'department_choices': DEPARTMENT_CHOICES,
        'view_mode': view_mode,
        'summary': _vgp_summary(passes),
    })


@login_required
def monthly_report(request):
    month = request.GET.get('month', date.today().strftime('%Y-%m'))
    year, mon = month.split('-')
    dept = request.GET.get('dept', '')
    view_mode = request.GET.get('view', 'detail')  # summary | detail
    passes = VisitorGatePass.objects.filter(visit_date__year=year, visit_date__month=mon)
    if dept: passes = passes.filter(person_to_meet__department=dept)
    from accounts.models import DEPARTMENT_CHOICES
    return render(request, 'visitor_pass/report.html', {
        'passes': passes, 'report_date': month, 'report_type': 'Monthly',
        'dept_filter': dept, 'department_choices': DEPARTMENT_CHOICES,
        'view_mode': view_mode,
        'summary': _vgp_summary(passes),
    })


def _vgp_summary(qs):
    """Build summary aggregates for visitor pass reports."""
    purpose_map = dict(VisitorGatePass._meta.get_field('visit_purpose').choices)
    status_map = dict(VisitorGatePass._meta.get_field('status').choices)

    total = qs.count()
    total_visitors = qs.aggregate(s=Sum('no_of_visitors'))['s'] or 0

    by_status = []
    for row in qs.values('status').annotate(count=Count('id')).order_by('-count'):
        key = row['status'] or ''
        by_status.append({'key': key, 'label': status_map.get(key, key or '—'), 'count': row['count']})

    by_purpose = []
    for row in qs.values('visit_purpose').annotate(count=Count('id')).order_by('-count'):
        key = row['visit_purpose'] or ''
        by_purpose.append({'key': key, 'label': purpose_map.get(key, key or '—'), 'count': row['count']})

    by_department = []
    for row in qs.values('person_to_meet__department').annotate(count=Count('id')).order_by('-count'):
        key = row['person_to_meet__department'] or ''
        by_department.append({'label': key or '—', 'count': row['count']})

    return {
        'total': total,
        'total_visitors': total_visitors,
        'by_status': by_status,
        'by_purpose': by_purpose,
        'by_department': by_department,
    }


@login_required
def export_report(request):
    report_type = request.GET.get('type', 'daily')
    report_date = request.GET.get('date', str(date.today()))
    if report_type == 'monthly':
        year, mon = report_date.split('-')
        passes = VisitorGatePass.objects.filter(visit_date__year=year, visit_date__month=mon)
    else:
        passes = VisitorGatePass.objects.filter(visit_date=report_date)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Visitor Gate Pass Report'
    headers = ['Pass No', 'Visitor Name', 'Company', 'Phone', 'ID Type', 'ID No',
               'No. of Visitors', 'Purpose', 'Person to Meet', 'Date', 'In Time',
               'Exp. Out', 'Actual Out', 'Vehicle No',
               'Approved By', 'Approval Date & Time', 'Status']
    hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18
    for vgp in passes:
        approved_at_str = vgp.approved_at.strftime('%d-%m-%Y %H:%M') if vgp.approved_at else ''
        ws.append([
            vgp.pass_number, vgp.visitor_name, vgp.visitor_company, vgp.visitor_phone,
            vgp.get_id_type_display(), vgp.id_number, vgp.no_of_visitors,
            vgp.get_visit_purpose_display(), vgp.person_to_meet.employee_name,
            str(vgp.visit_date), str(vgp.in_time), str(vgp.expected_out_time),
            str(vgp.actual_out_time or ''), vgp.vehicle_number,
            vgp.approved_by.employee_name if vgp.approved_by else '',
            approved_at_str, vgp.get_status_display()
        ])
    from accounts.report_utils import add_excel_logo_and_note
    add_excel_logo_and_note(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="VGP_Report_%s.xlsx"' % report_date
    return response


@login_required
def export_report_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    report_type = request.GET.get('type', 'daily')
    report_date = request.GET.get('date', str(date.today()))
    if report_type == 'monthly':
        year, mon = report_date.split('-')
        passes = VisitorGatePass.objects.filter(visit_date__year=year, visit_date__month=mon)
    else:
        passes = VisitorGatePass.objects.filter(visit_date=report_date)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=12*mm, bottomMargin=10*mm,
                            leftMargin=10*mm, rightMargin=10*mm)

    navy  = colors.HexColor('#1e3a5f')
    alt   = colors.HexColor('#f8f9fa')

    title_style = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=14,
                                 alignment=TA_CENTER, textColor=navy, spaceAfter=2)
    sub_style   = ParagraphStyle('s', fontName='Helvetica', fontSize=9,
                                 alignment=TA_CENTER, textColor=colors.grey, spaceAfter=6)

    elements = []
    from accounts.report_utils import build_pdf_header_table
    elements.append(build_pdf_header_table(
        'VISITOR GATE PASS REPORT',
        '%s Report \u2014 %s  |  Total Records: %d  |  Generated: %s' % (
            report_type, report_date, passes.count(), timezone.now().strftime('%d %b %Y %H:%M')),
        277
    ))
    elements.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=4*mm))

    headers = ['#', 'Pass No', 'Visitor Name', 'Company', 'Phone',
               'Purpose', 'Contact Person', 'Date', 'In', 'Exp.Out', 'Act.Out', 'Visitors', 'Status']
    col_w = [8*mm, 22*mm, 32*mm, 30*mm, 24*mm,
             26*mm, 30*mm, 20*mm, 14*mm, 14*mm, 14*mm, 14*mm, 18*mm]

    data = [headers]
    for i, vgp in enumerate(passes, 1):
        data.append([
            str(i), vgp.pass_number, vgp.visitor_name,
            vgp.visitor_company or '-', vgp.visitor_phone,
            vgp.get_visit_purpose_display(), vgp.person_to_meet.employee_name,
            str(vgp.visit_date), str(vgp.in_time), str(vgp.expected_out_time),
            str(vgp.actual_out_time or '-'), str(vgp.no_of_visitors),
            vgp.get_status_display(),
        ])

    if len(data) == 1:
        data.append(['', 'No records found for this period.'] + [''] * 11)

    table = Table(data, colWidths=col_w, repeatRows=1)
    table.setStyle(TableStyle([
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0),  8),
        ('BACKGROUND',  (0, 0), (-1, 0),  navy),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 1), (-1, -1), 7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt]),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#dee2e6')),
        ('LINEABOVE',   (0, 0), (-1, 0),  1.5, navy),
        ('LINEBELOW',   (0, 0), (-1, 0),  1.5, navy),
        ('ALIGN',       (0, 0), (0, -1),  'CENTER'),
        ('ALIGN',       (7, 0), (12, -1), 'CENTER'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',     (0, 0), (-1, -1), 3),
        ('WORDWRAP',    (0, 0), (-1, -1), True),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph('This is a computer generated report. \u2014 Unity Cement ERP System', sub_style))

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="VGP_Report_%s.pdf"' % report_date
    return response
