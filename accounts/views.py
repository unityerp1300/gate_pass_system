from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import secrets
from .models import Employee, Notification
from .forms import LoginForm, EmployeeForm, PasswordResetForm, ForgotPasswordForm, ForgotPasswordSetForm
from django.core.mail import send_mail
from django.conf import settings as django_settings
from .models import SystemSetting, AdminOTPSession, AuditLog, RolePermissionTemplate, RoleModuleAccess, DEPARTMENT_CHOICES
from .models import ROLE_CHOICES
from .models import EmailLog


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard:index')
    reason = request.GET.get('reason', '')
    form = LoginForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        if not user.is_active:
            messages.error(request, 'Your account has been disabled. Please contact the administrator.')
            AuditLog.log(request, 'login_failed', 'System', f'Disabled account login attempt: {user.username}')
            return render(request, 'accounts/login.html', {'form': form, 'reason': reason})
        if user.has_role('administrator') or user.is_superuser:
            if user.must_change_password:
                login(request, user)
                return redirect('accounts:reset_password')
            otp_session = AdminOTPSession.generate_for(user)
            _send_admin_otp_mail(user, otp_session.otp_code)
            request.session['otp_user_id'] = user.pk
            return redirect('accounts:admin_otp_verify')
        login(request, user)
        if user.must_change_password:
            return redirect('accounts:reset_password')
        Employee.objects.filter(pk=user.pk).update(session_key=request.session.session_key)
        request.session['last_activity'] = timezone.now().timestamp()
        AuditLog.log(request, 'login', 'System', f'{user.username} logged in successfully', user=user)
        return redirect('dashboard:index')
    if request.method == 'POST':
        AuditLog.log(request, 'login_failed', 'System',
                     f'Failed login attempt: {request.POST.get("username", "")}')
    return render(request, 'accounts/login.html', {'form': form, 'reason': reason})


def logout_view(request):
    if request.user.is_authenticated:
        AuditLog.log(request, 'logout', 'System', f'{request.user.username} logged out')
        Employee.objects.filter(pk=request.user.pk).update(session_key=None)
    logout(request)
    return redirect('accounts:login')


def session_check(request):
    """AJAX endpoint polled every 30s to detect server-side kick or timeout."""
    if request.user.is_authenticated:
        return JsonResponse({'ok': True})
    return JsonResponse({'timeout': True}, status=401)


@login_required
def get_notifications(request):
    """AJAX endpoint to get user's recent notifications respecting permissions."""
    from accounts.models import NotificationPermission
    
    # Get notifications visible to this user based on permissions
    visible_notifications = NotificationPermission.get_visible_notifications(request.user)
    
    # Count unread notifications visible to user
    unread_count = visible_notifications.filter(is_read=False).count()
    notifications = visible_notifications[:10]
    
    data = {
        'unread_count': min(unread_count, 99),  # Cap at 99 for UI
        'notifications': [
            {
                'id': n.id,
                'type': n.notification_type,
                'title': n.title,
                'description': n.description,
                'module': n.related_module,
                'is_read': n.is_read,
                'created_at': n.created_at.isoformat(),
                'created_at_display': n.created_at.strftime('%b %d, %I:%M %p'),
            }
            for n in notifications
        ]
    }
    return JsonResponse(data)


@login_required
def mark_notifications_read(request):
    """AJAX endpoint to mark notifications as read."""
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return JsonResponse({'ok': True})


@login_required
def clear_notifications(request):
    """AJAX endpoint to clear all notifications for current user."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Method not allowed'}, status=405)
    Notification.objects.filter(recipient=request.user).delete()
    return JsonResponse({'ok': True})


def set_hostname(request):
    """AJAX endpoint to store client hostname in session."""
    if request.method == 'POST':
        import json
        try:
            data = json.loads(request.body)
            hostname = data.get('hostname', '').strip()[:200]
            if hostname:
                request.session['client_hostname'] = hostname
        except Exception:
            pass
    return JsonResponse({'ok': True})


@login_required
def reset_password(request):
    if not request.user.must_change_password:
        return redirect('dashboard:index')
    form = PasswordResetForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        new_pwd = form.cleaned_data['new_password']
        request.user.set_password(new_pwd)
        request.user.password_plain = new_pwd  # visible only to admin
        request.user.must_change_password = False
        request.user.save()
        login(request, request.user)  # re-authenticate after password change
        messages.success(request, 'Password updated successfully. Welcome!')
        return redirect('dashboard:index')
    return render(request, 'accounts/reset_password.html', {'form': form})


# ── Forgot Password ──────────────────────────────────────────────────────────

# In-memory token store: {token: employee_pk} — simple, no extra model needed
_forgot_tokens = {}


def forgot_password(request):
    form = ForgotPasswordForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        username = form.cleaned_data['username'].strip()
        email    = form.cleaned_data['email'].strip()
        try:
            emp = Employee.objects.get(username=username, email=email, is_active=True)
            token = secrets.token_urlsafe(32)
            _forgot_tokens[token] = emp.pk
            reset_url = f"{request.scheme}://{request.get_host()}/accounts/forgot-password/set/{token}/"
            send_mail(
                subject='Password Reset — ERP Department',
                message=f"""Dear {emp.employee_name},

A password reset was requested for your account.

Click the link below to set a new password (valid for this session only):

  {reset_url}

If you did not request this, please ignore this email.

Regards,
ERP Department — Unity Cement""",
                from_email=django_settings.DEFAULT_FROM_EMAIL,
                recipient_list=[emp.email],
                fail_silently=True,
            )
        except Employee.DoesNotExist:
            pass  # Don't reveal if user exists
        messages.success(request, 'If the details match, a reset link has been sent to your email.')
        return redirect('accounts:login')
    return render(request, 'accounts/forgot_password.html', {'form': form})


def forgot_password_set(request, token):
    emp_pk = _forgot_tokens.get(token)
    if not emp_pk:
        messages.error(request, 'This reset link is invalid or has already been used.')
        return redirect('accounts:login')
    emp = get_object_or_404(Employee, pk=emp_pk)
    form = ForgotPasswordSetForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        new_pwd = form.cleaned_data['new_password']
        emp.set_password(new_pwd)
        emp.password_plain = new_pwd
        emp.must_change_password = False
        emp.save()
        _forgot_tokens.pop(token, None)
        AuditLog.log(request, 'password_reset', 'Accounts', f'Password reset via forgot-password: {emp.username}')
        messages.success(request, 'Password updated successfully. Please login with your new password.')
        return redirect('accounts:login')
    return render(request, 'accounts/forgot_password_set.html', {'form': form, 'emp': emp})


@login_required
def employee_list(request):
    if not request.user.perm_accounts_view and not request.user.is_superuser:
        messages.error(request, 'Access Denied. You do not have View permission for Accounts.')
        return redirect('dashboard:index')
    q    = request.GET.get('q', '')
    dept = request.GET.get('dept', '')
    employees = Employee.objects.all()
    if q:
        employees = employees.filter(
            Q(employee_name__icontains=q) | Q(employee_code__icontains=q) | Q(department__icontains=q)
        )
    if dept:
        employees = employees.filter(
            Q(department=dept) | Q(additional_departments__contains=f'|{dept}|')
        )
    from .models import DEPARTMENT_CHOICES
    from django.core.paginator import Paginator
    paginator = Paginator(employees, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    
    # Calculate correct counts from all employees (filtered), not just current page
    all_employees = employees
    active_count = all_employees.filter(is_active=True).count()
    total_count = all_employees.count()
    
    return render(request, 'accounts/employee_list.html', {
        'employees': page_obj, 'page_obj': page_obj,
        'q': q, 'dept': dept,
        'department_choices': DEPARTMENT_CHOICES,
        'can_write':  request.user.perm_accounts_write  or request.user.is_superuser,
        'can_delete': request.user.perm_accounts_delete or request.user.is_superuser,
        'can_export': request.user.perm_accounts_export or request.user.is_superuser,
        'active_count': active_count,
        'total_count': total_count,
    })


def _build_perm_summary(emp):
    """Build a readable permission summary for the employee form display."""
    MODULES = [
        ('Dashboard',    [('View', 'perm_dashboard_view')]),
        ('Accounts',     [('View', 'perm_accounts_view'), ('Write', 'perm_accounts_write'), ('Delete', 'perm_accounts_delete'), ('Export', 'perm_accounts_export')]),
        ('IGP',          [('View', 'perm_igp_view'), ('Write', 'perm_igp_write'), ('Approve', 'perm_igp_approve'), ('Export', 'perm_igp_export')]),
        ('VGP',          [('View', 'perm_vgp_view'), ('Write', 'perm_vgp_write'), ('Approve', 'perm_vgp_approve'), ('Export', 'perm_vgp_export')]),
        ('Help Desk',    [('View', 'perm_helpdesk_view'), ('Raise', 'perm_hd_raise'), ('Manage', 'perm_helpdesk_manage')]),
        ('MGP',          [('View', 'perm_mgp_view'), ('Write', 'perm_mgp_write'), ('Approve', 'perm_mgp_approve'), ('Request', 'perm_mgp_request'), ('Export', 'perm_mgp_export')]),
        ('Reports',      [('IGP', 'perm_reports_igp'), ('VGP', 'perm_reports_vgp'), ('MGP', 'perm_reports_mgp'), ('Audit', 'perm_reports_audit')]),
        ('Grievance',    [('View', 'perm_grv_view'), ('Raise', 'perm_grv_write'), ('Manage', 'perm_grv_manage')]),
    ]
    result = []
    for module, perms in MODULES:
        result.append((module, [(lbl, getattr(emp, field, False)) for lbl, field in perms]))
    return result


@login_required
def employee_create(request):
    if not request.user.perm_accounts_write and not request.user.is_superuser:
        messages.error(request, 'Access Denied. You do not have Write permission for Accounts.')
        return redirect('accounts:employee_list')
    form = EmployeeForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        emp = form.save(commit=False)
        # Auto-apply role template permissions
        try:
            tpl = RolePermissionTemplate.objects.get(role=emp.role)
            tpl.apply_to(emp)
        except RolePermissionTemplate.DoesNotExist:
            pass
        emp.save()
        AuditLog.log(request, 'employee_create', 'Accounts', f'Employee created: {emp.employee_name} ({emp.employee_code})')
        messages.success(request, 'Employee created successfully.')
        return redirect('accounts:employee_list')
    return render(request, 'accounts/employee_form.html', {'form': form, 'title': 'Create Employee', 'perm_summary': []})


@login_required
def employee_edit(request, pk):
    if not request.user.perm_accounts_write and not request.user.is_superuser:
        messages.error(request, 'Access Denied. You do not have Write permission for Accounts.')
        return redirect('accounts:employee_list')
    emp = get_object_or_404(Employee, pk=pk)
    form = EmployeeForm(request.POST or None, instance=emp)
    if request.method == 'POST' and form.is_valid():
        updated = form.save(commit=False)
        # If role changed, re-apply role template
        if 'role' in form.changed_data:
            try:
                tpl = RolePermissionTemplate.objects.get(role=updated.role)
                tpl.apply_to(updated)
            except RolePermissionTemplate.DoesNotExist:
                pass
        updated.save()
        AuditLog.log(request, 'employee_edit', 'Accounts', f'Employee edited: {emp.employee_name} ({emp.employee_code})')
        messages.success(request, 'Employee updated successfully.')
        return redirect('accounts:employee_list')
    return render(request, 'accounts/employee_form.html', {
        'form': form, 'title': 'Edit Employee',
        'perm_summary': _build_perm_summary(emp),
    })


@login_required
def toggle_status(request, pk):
    if not request.user.perm_accounts_write and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('accounts:employee_list')
    emp = get_object_or_404(Employee, pk=pk)
    if emp.pk == request.user.pk:
        messages.error(request, 'You cannot disable your own account.')
        return redirect('accounts:employee_list')
    emp.is_active = not emp.is_active
    emp.save()
    status = 'enabled' if emp.is_active else 'disabled'
    messages.success(request, f'{emp.employee_name} has been {status}.')
    return redirect('accounts:employee_list')


@login_required
def employee_delete(request, pk):
    if not request.user.perm_accounts_delete and not request.user.is_superuser:
        messages.error(request, 'Access Denied. You do not have Delete permission for Accounts.')
        return redirect('accounts:employee_list')
    emp = get_object_or_404(Employee, pk=pk)
    AuditLog.log(request, 'employee_delete', 'Accounts', f'Employee deleted: {emp.employee_name} ({emp.employee_code})')
    emp.delete()
    messages.success(request, 'Employee deleted.')
    return redirect('accounts:employee_list')


@login_required
def import_employees(request):
    if not request.user.perm_accounts_write and not request.user.is_superuser:
        messages.error(request, 'Access Denied. You do not have Write permission for Accounts.')
        return redirect('dashboard:index')
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            wb = openpyxl.load_workbook(request.FILES['excel_file'])
            ws = wb.active
            created, updated = 0, 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                emp_data = {
                    'employee_name': row[0] or '',
                    'employee_code': row[1] or '',
                    'department': row[2] or 'Other',
                    'designation': row[3] or '',
                    'role': row[4] or 'employee',
                    'email': row[5] or '',
                    'username': row[6] or '',
                }
                password = str(row[7]) if row[7] else 'Pass@123'
                emp, created_flag = Employee.objects.get_or_create(
                    employee_code=emp_data['employee_code'],
                    defaults=emp_data
                )
                if created_flag:
                    emp.set_password(password)
                    emp.password_plain = password
                    emp.save()
                    created += 1
                else:
                    for k, v in emp_data.items():
                        setattr(emp, k, v)
                    emp.save()
                    updated += 1
            messages.success(request, f'Import complete: {created} created, {updated} updated.')
        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')
    return redirect('accounts:employee_list')


@login_required
def export_employees(request):
    if not request.user.perm_accounts_export and not request.user.is_superuser:
        messages.error(request, 'Access Denied. You do not have Export permission for Accounts.')
        return redirect('dashboard:index')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'
    headers = ['Employee Name', 'Employee Code', 'Department', 'Designation', 'Role',
               'Email', 'Username', 'Password', 'Joining Date', 'Is Active']
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20
    for emp in Employee.objects.all():
        ws.append([
            emp.employee_name, emp.employee_code, emp.get_departments_display(), emp.designation,
            emp.get_roles_display(), emp.email, emp.username, emp.password_plain,
            str(emp.joining_date) if emp.joining_date else '', 'Yes' if emp.is_active else 'No'
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employees.xlsx"'
    return response


@login_required
def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Import Template'
    headers = ['Employee Name*', 'Employee Code*', 'Department', 'Designation', 'Role', 'Email*', 'Username*', 'Password']
    header_fill = PatternFill(start_color='2d6a4f', end_color='2d6a4f', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 22
    ws.append(['John Doe', 'EMP001', 'HR', 'Manager', 'manager', 'john@example.com', 'johndoe', 'Pass@123'])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="import_template.xlsx"'
    return response


# ── OTP Helper ────────────────────────────────────────────────────────────────

def _send_admin_otp_mail(user, otp_code):
    try:
        send_mail(
            subject='Admin Login OTP — ERP Department',
            message=f"""Dear {user.employee_name},

Your One-Time Password (OTP) for Administrator login is:

  {otp_code}

This OTP is valid for 5 minutes. Do not share it with anyone.

If you did not attempt to login, please contact IT immediately.

Regards,
ERP Department — Unity Cement""",
            from_email=django_settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            fail_silently=True,
        )
    except Exception:
        pass


# ── Admin OTP Verify ──────────────────────────────────────────────────────────

@ensure_csrf_cookie
def admin_otp_verify(request):
    user_id = request.session.get('otp_user_id')
    if not user_id:
        messages.error(request, 'Session expired. Please login again.')
        return redirect('accounts:login')
    try:
        user = Employee.objects.get(pk=user_id)
    except Employee.DoesNotExist:
        return redirect('accounts:login')

    # Mask email: e***@domain.com
    email = user.email
    parts = email.split('@')
    masked_email = parts[0][0] + '*' * (len(parts[0]) - 1) + '@' + parts[1] if len(parts) == 2 else email

    if request.method == 'POST':
        entered = request.POST.get('otp_code', '').strip()
        try:
            otp_session = AdminOTPSession.objects.get(employee=user)
        except AdminOTPSession.DoesNotExist:
            messages.error(request, 'OTP not found. Please login again.')
            return redirect('accounts:login')

        if otp_session.is_expired():
            messages.error(request, 'OTP has expired. Please login again.')
            return redirect('accounts:login')

        if len(entered) != 6 or not entered.isdigit():
            messages.error(request, 'Please enter a valid 6-digit OTP.')
            return render(request, 'accounts/admin_otp_verify.html', {'masked_email': masked_email})

        if otp_session.otp_code != entered:
            messages.error(request, 'Incorrect OTP. Please try again.')
            return render(request, 'accounts/admin_otp_verify.html', {'masked_email': masked_email})

        # OTP correct — log the user in
        otp_session.verified = True
        otp_session.save()
        del request.session['otp_user_id']
        user.backend = 'django.contrib.auth.backends.ModelBackend'
        login(request, user)
        Employee.objects.filter(pk=user.pk).update(session_key=request.session.session_key)
        request.session['last_activity'] = timezone.now().timestamp()
        AuditLog.log(request, 'otp_verified', 'System', f'Admin OTP verified: {user.username}', user=user)
        messages.success(request, f'Welcome back, {user.employee_name}!')
        return redirect('dashboard:index')

    # GET — resend OTP
    if request.GET.get('resend'):
        otp_session = AdminOTPSession.generate_for(user)
        _send_admin_otp_mail(user, otp_session.otp_code)
        messages.success(request, 'A new OTP has been sent to your email.')
        return redirect('accounts:admin_otp_verify')

    return render(request, 'accounts/admin_otp_verify.html', {'masked_email': masked_email})


# ── Settings ──────────────────────────────────────────────────────────────────

@login_required
def system_settings(request):
    if not request.user.has_role('administrator') and not request.user.is_superuser:
        messages.error(request, 'Access Denied. Administrator only.')
        return redirect('dashboard:index')

    setting = SystemSetting.get()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'save_series':
            setting.igp_prefix = request.POST.get('igp_prefix', 'IGP').strip().upper()
            setting.vgp_prefix = request.POST.get('vgp_prefix', 'VGP').strip().upper()
            setting.tkt_prefix = request.POST.get('tkt_prefix', 'TKT').strip().upper()
            setting.mgp_prefix = request.POST.get('mgp_prefix', 'MGP').strip().upper()
            setting.save()
            AuditLog.log(request, 'settings_change', 'Settings', f'Pass prefixes updated: IGP={setting.igp_prefix}, VGP={setting.vgp_prefix}, TKT={setting.tkt_prefix}, MGP={setting.mgp_prefix}')
            messages.success(request, 'Pass series prefixes updated successfully.')

        elif action == 'reset_igp':
            new_num = int(request.POST.get('igp_reset_number', 1))
            setting.igp_next_number = new_num
            setting.save(update_fields=['igp_next_number'])
            AuditLog.log(request, 'settings_change', 'Settings', f'IGP series reset to {setting.igp_prefix}-{new_num:05d}')
            messages.success(request, f'IGP series reset to {setting.igp_prefix}-{new_num:05d}.')

        elif action == 'reset_vgp':
            new_num = int(request.POST.get('vgp_reset_number', 1))
            setting.vgp_next_number = new_num
            setting.save(update_fields=['vgp_next_number'])
            AuditLog.log(request, 'settings_change', 'Settings', f'VGP series reset to {setting.vgp_prefix}-{new_num:05d}')
            messages.success(request, f'VGP series reset to {setting.vgp_prefix}-{new_num:05d}.')

        elif action == 'reset_tkt':
            new_num = int(request.POST.get('tkt_reset_number', 1))
            setting.tkt_next_number = new_num
            setting.save(update_fields=['tkt_next_number'])
            AuditLog.log(request, 'settings_change', 'Settings', f'Ticket series reset to {setting.tkt_prefix}-{new_num:05d}')
            messages.success(request, f'Ticket series reset to {setting.tkt_prefix}-{new_num:05d}.')

        elif action == 'reset_mgp':
            new_num = int(request.POST.get('mgp_reset_number', 1))
            setting.mgp_next_number = new_num
            setting.save(update_fields=['mgp_next_number'])
            AuditLog.log(request, 'settings_change', 'Settings', f'MGP series reset to {setting.mgp_prefix}-{new_num:05d}')
            messages.success(request, f'MGP series reset to {setting.mgp_prefix}-{new_num:05d}.')

        elif action == 'reset_db_hd':
            confirm_text = request.POST.get('confirm_text', '').strip()
            if confirm_text == 'DELETE HD':
                from helpdesk.models import Ticket, TicketComment
                TicketComment.objects.all().delete()
                count = Ticket.objects.count()
                Ticket.objects.all().delete()
                setting.tkt_next_number = 1
                setting.save(update_fields=['tkt_next_number'])
                AuditLog.log(request, 'settings_change', 'Settings', f'HD database reset — {count} tickets deleted by {request.user.username}')
                messages.success(request, f'Help Desk database cleared. {count} tickets deleted and series reset to {setting.tkt_prefix}-00001.')
            else:
                messages.error(request, 'Confirmation text did not match. HD data was NOT deleted.')

        elif action == 'reset_db_mgp':
            confirm_text = request.POST.get('confirm_text', '').strip()
            if confirm_text == 'DELETE MGP':
                from material_pass.models import MaterialGatePass, MaterialItem, MaterialAttachment
                MaterialAttachment.objects.all().delete()
                MaterialItem.objects.all().delete()
                count = MaterialGatePass.objects.count()
                MaterialGatePass.objects.all().delete()
                setting.mgp_next_number = 1
                setting.save(update_fields=['mgp_next_number'])
                AuditLog.log(request, 'settings_change', 'Settings', f'MGP database reset — {count} records deleted by {request.user.username}')
                messages.success(request, f'Material Gate Pass database cleared. {count} records deleted and series reset to {setting.mgp_prefix}-00001.')
            else:
                messages.error(request, 'Confirmation text did not match. MGP data was NOT deleted.')

        elif action == 'reset_db_igp':
            confirm_text = request.POST.get('confirm_text', '').strip()
            if confirm_text == 'DELETE IGP':
                from internal_pass.models import InternalGatePass
                count = InternalGatePass.objects.count()
                InternalGatePass.objects.all().delete()
                setting.igp_next_number = 1
                setting.save(update_fields=['igp_next_number'])
                AuditLog.log(request, 'settings_change', 'Settings', f'IGP database reset — {count} records deleted by {request.user.username}')
                messages.success(request, f'IGP database cleared. {count} records deleted and series reset to 1.')
            else:
                messages.error(request, 'Confirmation text did not match. IGP data was NOT deleted.')

        elif action == 'reset_db_vgp':
            confirm_text = request.POST.get('confirm_text', '').strip()
            if confirm_text == 'DELETE VGP':
                from visitor_pass.models import VisitorGatePass
                count = VisitorGatePass.objects.count()
                VisitorGatePass.objects.all().delete()
                setting.vgp_next_number = 1
                setting.save(update_fields=['vgp_next_number'])
                AuditLog.log(request, 'settings_change', 'Settings', f'VGP database reset — {count} records deleted by {request.user.username}')
                messages.success(request, f'VGP database cleared. {count} records deleted and series reset to 1.')
            else:
                messages.error(request, 'Confirmation text did not match. VGP data was NOT deleted.')

        elif action == 'save_smtp':
            setting.smtp_host     = request.POST.get('smtp_host', '').strip()
            setting.smtp_port     = int(request.POST.get('smtp_port', 587))
            setting.smtp_use_tls  = request.POST.get('smtp_use_tls') == 'on'
            setting.smtp_user     = request.POST.get('smtp_user', '').strip()
            setting.smtp_from     = request.POST.get('smtp_from', '').strip()
            new_pwd = request.POST.get('smtp_password', '').strip()
            if new_pwd:
                setting.smtp_password = new_pwd
            setting.save()
            # Apply to live Django settings
            django_settings.EMAIL_HOST          = setting.smtp_host
            django_settings.EMAIL_PORT          = setting.smtp_port
            django_settings.EMAIL_USE_TLS       = setting.smtp_use_tls
            django_settings.EMAIL_HOST_USER     = setting.smtp_user
            django_settings.EMAIL_HOST_PASSWORD = setting.smtp_password
            django_settings.DEFAULT_FROM_EMAIL  = setting.smtp_from or setting.smtp_user
            AuditLog.log(request, 'settings_change', 'Settings', f'SMTP settings updated: {setting.smtp_host}:{setting.smtp_port}')
            messages.success(request, 'SMTP mail settings saved successfully.')

        elif action == 'test_smtp':
            try:
                send_mail(
                    subject='SMTP Test — ERP Department',
                    message='This is a test email from ERP Department to verify SMTP configuration.',
                    from_email=django_settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[request.user.email],
                    fail_silently=False,
                )
                messages.success(request, f'Test email sent successfully to {request.user.email}.')
            except Exception as e:
                messages.error(request, f'SMTP test failed: {e}')

        elif action == 'save_maintenance':
            setting.maintenance_mode    = request.POST.get('maintenance_mode') == 'on'
            setting.maintenance_message = request.POST.get('maintenance_message', '').strip()
            setting.save()
            status = 'ENABLED' if setting.maintenance_mode else 'DISABLED'
            AuditLog.log(request, 'settings_change', 'Settings', f'Maintenance mode {status}')
            messages.success(request, f'Maintenance mode {status.lower()} successfully.')

        elif action == 'save_session':
            try:
                minutes = int(request.POST.get('session_timeout_minutes', setting.session_timeout_minutes or 20))
            except Exception:
                minutes = 20
            minutes = max(1, min(24 * 60, minutes))
            setting.session_timeout_minutes = minutes
            setting.save(update_fields=['session_timeout_minutes'])
            AuditLog.log(request, 'settings_change', 'Settings', f'Session timeout updated to {minutes} minutes')
            messages.success(request, 'Session timeout updated successfully.')

        elif action == 'save_inauguration':
            setting.welcome_enabled = request.POST.get('welcome_enabled') == 'on'
            setting.welcome_title = (request.POST.get('welcome_title', '') or '').strip()[:120]
            setting.welcome_message_management = (request.POST.get('welcome_message_management', '') or '').strip()
            setting.welcome_message_president = (request.POST.get('welcome_message_president', '') or '').strip()
            bump = request.POST.get('bump_version') == 'on'
            if bump:
                setting.welcome_version = int(setting.welcome_version or 1) + 1
            setting.save()
            AuditLog.log(
                request,
                'settings_change',
                'Settings',
                f'Inauguration page updated: enabled={setting.welcome_enabled}, version={setting.welcome_version}'
            )
            messages.success(request, 'Inauguration page settings saved.')

        elif action == 'save_workflow_recipients':
            import json
            module_key = request.POST.get('module_key', '')
            roles = request.POST.getlist('recipient_roles')
            try:
                current = json.loads(setting.workflow_email_recipients or '{}')
            except Exception:
                current = {}
            current[module_key] = roles
            setting.workflow_email_recipients = json.dumps(current)
            setting.save(update_fields=['workflow_email_recipients'])
            AuditLog.log(request, 'settings_change', 'Settings', f'Workflow email recipients updated for {module_key}: {roles}')
            messages.success(request, f'Email recipients saved for {module_key.upper()}.')

        elif action == 'save_print_format':
            import json
            pf_module = request.POST.get('pf_module', '')
            enabled_fields = request.POST.getlist('pf_fields')
            field_map = {'igp': 'igp_print_fields', 'vgp': 'vgp_print_fields', 'mgp': 'mgp_print_fields'}
            attr = field_map.get(pf_module)
            if attr:
                setting.__dict__[attr] = json.dumps(enabled_fields)
                setting.save(update_fields=[attr])
                AuditLog.log(request, 'settings_change', 'Settings', f'Print format updated for {pf_module.upper()}')
                messages.success(request, f'{pf_module.upper()} print format saved.')

        elif action == 'restore_backup':
            import json as _json
            backup_file = request.FILES.get('backup_file')
            if backup_file:
                try:
                    data = _json.loads(backup_file.read().decode('utf-8'))
                    restored = []
                    if 'settings' in data:
                        s = data['settings']
                        for k, v in s.items():
                            if hasattr(setting, k):
                                setattr(setting, k, v)
                        setting.save()
                        restored.append('Settings')
                    AuditLog.log(request, 'settings_change', 'Settings', f'Backup restored: {restored}')
                    messages.success(request, f'Backup restored successfully: {", ".join(restored)}.')
                except Exception as e:
                    messages.error(request, f'Restore failed: {e}')
            else:
                messages.error(request, 'No backup file provided.')

        elif action == 'save_role_template':
            role = request.POST.get('role')
            if role:
                tpl, _ = RolePermissionTemplate.objects.get_or_create(role=role)
                for f in RolePermissionTemplate.PERM_FIELDS:
                    setattr(tpl, f, request.POST.get(f) == 'on')
                tpl.save()
                
                # Apply the template to all employees with this role
                employees_with_role = Employee.objects.filter(role=role)
                for emp in employees_with_role:
                    tpl.apply_to(emp)
                    emp.save()
                
                AuditLog.log(request, 'settings_change', 'Settings', f'Role template saved for: {role}')
                emp_count = employees_with_role.count()
                if emp_count > 0:
                    messages.success(request, f'Permission template saved for {tpl.get_role_display()}. Applied to {emp_count} employee(s).')
                else:
                    messages.success(request, f'Permission template saved for {tpl.get_role_display()}.')

        elif action == 'save_notification_permissions':
            from accounts.models import NotificationPermission
            
            scope = request.POST.get('notification_scope')
            notification_type = request.POST.get('notification_type')
            
            if scope and notification_type:
                role = request.POST.get('notification_role', '')
                employee_id = request.POST.get('notification_employee')
                department = request.POST.get('notification_department', '')
                
                # Determine the target based on scope
                if scope == 'role' and role:
                    filter_kwargs = {'scope': 'role', 'role': role, 'notification_type': notification_type}
                elif scope == 'employee' and employee_id:
                    employee = get_object_or_404(Employee, pk=employee_id)
                    filter_kwargs = {'scope': 'employee', 'employee': employee, 'notification_type': notification_type}
                elif scope == 'department' and department:
                    filter_kwargs = {'scope': 'department', 'department': department, 'notification_type': notification_type}
                else:
                    messages.error(request, 'Invalid notification permission configuration.')
                    return redirect('accounts:system_settings')
                
                perm, created = NotificationPermission.objects.get_or_create(**filter_kwargs)
                perm.can_receive = request.POST.get('can_receive') == 'on'
                perm.can_view_own = request.POST.get('can_view_own') == 'on'
                perm.can_view_department = request.POST.get('can_view_department') == 'on'
                perm.can_view_all = request.POST.get('can_view_all') == 'on'
                perm.save()
                
                AuditLog.log(request, 'settings_change', 'Settings', f'Notification permission saved: {perm}')
                messages.success(request, 'Notification permission saved.')

        elif action == 'save_workflow_notifications':
            setting.notif_igp_popup = request.POST.get('notif_igp_popup') == 'on'
            setting.notif_igp_email = request.POST.get('notif_igp_email') == 'on'
            setting.notif_vgp_popup = request.POST.get('notif_vgp_popup') == 'on'
            setting.notif_vgp_email = request.POST.get('notif_vgp_email') == 'on'
            setting.notif_mgp_popup = request.POST.get('notif_mgp_popup') == 'on'
            setting.notif_mgp_email = request.POST.get('notif_mgp_email') == 'on'
            setting.notif_hd_popup = request.POST.get('notif_hd_popup') == 'on'
            setting.notif_hd_email = request.POST.get('notif_hd_email') == 'on'
            setting.save(update_fields=[
                'notif_igp_popup', 'notif_igp_email',
                'notif_vgp_popup', 'notif_vgp_email',
                'notif_mgp_popup', 'notif_mgp_email',
                'notif_hd_popup', 'notif_hd_email',
            ])
            AuditLog.log(request, 'settings_change', 'Settings', 'Workflow notification channel settings updated')
            messages.success(request, 'Workflow notification settings saved. Changes apply immediately.')

        elif action == 'save_workflow_settings':
            setting.notif_igp_popup = request.POST.get('notif_igp_popup') == 'on'
            setting.notif_igp_email = request.POST.get('notif_igp_email') == 'on'
            setting.notif_vgp_popup = request.POST.get('notif_vgp_popup') == 'on'
            setting.notif_vgp_email = request.POST.get('notif_vgp_email') == 'on'
            setting.notif_mgp_popup = request.POST.get('notif_mgp_popup') == 'on'
            setting.notif_mgp_email = request.POST.get('notif_mgp_email') == 'on'
            setting.notif_hd_popup = request.POST.get('notif_hd_popup') == 'on'
            setting.notif_hd_email = request.POST.get('notif_hd_email') == 'on'
            setting.skip_management_notifications = request.POST.get('skip_management_notifications') == 'on'
            setting.skip_plant_head_notifications = request.POST.get('skip_plant_head_notifications') == 'on'
            setting.save(update_fields=[
                'notif_igp_popup', 'notif_igp_email',
                'notif_vgp_popup', 'notif_vgp_email',
                'notif_mgp_popup', 'notif_mgp_email',
                'notif_hd_popup', 'notif_hd_email',
                'skip_management_notifications',
                'skip_plant_head_notifications',
            ])
            AuditLog.log(request, 'settings_change', 'Settings', 
                        f'Workflow settings updated: Skip Mgmt={setting.skip_management_notifications}, Skip Head={setting.skip_plant_head_notifications}')
            messages.success(request, 'Workflow settings saved. Changes apply immediately.')

        elif action == 'save_module_access':
            role = request.POST.get('role')
            if role and role not in RoleModuleAccess.EXCLUDED_ROLES:
                obj, _ = RoleModuleAccess.objects.get_or_create(role=role)
                obj.show_igp = request.POST.get('show_igp') == 'on'
                obj.show_vgp = request.POST.get('show_vgp') == 'on'
                obj.show_hd  = request.POST.get('show_hd')  == 'on'
                obj.save()
                AuditLog.log(request, 'settings_change', 'Settings', f'Module access saved for role: {role}')
                messages.success(request, f'Module access updated for {dict(ROLE_CHOICES).get(role, role)}.')

        return redirect('accounts:system_settings')

    from internal_pass.models import InternalGatePass
    from visitor_pass.models import VisitorGatePass
    from helpdesk.models import Ticket
    from material_pass.models import MaterialGatePass
    from accounts.models import NotificationPermission

    # Email log filters
    email_q = request.GET.get('email_q', '').strip()
    email_status = request.GET.get('email_status', '').strip()
    email_channel = request.GET.get('email_channel', '').strip()
    email_date = request.GET.get('email_date', '').strip()
    email_logs_qs = EmailLog.objects.all()
    if email_q:
        email_logs_qs = email_logs_qs.filter(
            Q(recipient__icontains=email_q) |
            Q(subject__icontains=email_q) |
            Q(message__icontains=email_q)
        )
    if email_status:
        email_logs_qs = email_logs_qs.filter(status=email_status)
    if email_channel:
        email_logs_qs = email_logs_qs.filter(channel=email_channel)
    if email_date:
        email_logs_qs = email_logs_qs.filter(created_at__date=email_date)

    # Notification log filters
    notif_q = request.GET.get('notif_q', '').strip()
    notif_read = request.GET.get('notif_read', '').strip()
    notif_module = request.GET.get('notif_module', '').strip()
    notif_date = request.GET.get('notif_date', '').strip()
    notification_logs_qs = Notification.objects.select_related('recipient').all()
    if notif_q:
        notification_logs_qs = notification_logs_qs.filter(
            Q(title__icontains=notif_q) |
            Q(description__icontains=notif_q) |
            Q(notification_type__icontains=notif_q) |
            Q(recipient__employee_name__icontains=notif_q) |
            Q(recipient__username__icontains=notif_q)
        )
    if notif_read == 'read':
        notification_logs_qs = notification_logs_qs.filter(is_read=True)
    elif notif_read == 'unread':
        notification_logs_qs = notification_logs_qs.filter(is_read=False)
    if notif_module:
        notification_logs_qs = notification_logs_qs.filter(related_module__iexact=notif_module)
    if notif_date:
        notification_logs_qs = notification_logs_qs.filter(created_at__date=notif_date)

    # Export Email Log
    export_kind = request.GET.get('export', '').strip()
    if export_kind == 'email_log':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Email Log'
        headers = ['Timestamp', 'Channel', 'Recipient', 'Subject', 'Status', 'Module', 'Reference ID', 'Error']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 24
        for row in email_logs_qs:
            ws.append([
                row.created_at.strftime('%d-%m-%Y %H:%M:%S'),
                row.channel.upper(),
                row.recipient,
                row.subject,
                row.status,
                row.related_module,
                row.related_id,
                row.error,
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Email_Log.xlsx"'
        return response

    # Export Notification Log
    if export_kind == 'notification_log':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Notification Log'
        headers = ['Timestamp', 'Recipient Name', 'Recipient Username', 'Type', 'Module', 'Title', 'Description', 'Read']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 24
        for row in notification_logs_qs:
            ws.append([
                row.created_at.strftime('%d-%m-%Y %H:%M:%S'),
                getattr(row.recipient, 'employee_name', ''),
                getattr(row.recipient, 'username', ''),
                row.notification_type,
                row.related_module,
                row.title,
                row.description,
                'Yes' if row.is_read else 'No',
            ])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Notification_Log.xlsx"'
        return response
    
    import json as _json
    try:
        wf_recipients = _json.loads(setting.workflow_email_recipients or '{}')
    except Exception:
        wf_recipients = {}

    # Print format modules and their fields
    IGP_PRINT_FIELDS = [
        ('pass_number','Pass Number'), ('employee_name','Employee Name'), ('employee_code','Employee Code'),
        ('department','Department'), ('designation','Designation'), ('out_date','Out Date'),
        ('return_date','Return Date'), ('purpose','Purpose'), ('destination','Destination'),
        ('vehicle_number','Vehicle Number'), ('status','Status'), ('approved_by','Approved By'),
    ]
    VGP_PRINT_FIELDS = [
        ('pass_number','Pass Number'), ('visitor_name','Visitor Name'), ('visitor_company','Company'),
        ('visitor_city','City'), ('purpose','Purpose'), ('person_to_meet','Person to Meet'),
        ('visit_date','Visit Date'), ('in_time','In Time'), ('out_time','Out Time'),
        ('access_card_no','Access Card No'), ('vehicle_number','Vehicle Number'), ('status','Status'),
    ]
    MGP_PRINT_FIELDS = [
        ('pass_number','Pass Number'), ('party_name','Party Name'), ('pass_date','Pass Date'),
        ('pass_type','Pass Type'), ('material_description','Material Description'),
        ('quantity','Quantity'), ('vehicle_number','Vehicle Number'), ('driver_name','Driver Name'),
        ('challan_number','Challan Number'), ('status','Status'), ('approved_by','Approved By'),
    ]

    def _pf_fields_with_state(fields_def, saved_json):
        try:
            enabled = _json.loads(saved_json or '[]')
        except Exception:
            enabled = [f[0] for f in fields_def]
        if not enabled:
            enabled = [f[0] for f in fields_def]
        return [(k, lbl, k in enabled) for k, lbl in fields_def]

    print_format_modules = [
        ('igp', 'Internal Gate Pass', IGP_PRINT_FIELDS, _pf_fields_with_state(IGP_PRINT_FIELDS, setting.igp_print_fields)),
        ('vgp', 'Visitor Gate Pass',  VGP_PRINT_FIELDS, _pf_fields_with_state(VGP_PRINT_FIELDS, setting.vgp_print_fields)),
        ('mgp', 'Material Gate Pass', MGP_PRINT_FIELDS, _pf_fields_with_state(MGP_PRINT_FIELDS, setting.mgp_print_fields)),
    ]
    # Flatten for template: (key, label, [(field_key, field_label, is_enabled)])
    print_format_modules_ctx = [(k, lbl, fields) for k, lbl, _, fields in print_format_modules]

    workflow_modules = [
        ('igp',  'Internal Gate Pass', 'bi-person-walking',    '#2A2A86'),
        ('vgp',  'Visitor Gate Pass',  'bi-person-badge-fill', '#27ae60'),
        ('mgp',  'Material Gate Pass', 'bi-box-seam',          '#c0392b'),
        ('hd',   'IT Help Desk',       'bi-headset',           '#f0a500'),
        ('grv',  'Grievance',          'bi-megaphone-fill',    '#7b4bb3'),
    ]

    # Workflow display data for System Settings
    igp_workflows = [
        ('Workflow 1', 'Employee',
         [('Creator', 'Employee', '\U0001f464'), ('Dept HOD', 'department_hod', '\U0001f4cb'), ('HR', 'hr', '\U0001f465'), ('Security', 'security', '\U0001f6e1')]),
        ('Workflow 2', 'Department HOD / HR',
         [('Creator', 'HOD/HR', '\U0001f464'), ('President/Plant Head', 'president_plant_head', '\U0001f3e2'), ('HR', 'hr', '\U0001f465'), ('Security', 'security', '\U0001f6e1')]),
        ('Workflow 3', 'President / Plant Head',
         [('Creator', 'President', '\U0001f464'), ('Management', 'management', '\U0001f4bc'), ('Security', 'security', '\U0001f6e1')]),
    ]
    vgp_workflow = [
        ('Creator', 'Any Role', '\U0001f464'),
        ('Contact Person', 'person_to_meet', '\U0001f91d'),
        ('Security', 'security', '\U0001f6e1'),
    ]
    mgp_workflows = [
        ('MGP Request', 'Any Role → Store HOD Approval',
         [('Creator', 'Any Role', '\U0001f464'), ('Store HOD', 'department_hod (Store)', '\U0001f3ea')]),
        ('MGP Creation', 'Acknowledgement Email Only',
         [('Creator', 'Any Role', '\U0001f464'), ('Acknowledgement', 'Creator Email', '\U0001f4e7')]),
    ]
    hd_workflow = [
        ('Creator', 'Any Role', '\U0001f464'),
        ('IT Department', 'IT Staff', '\U0001f4bb'),
        ('Resolve ACK', 'Auto Email', '\U0001f4e7'),
    ]
    grv_workflow = [
        ('Creator', 'Any Employee', '\U0001f464'),
        ('Management', 'management', '\U0001f4bc'),
        ('President/HOD/HR', 'Notified', '\U0001f465'),
    ]

    templates = {t.role: t for t in RolePermissionTemplate.objects.all()}
    module_access_roles = [(v, l) for v, l in ROLE_CHOICES if v not in RoleModuleAccess.EXCLUDED_ROLES]
    module_access = {obj.role: obj for obj in RoleModuleAccess.objects.all()}
    notification_permissions = NotificationPermission.objects.all().order_by('scope', 'notification_type')
    
    from accounts.models import NotificationWorkflow
    workflows = NotificationWorkflow.objects.prefetch_related('stages').order_by('pass_type', 'creator_role')
    
    context = {
        'setting': setting,
        'igp_last': InternalGatePass.objects.order_by('-id').first(),
        'vgp_last': VisitorGatePass.objects.order_by('-id').first(),
        'igp_total': InternalGatePass.objects.count(),
        'vgp_total': VisitorGatePass.objects.count(),
        'tkt_last': Ticket.objects.order_by('-id').first(),
        'tkt_total': Ticket.objects.count(),
        'mgp_last': MaterialGatePass.objects.order_by('-id').first(),
        'mgp_total': MaterialGatePass.objects.count(),
        'role_choices': ROLE_CHOICES,
        'templates': templates,
        'perm_fields': RolePermissionTemplate.PERM_FIELDS,
        'all_employees': Employee.objects.filter(is_active=True).order_by('role', 'employee_name'),
        'module_access_roles': module_access_roles,
        'module_access': module_access,
        'notification_permissions': notification_permissions,
        'notification_types': NotificationPermission.NOTIFICATION_TYPES,
        'notification_scopes': NotificationPermission.SCOPE_CHOICES,
        'department_choices': DEPARTMENT_CHOICES,
        'email_logs': email_logs_qs[:200],
        'notification_logs': notification_logs_qs[:200],
        'email_q': email_q,
        'email_status': email_status,
        'email_channel': email_channel,
        'email_date': email_date,
        'notif_q': notif_q,
        'notif_read': notif_read,
        'notif_module': notif_module,
        'notif_date': notif_date,
        'workflows': workflows,
        'workflow_modules': workflow_modules,
        'workflow_recipients': wf_recipients,
        'print_format_modules': print_format_modules_ctx,
        'igp_workflows': igp_workflows,
        'vgp_workflow': vgp_workflow,
        'mgp_workflows': mgp_workflows,
        'hd_workflow': hd_workflow,
        'grv_workflow': grv_workflow,
    }

    # List auto-backups
    import os as _os
    backup_dir = _os.path.join(django_settings.BASE_DIR, 'backups')
    auto_backups = []
    if _os.path.exists(backup_dir):
        for fn in sorted(_os.listdir(backup_dir), reverse=True):
            if fn.endswith('.json'):
                fp = _os.path.join(backup_dir, fn)
                size_kb = _os.path.getsize(fp) // 1024
                auto_backups.append({'name': fn, 'size_kb': size_kb})
    context['auto_backups'] = auto_backups[:30]

    return render(request, 'accounts/system_settings.html', context)


@login_required
def backup_download(request):
    if not request.user.has_role('administrator') and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')
    import json as _json
    from django.core import serializers as dj_serializers
    backup_type = request.GET.get('type', 'full')

    # Download a specific auto-backup file
    filename = request.GET.get('file', '')
    if filename:
        backup_dir = os.path.join(django_settings.BASE_DIR, 'backups')
        filepath = os.path.join(backup_dir, filename)
        # Security: only allow files in backups dir
        if os.path.exists(filepath) and os.path.dirname(os.path.abspath(filepath)) == os.path.abspath(backup_dir):
            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/json')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        messages.error(request, 'Backup file not found.')
        return redirect('accounts:system_settings')

    data = {}
    if backup_type in ('full', 'settings'):
        setting = SystemSetting.get()
        data['settings'] = {
            f.name: getattr(setting, f.name)
            for f in setting._meta.fields
            if f.name not in ('id',)
        }
        for k, v in data['settings'].items():
            if hasattr(v, 'isoformat'):
                data['settings'][k] = v.isoformat()
    if backup_type in ('full', 'employees'):
        data['employees'] = _json.loads(
            dj_serializers.serialize('json', Employee.objects.all())
        )
    data['_meta'] = {'backup_type': backup_type, 'created_at': timezone.now().isoformat()}
    filename_out = f'erp_backup_{backup_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.json'
    response = HttpResponse(
        _json.dumps(data, indent=2, default=str),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_out}"'
    AuditLog.log(request, 'settings_change', 'Settings', f'Backup downloaded: {backup_type}')
    return response


@login_required
def inauguration_page(request):
    setting = SystemSetting.get()
    if not setting.welcome_enabled:
        return redirect('dashboard:index')

    if request.method == 'POST':
        cur = int(setting.welcome_version or 1)
        Employee.objects.filter(pk=request.user.pk).update(welcome_seen_version=cur)
        AuditLog.log(request, 'page_view', 'System', f'Inauguration accepted (v{cur})')
        return redirect('dashboard:index')

    return render(request, 'accounts/inauguration.html', {'setting': setting})


# ── User Rights Matrix ─────────────────────────────────────────────────────────

@login_required
def user_rights_matrix(request):
    if not request.user.has_role('administrator') and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')

    employees = Employee.objects.filter(is_active=True).order_by('role', 'employee_name')

    PERM_COLS = [
        ('Dashboard', [
            ('perm_dashboard_view', 'View'),
        ]),
        ('Accounts', [
            ('perm_accounts_view',   'View'),
            ('perm_accounts_write',  'Write'),
            ('perm_accounts_delete', 'Delete'),
            ('perm_accounts_export', 'Export'),
        ]),
        ('Help Desk', [
            ('perm_helpdesk_view',   'View'),
            ('perm_helpdesk_write',  'Raise'),
            ('perm_helpdesk_manage', 'Manage'),
        ]),
        ('IGP', [
            ('perm_igp_view',    'View'),
            ('perm_igp_write',   'Write'),
            ('perm_igp_delete',  'Delete'),
            ('perm_igp_approve', 'Approve'),
            ('perm_igp_bypass',  'Bypass'),
            ('perm_igp_export',  'Export'),
        ]),
        ('VGP', [
            ('perm_vgp_view',    'View'),
            ('perm_vgp_write',   'Write'),
            ('perm_vgp_delete',  'Delete'),
            ('perm_vgp_approve', 'Approve'),
            ('perm_vgp_bypass',  'Bypass'),
            ('perm_vgp_export',  'Export'),
        ]),
        ('MGP', [
            ('perm_mgp_view',    'View'),
            ('perm_mgp_write',   'Write'),
            ('perm_mgp_delete',  'Delete'),
            ('perm_mgp_approve', 'Approve'),
            ('perm_mgp_export',  'Export'),
        ]),
        ('Reports', [
            ('perm_reports_igp',   'IGP'),
            ('perm_reports_vgp',   'VGP'),
            ('perm_reports_mgp',   'MGP'),
            ('perm_reports_audit', 'Audit'),
        ]),
    ]
    flat_cols = [(mod, field, label) for mod, perms in PERM_COLS for field, label in perms]

    # ── Excel Export ──
    if request.GET.get('export') == 'excel':
        import openpyxl as xl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = xl.Workbook()
        ws = wb.active
        ws.title = 'User Rights Matrix'

        navy   = PatternFill('solid', fgColor='1E3A5F')
        blue   = PatternFill('solid', fgColor='2D5A8E')
        green  = PatternFill('solid', fgColor='0E6655')
        purple = PatternFill('solid', fgColor='6C3483')
        alt    = PatternFill('solid', fgColor='F8F9FA')
        white_font = Font(bold=True, color='FFFFFF', size=9)
        thin = Border(
            left=Side(style='thin', color='DEE2E6'),
            right=Side(style='thin', color='DEE2E6'),
            top=Side(style='thin', color='DEE2E6'),
            bottom=Side(style='thin', color='DEE2E6'),
        )

        # Row 1: module group headers
        ws.cell(1, 1, 'Employee').font = white_font
        ws.cell(1, 1).fill = navy
        ws.cell(1, 2, 'Role').font = white_font
        ws.cell(1, 2).fill = navy
        ws.cell(1, 3, 'Department').font = white_font
        ws.cell(1, 3).fill = navy

        col = 4
        mod_colors = {'Dashboard': navy, 'Accounts': blue, 'IGP': green, 'VGP': purple, 'MGP': PatternFill('solid', fgColor='B7410E'), 'Reports': PatternFill('solid', fgColor='1A6B3C')}
        for mod, perms in PERM_COLS:
            start = col
            for _ in perms:
                c = ws.cell(1, col, mod if col == start else '')
                c.fill = mod_colors[mod]
                c.font = white_font
                c.alignment = Alignment(horizontal='center')
                col += 1
            if len(perms) > 1:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=col - 1)

        # Row 2: permission labels
        for c in range(1, 4):
            ws.cell(2, c).fill = navy
            ws.cell(2, c).font = white_font
        col = 4
        for mod, perms in PERM_COLS:
            for _, label in perms:
                c = ws.cell(2, col, label)
                c.fill = mod_colors[mod]
                c.font = white_font
                c.alignment = Alignment(horizontal='center')
                col += 1

        # Data rows
        for i, emp in enumerate(employees, 3):
            fill = alt if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
            ws.cell(i, 1, emp.employee_name).fill = fill
            ws.cell(i, 2, emp.get_role_display()).fill = fill
            ws.cell(i, 3, emp.department).fill = fill
            col = 4
            for _, field, _ in flat_cols:
                val = '✓' if getattr(emp, field, False) else '✗'
                c = ws.cell(i, col, val)
                c.fill = fill
                c.alignment = Alignment(horizontal='center')
                c.font = Font(color='27AE60' if val == '✓' else 'C0392B', bold=(val == '✓'))
                col += 1

        # Column widths
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        for i in range(4, col):
            ws.column_dimensions[get_column_letter(i)].width = 9

        # Apply borders
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=col - 1):
            for cell in row:
                cell.border = thin

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="UserRightsMatrix.xlsx"'
        return resp

    # ── PDF Export ──
    if request.GET.get('export') == 'pdf':
        from io import BytesIO as _BIO
        from reportlab.lib.pagesizes import A3, landscape
        from reportlab.lib import colors as _c
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        buf = _BIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                                topMargin=10*mm, bottomMargin=8*mm,
                                leftMargin=8*mm, rightMargin=8*mm)
        navy_c  = _c.HexColor('#1e3a5f')
        blue_c  = _c.HexColor('#2d5a8e')
        green_c = _c.HexColor('#0e6655')
        purp_c  = _c.HexColor('#6c3483')
        alt_c   = _c.HexColor('#f8f9fa')

        ts = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=13,
                            alignment=TA_CENTER, textColor=navy_c, spaceAfter=2)
        ss = ParagraphStyle('s', fontName='Helvetica', fontSize=8,
                            alignment=TA_CENTER, textColor=_c.grey, spaceAfter=4)

        elems = []
        elems.append(Paragraph('USER RIGHTS MATRIX REPORT', ts))
        elems.append(Paragraph(f'Generated: {timezone.now().strftime("%d %b %Y %H:%M")}  |  Total Employees: {employees.count()}', ss))
        elems.append(HRFlowable(width='100%', thickness=2, color=navy_c, spaceAfter=4*mm))

        # Header row 1 — module groups
        h1 = ['Employee', 'Role', 'Dept']
        mod_spans = []
        for mod, perms in PERM_COLS:
            h1.append(mod)
            for _ in perms[1:]:
                h1.append('')
            mod_spans.append((mod, len(perms)))

        # Header row 2 — permission labels
        h2 = ['', '', '']
        for _, perms in PERM_COLS:
            for _, label in perms:
                h2.append(label)

        data = [h1, h2]
        for emp in employees:
            row = [emp.employee_name, emp.get_role_display(), emp.department]
            for _, field, _ in flat_cols:
                row.append('✓' if getattr(emp, field, False) else '✗')
            data.append(row)

        total_cols = 3 + len(flat_cols)
        col_w = [38*mm, 28*mm, 32*mm] + [10*mm] * len(flat_cols)

        tbl = Table(data, colWidths=col_w, repeatRows=2)

        style_cmds = [
            ('FONTNAME',    (0,0), (-1,1),  'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,1),  7),
            ('FONTNAME',    (0,2), (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,2), (-1,-1), 7),
            ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
            ('ALIGN',       (0,0), (2,-1),  'LEFT'),
            ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING',     (0,0), (-1,-1), 3),
            ('GRID',        (0,0), (-1,-1), 0.3, _c.HexColor('#dee2e6')),
            ('ROWBACKGROUNDS', (0,2), (-1,-1), [_c.white, alt_c]),
            # Header row 1 backgrounds
            ('BACKGROUND',  (0,0), (2,1),   navy_c),
            ('TEXTCOLOR',   (0,0), (-1,1),  _c.white),
        ]

        # Module group colors in header row 1
        col_idx = 3
        mod_color_map = {
            'Dashboard': navy_c, 'Accounts': blue_c, 'IGP': green_c, 'VGP': purp_c,
            'MGP': _c.HexColor('#b7410e'),
            'Reports': _c.HexColor('#1a6b3c'),
        }
        for mod, perms in PERM_COLS:
            end = col_idx + len(perms) - 1
            style_cmds.append(('BACKGROUND', (col_idx, 0), (end, 1), mod_color_map[mod]))
            if len(perms) > 1:
                style_cmds.append(('SPAN', (col_idx, 0), (end, 0)))
            col_idx += len(perms)

        # Color ✓ green, ✗ red in data rows
        for ri, emp in enumerate(employees, 2):
            ci = 3
            for _, field, _ in flat_cols:
                val = getattr(emp, field, False)
                color = _c.HexColor('#27ae60') if val else _c.HexColor('#c0392b')
                style_cmds.append(('TEXTCOLOR', (ci, ri), (ci, ri), color))
                if val:
                    style_cmds.append(('FONTNAME', (ci, ri), (ci, ri), 'Helvetica-Bold'))
                ci += 1

        tbl.setStyle(TableStyle(style_cmds))
        elems.append(tbl)
        elems.append(Spacer(1, 4*mm))
        elems.append(Paragraph('✓ = Granted  |  ✗ = Not Granted  |  Bypass = Temporary Admin Override', ss))
        elems.append(Paragraph('This is a computer generated report.', ss))

        doc.build(elems)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/pdf')
        resp['Content-Disposition'] = 'attachment; filename="UserRightsMatrix.pdf"'
        return resp

    return render(request, 'accounts/user_rights_matrix.html', {
        'employees': employees,
        'PERM_COLS': PERM_COLS,
        'flat_cols': flat_cols,
    })


# ── Audit Log ─────────────────────────────────────────────────────────────────

@login_required
def audit_log(request):
    if not request.user.has_role('administrator') and not request.user.is_superuser:
        messages.error(request, 'Access Denied. Administrator only.')
        return redirect('dashboard:index')

    logs = AuditLog.objects.select_related('user').all()

    # Filters
    f_user     = request.GET.get('user', '')
    f_action   = request.GET.get('action', '')
    f_module   = request.GET.get('module', '')
    f_ip       = request.GET.get('ip', '')
    f_date     = request.GET.get('date', '')
    f_dept     = request.GET.get('dept', '')

    if f_user:   logs = logs.filter(username__icontains=f_user)
    if f_action: logs = logs.filter(action=f_action)
    if f_module: logs = logs.filter(module=f_module)
    if f_ip:     logs = logs.filter(ip_address__icontains=f_ip)
    if f_date:   logs = logs.filter(timestamp__date=f_date)
    if f_dept:   logs = logs.filter(user__department=f_dept)

    # Export Excel
    if request.GET.get('export') == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Audit Log'
        headers = ['#', 'Timestamp', 'Username', 'IP Address', 'Action', 'Module', 'Description']
        hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 22
        for i, log in enumerate(logs, 1):
            ws.append([i, log.timestamp.strftime('%d-%m-%Y %H:%M:%S'),
                       log.username, log.ip_address or '', log.action,
                       log.module, log.description])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="AuditLog_{f_date or "all"}.xlsx"'
        return resp

    # Export PDF
    if request.GET.get('export') == 'pdf':
        from io import BytesIO as _BytesIO
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors as _colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        buf = _BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                topMargin=12*mm, bottomMargin=10*mm,
                                leftMargin=10*mm, rightMargin=10*mm)
        navy = _colors.HexColor('#1e3a5f')
        alt  = _colors.HexColor('#f8f9fa')
        ts   = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=13,
                              alignment=TA_CENTER, textColor=navy, spaceAfter=2)
        ss   = ParagraphStyle('s', fontName='Helvetica', fontSize=8,
                              alignment=TA_CENTER, textColor=_colors.grey, spaceAfter=4)
        elems = []
        elems.append(Paragraph('SYSTEM AUDIT LOG REPORT', ts))
        elems.append(Paragraph(f'Generated on {timezone.now().strftime("%d %b %Y %H:%M")}  |  Total: {logs.count()} records', ss))
        elems.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=4*mm))
        headers = ['#', 'Timestamp', 'Username', 'IP Address', 'Action', 'Module', 'Description']
        col_w   = [10*mm, 38*mm, 30*mm, 28*mm, 30*mm, 20*mm, 110*mm]
        data = [headers]
        for i, log in enumerate(logs, 1):
            data.append([str(i), log.timestamp.strftime('%d-%m-%Y %H:%M:%S'),
                         log.username, log.ip_address or '-', log.action,
                         log.module, log.description])
        if len(data) == 1:
            data.append(['', 'No records found.', '', '', '', '', ''])
        tbl = Table(data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('FONTNAME',    (0,0),(-1,0),  'Helvetica-Bold'),
            ('FONTSIZE',    (0,0),(-1,0),  8),
            ('BACKGROUND',  (0,0),(-1,0),  navy),
            ('TEXTCOLOR',   (0,0),(-1,0),  _colors.white),
            ('FONTNAME',    (0,1),(-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,1),(-1,-1), 7),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[_colors.white, alt]),
            ('GRID',        (0,0),(-1,-1), 0.3, _colors.HexColor('#dee2e6')),
            ('VALIGN',      (0,0),(-1,-1), 'MIDDLE'),
            ('PADDING',     (0,0),(-1,-1), 3),
            ('WORDWRAP',    (0,0),(-1,-1), True),
        ]))
        elems.append(tbl)
        elems.append(Spacer(1, 4*mm))
        elems.append(Paragraph('This is a computer generated report.', ss))
        doc.build(elems)
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="AuditLog_{f_date or "all"}.pdf"'
        return resp

    from accounts.models import ACTION_CHOICES, DEPARTMENT_CHOICES
    from django.core.paginator import Paginator
    paginator = Paginator(logs, 1000)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    context = {
        'logs': page_obj, 'page_obj': page_obj,
        'total': logs.count(),
        'action_choices': ACTION_CHOICES,
        'department_choices': DEPARTMENT_CHOICES,
        'modules': AuditLog.objects.values_list('module', flat=True).distinct(),
        'f_user': f_user, 'f_action': f_action, 'f_module': f_module,
        'f_ip': f_ip, 'f_date': f_date, 'f_dept': f_dept,
    }
    return render(request, 'accounts/audit_log.html', context)


@login_required
def bulk_action(request):
    if request.method != 'POST':
        return redirect('accounts:employee_list')
    if not request.user.has_role('administrator') and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('accounts:employee_list')
    action = request.POST.get('action')
    ids    = request.POST.getlist('selected_ids')
    if not ids:
        messages.warning(request, 'No employees selected.')
        return redirect('accounts:employee_list')
    employees = Employee.objects.filter(pk__in=ids).exclude(pk=request.user.pk)
    if action == 'delete':
        count = employees.count()
        employees.delete()
        AuditLog.log(request, 'employee_bulk_delete', 'Employee', f'{count} Employees deleted in bulk by {request.user.username}')
        messages.success(request, f'{count} employee(s) deleted.')
    elif action == 'duplicate':
        for emp in employees:
            emp.pk = None
            emp.username = emp.username + '_copy'
            emp.email = None
            emp.save()
        AuditLog.log(request, 'employee_bulk_duplicate', 'Employee', f'{len(ids)} Employees duplicated in bulk by {request.user.username}')
        messages.success(request, f'{len(ids)} employee(s) duplicated.')
    else:
        messages.error(request, 'Unknown action.')
    return redirect('accounts:employee_list')
