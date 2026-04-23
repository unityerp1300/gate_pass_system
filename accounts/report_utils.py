"""
Shared report utilities for Excel and PDF exports.
- Excel: Unity logo in top-right, approval timestamp, computer-generated note
- PDF: Unity logo small in top-right corner
"""
import os
from io import BytesIO
from django.conf import settings as django_settings
from django.utils import timezone


LOGO_PATH = os.path.join(django_settings.BASE_DIR, 'static', 'images', 'Unity_Logo_Horizontal.png')
GENERATED_NOTE = 'This is a computer generated report. — Unity Cement ERP System'


def add_excel_logo_and_note(ws, total_cols):
    """
    Add Unity logo to top-right of Excel sheet and a note at the bottom.
    Call AFTER writing all data rows.
    total_cols: number of data columns (for note placement).
    """
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, Alignment, PatternFill
        if os.path.exists(LOGO_PATH):
            img = XLImage(LOGO_PATH)
            img.width = 120
            img.height = 32
            # Place logo at column total_cols (right side), row 1
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(max(total_cols, 1))
            ws.add_image(img, col_letter + '1')
    except Exception:
        pass

    try:
        from openpyxl.styles import Font, Alignment
        # Add note 2 rows after last data row
        last_row = ws.max_row + 2
        note_cell = ws.cell(row=last_row, column=1, value=GENERATED_NOTE)
        note_cell.font = Font(italic=True, color='718096', size=9)
        note_cell.alignment = Alignment(horizontal='left')
        # Generated timestamp
        ts_cell = ws.cell(row=last_row + 1, column=1,
                          value=f'Generated: {timezone.now().strftime("%d %b %Y %H:%M:%S")}')
        ts_cell.font = Font(italic=True, color='718096', size=9)
    except Exception:
        pass


def build_pdf_header_table(title, subtitle, page_width_mm):
    """
    Build a PDF header table with logo on the RIGHT and title on the left.
    Returns a reportlab Table element.
    """
    from reportlab.platypus import Table, TableStyle, Paragraph, Image, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT
    from reportlab.lib import colors
    from reportlab.lib.units import mm

    navy = colors.HexColor('#1e3a5f')

    title_sty = ParagraphStyle('ht', fontName='Helvetica-Bold', fontSize=13,
                               textColor=navy, alignment=TA_LEFT, leading=16)
    sub_sty = ParagraphStyle('hs', fontName='Helvetica', fontSize=8,
                             textColor=colors.grey, alignment=TA_LEFT, leading=11)

    title_cell = [Paragraph(title, title_sty), Spacer(1, 1*mm), Paragraph(subtitle, sub_sty)]

    LOGO_W = 36 * mm
    LOGO_H = 10 * mm
    if os.path.exists(LOGO_PATH):
        logo_cell = Image(LOGO_PATH, width=LOGO_W, height=LOGO_H)
    else:
        logo_cell = Paragraph('Unity Cement', ParagraphStyle('lc', fontName='Helvetica-Bold',
                              fontSize=10, textColor=navy, alignment=TA_RIGHT))

    tbl = Table([[title_cell, logo_cell]],
                colWidths=[page_width_mm * mm - LOGO_W - 4*mm, LOGO_W])
    tbl.setStyle(TableStyle([
        ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN',         (1, 0), (1, 0),   'RIGHT'),
        ('LEFTPADDING',   (0, 0), (-1, -1), 0),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 0),
        ('TOPPADDING',    (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    return tbl
