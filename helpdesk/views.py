from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from django.utils import timezone
from .models import Ticket, TicketComment
from .forms import TicketForm, TicketUpdateForm, CommentForm
from accounts.notification_service import send_workflow_notification


def _is_it(user):
    return user.perm_helpdesk_manage or user.has_role('administrator') or user.is_superuser or user.has_department('IT')


def _send_hd_it_notification(ticket):
    """Send extraordinary HTML email to IT department when a ticket is raised."""
    try:
        from accounts.models import Employee
        from django.conf import settings
        from django.core.mail import EmailMultiAlternatives
        it_staff = Employee.objects.filter(
            is_active=True, email__gt=''
        ).filter(
            Q(department='IT') | Q(additional_departments__contains='|IT|') |
            Q(perm_helpdesk_manage=True)
        ).distinct()
        if not it_staff.exists():
            return

        priority_colors = {
            'critical': ('#fee2e2', '#991b1b', '#dc2626'),
            'high':     ('#fef3c7', '#92400e', '#d97706'),
            'medium':   ('#dbeafe', '#1e40af', '#3b82f6'),
            'low':      ('#f0fdf4', '#166534', '#22c55e'),
        }
        p_bg, p_fg, p_border = priority_colors.get(ticket.priority, ('#f1f5f9', '#475569', '#94a3b8'))

        subject = f'[New Ticket] {ticket.ticket_number} — {ticket.title}'
        html_body = (
            '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"></head>'
            '<body style="margin:0;padding:0;background:#f0f4f8;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Inter,Arial,sans-serif;">'
            '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:32px 0;"><tr><td align="center">'
            '<table width="620" cellpadding="0" cellspacing="0" style="background:white;border-radius:20px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,0.12);">'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:4px;"></td></tr>'
            '<tr><td style="background:linear-gradient(135deg,#0f172a,#1e3a5f,#7c3aed);padding:36px 40px;text-align:center;">'
            '<div style="font-size:11px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;">Unity Cement — IT Help Desk</div>'
            '<div style="width:64px;height:64px;border-radius:50%;background:rgba(255,255,255,0.12);border:2px solid rgba(255,255,255,0.25);display:inline-flex;align-items:center;justify-content:center;margin-bottom:16px;font-size:28px;">&#127381;</div>'
            '<div style="font-size:26px;font-weight:800;color:white;margin-bottom:6px;">New Support Ticket</div>'
            '<div style="font-size:14px;color:rgba(255,255,255,0.7);margin-bottom:16px;">Requires IT Department Attention</div>'
            f'<div style="display:inline-block;background:rgba(240,165,0,0.2);border:1.5px solid rgba(240,165,0,0.5);border-radius:50px;padding:6px 20px;font-size:14px;color:#fbbf24;font-weight:700;">{ticket.ticket_number}</div>'
            '</td></tr>'
            '<tr><td style="padding:28px 40px 0;">'
            '<p style="font-size:16px;color:#1e293b;margin:0 0 8px;">Dear <strong style="color:#1e3a5f;">IT Team</strong>,</p>'
            '<p style="font-size:14px;color:#64748b;margin:0;line-height:1.6;">A new support ticket has been raised and requires your attention. Please review and respond promptly.</p>'
            '</td></tr>'
            '<tr><td style="padding:20px 40px;">'
            '<div style="border-radius:12px;overflow:hidden;border:1px solid #e2e8f0;">'
            '<div style="background:linear-gradient(135deg,#1e3a5f,#7c3aed);padding:12px 16px;"><span style="font-size:11px;font-weight:700;color:white;letter-spacing:1.5px;text-transform:uppercase;">&#127381; Ticket Details</span></div>'
            '<table width="100%" cellpadding="0" cellspacing="0">'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;width:38%;border-bottom:1px solid #f1f5f9;">Ticket No.</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;"><strong>{ticket.ticket_number}</strong></td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Title</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{ticket.title}</td></tr>'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Raised By</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{ticket.raised_by.employee_name} ({ticket.raised_by.department})</td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Category</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{ticket.get_category_display()}</td></tr>'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Priority</td><td style="padding:9px 16px;font-size:13px;border-bottom:1px solid #f1f5f9;"><span style="background:{p_bg};color:{p_fg};border:1px solid {p_border};padding:3px 12px;border-radius:20px;font-size:12px;font-weight:700;">{ticket.get_priority_display()}</span></td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;">Description</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;">{ticket.description}</td></tr>'
            '</table></div></td></tr>'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:3px;"></td></tr>'
            '<tr><td style="background:#f8fafc;padding:20px 40px;text-align:center;">'
            '<p style="font-size:11px;color:#94a3b8;margin:0;">This is an automated email from <strong>ERP Department — Unity Cement (PMC Cement Ltd.)</strong></p>'
            '<p style="font-size:11px;color:#94a3b8;margin:6px 0 0;">Do not reply to this email • For support contact IT/ERP Department</p>'
            '</td></tr>'
            '</table></td></tr></table></body></html>'
        )
        plain = f"""New Ticket: {ticket.ticket_number}
Title: {ticket.title}
Raised By: {ticket.raised_by.employee_name} ({ticket.raised_by.department})
Category: {ticket.get_category_display()}
Priority: {ticket.get_priority_display()}
Description: {ticket.description}

Regards,
ERP Department — Unity Cement"""

        for staff in it_staff:
            msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [staff.email])
            msg.attach_alternative(html_body, 'text/html')
            msg.send(fail_silently=True)
    except Exception:
        pass


def _send_hd_resolve_ack(ticket):
    """Send extraordinary acknowledgement email to ticket raiser when resolved."""
    try:
        if not ticket.raised_by.email:
            return
        from django.conf import settings
        from django.core.mail import EmailMultiAlternatives
        subject = f'Ticket Resolved — {ticket.ticket_number}'
        html_body = (
            '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"></head>'
            '<body style="margin:0;padding:0;background:#f0f4f8;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Inter,Arial,sans-serif;">'
            '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:32px 0;"><tr><td align="center">'
            '<table width="620" cellpadding="0" cellspacing="0" style="background:white;border-radius:20px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,0.12);">'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:4px;"></td></tr>'
            '<tr><td style="background:linear-gradient(135deg,#064e3b,#059669);padding:36px 40px;text-align:center;">'
            '<div style="font-size:11px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;">Unity Cement — IT Help Desk</div>'
            '<div style="width:72px;height:72px;border-radius:50%;background:rgba(255,255,255,0.15);border:2px solid rgba(255,255,255,0.3);display:inline-flex;align-items:center;justify-content:center;margin-bottom:16px;font-size:32px;">&#10003;</div>'
            '<div style="font-size:26px;font-weight:800;color:white;margin-bottom:8px;">Ticket Resolved</div>'
            '<div style="display:inline-block;background:#d1fae5;color:#065f46;border-radius:50px;padding:6px 20px;font-size:13px;font-weight:700;letter-spacing:1px;">RESOLVED</div>'
            f'<div style="margin-top:14px;display:inline-block;background:rgba(255,255,255,0.12);border:1.5px solid rgba(255,255,255,0.25);border-radius:50px;padding:5px 18px;font-size:13px;color:rgba(255,255,255,0.9);font-weight:600;">{ticket.ticket_number}</div>'
            '</td></tr>'
            f'<tr><td style="padding:28px 40px 0;"><p style="font-size:16px;color:#1e293b;margin:0 0 8px;">Dear <strong style="color:#1e3a5f;">{ticket.raised_by.employee_name}</strong>,</p>'
            '<p style="font-size:14px;color:#64748b;margin:0;line-height:1.6;">Your IT support ticket has been <strong>resolved</strong> by the IT Department. We hope your issue has been addressed satisfactorily.</p>'
            '</td></tr>'
            '<tr><td style="padding:20px 40px;">'
            '<div style="border-radius:12px;overflow:hidden;border:1px solid #e2e8f0;">'
            '<div style="background:linear-gradient(135deg,#1e3a5f,#059669);padding:12px 16px;"><span style="font-size:11px;font-weight:700;color:white;letter-spacing:1.5px;text-transform:uppercase;">&#127381; Ticket Summary</span></div>'
            '<table width="100%" cellpadding="0" cellspacing="0">'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;width:38%;border-bottom:1px solid #f1f5f9;">Ticket No.</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;"><strong>{ticket.ticket_number}</strong></td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Title</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{ticket.title}</td></tr>'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Category</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{ticket.get_category_display()}</td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Resolved By</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{ticket.assigned_to.employee_name if ticket.assigned_to else "IT Department"}</td></tr>'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;">Resolution Note</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;">{ticket.resolution_note or "Issue resolved by IT team."}</td></tr>'
            '</table></div></td></tr>'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:3px;"></td></tr>'
            '<tr><td style="background:#f8fafc;padding:20px 40px;text-align:center;">'
            '<p style="font-size:11px;color:#94a3b8;margin:0;">This is an automated email from <strong>ERP Department — Unity Cement (PMC Cement Ltd.)</strong></p>'
            '<p style="font-size:11px;color:#94a3b8;margin:6px 0 0;">Do not reply to this email • For support contact IT/ERP Department</p>'
            '</td></tr>'
            '</table></td></tr></table></body></html>'
        )
        plain = f"""Dear {ticket.raised_by.employee_name},

Your ticket ({ticket.ticket_number}) has been resolved.

Title: {ticket.title}
Resolved By: {ticket.assigned_to.employee_name if ticket.assigned_to else 'IT Department'}
Resolution: {ticket.resolution_note or 'Issue resolved by IT team.'}

Regards,
ERP Department — Unity Cement"""
        msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [ticket.raised_by.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


@login_required
def dashboard(request):
    return redirect('helpdesk:ticket_list')


@login_required
def ticket_list(request):
    user = request.user
    is_it = _is_it(user)
    FULL_ROLES = ('administrator', 'management', 'president_plant_head')

    if is_it:
        tickets = Ticket.objects.select_related('raised_by', 'assigned_to').all()
    elif any(user.has_role(role) for role in FULL_ROLES) or user.is_superuser:
        tickets = Ticket.objects.select_related('raised_by', 'assigned_to').all()
    else:
        # Department-based: only see tickets from own department
        dept_q = Q(raised_by__department__in=user.get_all_departments())
        extra_dept_q = Q()
        for department in user.get_all_departments():
            extra_dept_q |= Q(raised_by__additional_departments__contains=f'|{department}|')
        tickets = Ticket.objects.select_related('raised_by', 'assigned_to').filter(dept_q | extra_dept_q)

    f_status   = request.GET.get('status', '')
    f_priority = request.GET.get('priority', '')
    f_category = request.GET.get('category', '')
    f_q        = request.GET.get('q', '')
    f_dept     = request.GET.get('dept', '')

    if f_status:   tickets = tickets.filter(status=f_status)
    if f_priority: tickets = tickets.filter(priority=f_priority)
    if f_category: tickets = tickets.filter(category=f_category)
    if f_dept and (is_it or any(user.has_role(role) for role in FULL_ROLES) or user.is_superuser):
        tickets = tickets.filter(raised_by__department=f_dept)
    if f_q: tickets = tickets.filter(
        Q(ticket_number__icontains=f_q) | Q(title__icontains=f_q) |
        Q(raised_by__employee_name__icontains=f_q)
    )

    from .models import STATUS_CHOICES, PRIORITY_CHOICES, CATEGORY_CHOICES
    from accounts.models import DEPARTMENT_CHOICES
    return render(request, 'helpdesk/ticket_list.html', {
        'tickets': tickets, 'is_it': is_it,
        'f_status': f_status, 'f_priority': f_priority,
        'f_category': f_category, 'f_q': f_q, 'f_dept': f_dept,
        'status_choices': STATUS_CHOICES,
        'priority_choices': PRIORITY_CHOICES,
        'category_choices': CATEGORY_CHOICES,
        'department_choices': DEPARTMENT_CHOICES,
        'show_dept_filter': is_it or any(user.has_role(role) for role in FULL_ROLES) or user.is_superuser,
    })


@login_required
def ticket_create(request):
    form = TicketForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        ticket = form.save(commit=False)
        ticket.raised_by = request.user
        ticket.save()
        _send_hd_it_notification(ticket)
        messages.success(request, f'Ticket {ticket.ticket_number} raised successfully. IT team will respond shortly.')
        return redirect('helpdesk:ticket_detail', pk=ticket.pk)
    return render(request, 'helpdesk/ticket_form.html', {'form': form, 'title': 'Raise New Ticket'})


@login_required
def ticket_detail(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    user = request.user
    is_it = _is_it(user)

    # Access control — non-IT can only see their own tickets
    if not is_it and ticket.raised_by != user:
        messages.error(request, 'Access Denied.')
        return redirect('helpdesk:ticket_list')

    comment_form = CommentForm(request.POST or None)
    update_form  = TicketUpdateForm(instance=ticket) if is_it else None

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'comment' and comment_form.is_valid():
            c = comment_form.save(commit=False)
            c.ticket = ticket
            c.author = user
            # Non-IT cannot post internal notes
            if not is_it:
                c.is_internal = False
            c.save()
            ticket.updated_at = timezone.now()
            ticket.save(update_fields=['updated_at'])
            messages.success(request, 'Comment added.')
            return redirect('helpdesk:ticket_detail', pk=pk)

        if action == 'update' and is_it:
            update_form = TicketUpdateForm(request.POST, instance=ticket)
            if update_form.is_valid():
                t = update_form.save(commit=False)
                old_status = ticket.status
                if t.status == 'resolved' and not ticket.resolved_at:
                    t.resolved_at = timezone.now()
                t.save()
                messages.success(request, f'Ticket {ticket.ticket_number} updated.')
                
                if old_status != 'resolved' and t.status == 'resolved':
                    _send_hd_resolve_ack(t)
                    send_workflow_notification(
                        module_key='hd',
                        notification_type='hd_resolved',
                        title=f'Ticket {ticket.ticket_number} Resolved',
                        description='Your help desk ticket has been resolved',
                        related_id=str(ticket.id),
                        related_module='HD',
                        requester=ticket.raised_by,
                        extra_users=[t.assigned_to] if t.assigned_to else None,
                    )
                elif t.assigned_to and ticket.assigned_to != t.assigned_to:
                    if t.assigned_to.perm_helpdesk_manage or t.assigned_to.has_role('administrator') or t.assigned_to.is_superuser:
                        send_workflow_notification(
                            module_key='hd',
                            notification_type='hd_assigned',
                            title=f'Ticket {ticket.ticket_number} Assigned to You',
                            description=f'A help desk ticket has been assigned to you: {ticket.title}',
                            related_id=str(ticket.id),
                            related_module='HD',
                            requester=ticket.raised_by,
                            extra_users=[t.assigned_to],
                        )
                
                return redirect('helpdesk:ticket_detail', pk=pk)

    comments = ticket.comments.select_related('author').all()
    if not is_it:
        comments = comments.filter(is_internal=False)

    return render(request, 'helpdesk/ticket_detail.html', {
        'ticket': ticket, 'comments': comments,
        'comment_form': comment_form, 'update_form': update_form,
        'is_it': is_it,
    })


@login_required
def ticket_close(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    if not _is_it(request.user) and ticket.raised_by != request.user:
        messages.error(request, 'Access Denied.')
        return redirect('helpdesk:ticket_list')
    ticket.status = 'closed'
    ticket.save(update_fields=['status', 'updated_at'])
    
    send_workflow_notification(
        module_key='hd',
        notification_type='hd_resolved',
        title=f'Ticket {ticket.ticket_number} Closed',
        description='Your help desk ticket has been closed',
        related_id=str(ticket.id),
        related_module='HD',
        requester=ticket.raised_by,
        extra_users=[ticket.assigned_to] if ticket.assigned_to else None,
    )
    
    messages.success(request, f'Ticket {ticket.ticket_number} closed.')
    return redirect('helpdesk:ticket_list')


@login_required
def bulk_action(request):
    if request.method != 'POST':
        return redirect('helpdesk:ticket_list')
    if not _is_it(request.user):
        messages.error(request, 'Access Denied.')
        return redirect('helpdesk:ticket_list')
    action = request.POST.get('action')
    ids    = request.POST.getlist('selected_ids')
    if not ids:
        messages.warning(request, 'No tickets selected.')
        return redirect('helpdesk:ticket_list')
    tickets = Ticket.objects.filter(pk__in=ids)
    if action == 'delete':
        count = tickets.count()
        tickets.delete()
        from accounts.models import AuditLog
        AuditLog.log(request, 'ticket_bulk_delete', 'TICKET', f'{count} Tickets deleted in bulk by {request.user.username}')
        messages.success(request, f'{count} ticket(s) deleted.')
    elif action == 'cancel':
        updated = tickets.exclude(status='closed').update(status='closed')
        from accounts.models import AuditLog
        AuditLog.log(request, 'ticket_bulk_cancel', 'TICKET', f'{updated} Tickets closed in bulk by {request.user.username}')
        messages.success(request, f'{updated} ticket(s) closed.')
    elif action == 'duplicate':
        for ticket in tickets:
            ticket.pk = None
            ticket.ticket_number = ''
            ticket.status = 'open'
            ticket.resolved_at = None
            ticket.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'ticket_bulk_duplicate', 'TICKET', f'{len(ids)} Tickets duplicated in bulk by {request.user.username}')
        messages.success(request, f'{len(ids)} ticket(s) duplicated.')
    else:
        messages.error(request, 'Unknown action.')
    return redirect('helpdesk:ticket_list')


@login_required
def export_report(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from django.http import HttpResponse
    from datetime import date

    status_f   = request.GET.get('status', '')
    priority_f = request.GET.get('priority', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    tickets = Ticket.objects.select_related('raised_by', 'assigned_to').all()
    if status_f:   tickets = tickets.filter(status=status_f)
    if priority_f: tickets = tickets.filter(priority=priority_f)
    if date_from:  tickets = tickets.filter(created_at__date__gte=date_from)
    if date_to:    tickets = tickets.filter(created_at__date__lte=date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HD Ticket Report'
    headers = ['Ticket No', 'Title', 'Category', 'Priority', 'Raised By', 'Department',
               'Assigned To', 'Status', 'Created', 'Resolved At', 'Resolution Note']
    hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 20
    for t in tickets:
        resolved_at_str = t.resolved_at.strftime('%d-%m-%Y %H:%M') if t.resolved_at else ''
        ws.append([
            t.ticket_number, t.title, t.get_category_display(), t.get_priority_display(),
            t.raised_by.employee_name, t.raised_by.department,
            t.assigned_to.employee_name if t.assigned_to else '',
            t.get_status_display(),
            t.created_at.strftime('%d-%m-%Y %H:%M'),
            resolved_at_str, t.resolution_note or '',
        ])
    from accounts.report_utils import add_excel_logo_and_note
    add_excel_logo_and_note(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="HD_Report_{date.today()}.xlsx"'
    return resp


@login_required
def export_report_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import date

    status_f   = request.GET.get('status', '')
    priority_f = request.GET.get('priority', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    tickets = Ticket.objects.select_related('raised_by', 'assigned_to').all()
    if status_f:   tickets = tickets.filter(status=status_f)
    if priority_f: tickets = tickets.filter(priority=priority_f)
    if date_from:  tickets = tickets.filter(created_at__date__gte=date_from)
    if date_to:    tickets = tickets.filter(created_at__date__lte=date_to)

    navy = colors.HexColor('#1e3a5f')
    alt  = colors.HexColor('#f8f9fa')
    sub_sty = ParagraphStyle('s', fontName='Helvetica', fontSize=8,
                             textColor=colors.grey, spaceAfter=4)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=12*mm, bottomMargin=10*mm,
                            leftMargin=10*mm, rightMargin=10*mm)
    elements = []
    from accounts.report_utils import build_pdf_header_table
    elements.append(build_pdf_header_table(
        'IT HELP DESK TICKET REPORT',
        f'Total: {tickets.count()}  |  Generated: {timezone.now().strftime("%d %b %Y %H:%M")}',
        277
    ))
    elements.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=4*mm))

    headers = ['#', 'Ticket No', 'Title', 'Category', 'Priority', 'Raised By', 'Dept', 'Status', 'Created', 'Resolved At']
    col_w = [8*mm, 22*mm, 50*mm, 28*mm, 20*mm, 30*mm, 28*mm, 20*mm, 30*mm, 30*mm]
    data = [headers]
    for i, t in enumerate(tickets, 1):
        data.append([
            str(i), t.ticket_number, t.title[:40], t.get_category_display(),
            t.get_priority_display(), t.raised_by.employee_name, t.raised_by.department,
            t.get_status_display(), t.created_at.strftime('%d-%m-%Y %H:%M'),
            t.resolved_at.strftime('%d-%m-%Y %H:%M') if t.resolved_at else '-',
        ])
    if len(data) == 1:
        data.append(['', 'No records found.'] + [''] * 8)
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('FONTNAME',       (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0), (-1,0),  8),
        ('BACKGROUND',     (0,0), (-1,0),  navy),
        ('TEXTCOLOR',      (0,0), (-1,0),  colors.white),
        ('FONTNAME',       (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',       (0,1), (-1,-1), 7.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt]),
        ('GRID',           (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
        ('VALIGN',         (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING',        (0,0), (-1,-1), 3),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph('This is a computer generated report. \u2014 Unity Cement ERP System', sub_sty))
    doc.build(elements)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="HD_Report_{date.today()}.pdf"'
    return resp
