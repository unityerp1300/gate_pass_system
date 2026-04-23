import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import os
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Count, Sum, Q
from datetime import date

from .models import MaterialGatePass, MaterialRequest
from .forms import MaterialGatePassForm, MaterialItemFormSet, ApprovalForm, ReturnForm, MaterialRequestForm, MaterialRequestItemFormSet, StoreReviewForm, HodReviewForm
from accounts.models import Employee, DEPARTMENT_CHOICES
from accounts.notification_service import send_workflow_notification


MGP_PRINT_COPIES = [
    ('original', 'Original For Buyer'),
    ('duplicate', 'Duplicate For Transporter'),
    ('triplicate', 'Triplicate For Supplier'),
]


def _base_url(request=None):
    # Always prefer SITE_BASE_URL setting for email links (works from any IP)
    site_url = getattr(settings, 'SITE_BASE_URL', None)
    if site_url:
        return site_url.rstrip('/')
    # Only fall back to request if SITE_BASE_URL is not configured
    return f"{request.scheme}://{request.get_host()}" if request else 'http://localhost:8000'


def _send_mgp_ack_mail(gp):
    """Send extraordinary acknowledgement email to MGP creator."""
    try:
        if not gp.employee.email:
            return
        from django.core.mail import EmailMultiAlternatives
        subject = f'MGP Submitted — {gp.pass_number}'
        html_body = (
            '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"></head>'
            '<body style="margin:0;padding:0;background:#f0f4f8;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',Inter,Arial,sans-serif;">'
            '<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:32px 0;"><tr><td align="center">'
            '<table width="620" cellpadding="0" cellspacing="0" style="background:white;border-radius:20px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,0.12);">'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:4px;"></td></tr>'
            '<tr><td style="background:linear-gradient(135deg,#0f172a,#1e3a5f,#dc2626);padding:36px 40px;text-align:center;">'
            '<div style="font-size:11px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:10px;">Unity Cement — ERP Department</div>'
            '<div style="width:64px;height:64px;border-radius:50%;background:rgba(255,255,255,0.12);border:2px solid rgba(255,255,255,0.25);display:inline-flex;align-items:center;justify-content:center;margin-bottom:16px;font-size:28px;">&#128230;</div>'
            '<div style="font-size:26px;font-weight:800;color:white;margin-bottom:6px;">MGP Submitted</div>'
            '<div style="font-size:14px;color:rgba(255,255,255,0.7);margin-bottom:16px;">Your Material Gate Pass has been submitted for approval</div>'
            f'<div style="display:inline-block;background:rgba(240,165,0,0.2);border:1.5px solid rgba(240,165,0,0.5);border-radius:50px;padding:6px 20px;font-size:14px;color:#fbbf24;font-weight:700;">{gp.pass_number}</div>'
            '</td></tr>'
            f'<tr><td style="padding:28px 40px 0;"><p style="font-size:16px;color:#1e293b;margin:0 0 8px;">Dear <strong style="color:#1e3a5f;">{gp.employee.employee_name}</strong>,</p>'
            '<p style="font-size:14px;color:#64748b;margin:0;line-height:1.6;">Your Material Gate Pass has been successfully submitted and is now <strong>pending Store HOD approval</strong>. You will receive an email once it is actioned.</p>'
            '</td></tr>'
            '<tr><td style="padding:20px 40px;">'
            '<div style="border-radius:12px;overflow:hidden;border:1px solid #e2e8f0;">'
            '<div style="background:linear-gradient(135deg,#1e3a5f,#dc2626);padding:12px 16px;"><span style="font-size:11px;font-weight:700;color:white;letter-spacing:1.5px;text-transform:uppercase;">&#128230; Pass Summary</span></div>'
            '<table width="100%" cellpadding="0" cellspacing="0">'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;width:38%;border-bottom:1px solid #f1f5f9;">Pass Number</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;"><strong>{gp.pass_number}</strong></td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Direction</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{gp.get_direction_display()}</td></tr>'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Type</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{gp.get_type_label()}</td></tr>'
            f'<tr style="background:#f8fafc;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Party</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{gp.party_name}</td></tr>'
            f'<tr style="background:white;"><td style="padding:9px 16px;font-size:12px;font-weight:600;color:#475569;">Date</td><td style="padding:9px 16px;font-size:13px;color:#1e293b;">{gp.pass_date}</td></tr>'
            '</table></div></td></tr>'
            '<tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22,#f0a500);height:3px;"></td></tr>'
            '<tr><td style="background:#f8fafc;padding:20px 40px;text-align:center;">'
            '<p style="font-size:11px;color:#94a3b8;margin:0;">This is an automated email from <strong>ERP Department — Unity Cement (PMC Cement Ltd.)</strong></p>'
            '</td></tr>'
            '</table></td></tr></table></body></html>'
        )
        plain = f"""Dear {gp.employee.employee_name},

Your MGP ({gp.pass_number}) has been submitted for Store HOD approval.

Direction: {gp.get_direction_display()}
Type: {gp.get_type_label()}
Party: {gp.party_name}
Date: {gp.pass_date}

Regards,
ERP Department — Unity Cement"""
        msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [gp.employee.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def _get_store_hod_approver():
    """Return the Store HOD / MGP approver for the MGP workflow."""
    return (
        Employee.objects.filter(
            is_active=True,
        ).filter(
            Q(department='Store') |
            Q(additional_departments__contains='|Store|')
        )
        .filter(
            Q(role='department_hod') |
            Q(additional_roles__contains='|department_hod|') |
            Q(perm_mgp_approve=True)
        )
        .order_by('-perm_mgp_approve', 'employee_name')
        .first()
        or Employee.objects.filter(is_active=True, role='administrator').first()
    )


def _can_mgp_approve(user):
    if not getattr(user, 'is_authenticated', False):
        return False
    if user.is_superuser or user.role == 'administrator':
        return True
    return user.is_active and user.has_department('Store') and (
        user.has_role('department_hod') or user.perm_mgp_approve
    )


def _send_approval_mail(gp, request=None):
    try:
        approver = _get_store_hod_approver()
        if not approver or not approver.email:
            return
        base         = _base_url(request)
        approve_url  = f"{base}/material-pass/token-action/{gp.approval_token}/approve/"
        reject_url   = f"{base}/material-pass/token-action/{gp.approval_token}/reject/"
        subject      = f'[Action Required] Material Gate Pass — {gp.pass_number}'

        html_body = f'''
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:Inter,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:30px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:white;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">

  <!-- Header -->
  <tr>
    <td style="background:linear-gradient(135deg,#1e3a5f,#2d5a8e);padding:28px 32px;text-align:center;">
      <div style="font-size:11px;color:rgba(255,255,255,0.6);letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;">Unity Cement — ERP Department</div>
      <div style="font-size:22px;font-weight:700;color:white;margin-bottom:4px;">Approval Required</div>
      <div style="font-size:13px;color:rgba(255,255,255,0.75);">Material Gate Pass</div>
      <div style="display:inline-block;background:rgba(255,255,255,0.15);border-radius:20px;padding:4px 16px;margin-top:10px;font-size:13px;color:white;font-weight:600;">{gp.pass_number}</div>
    </td>
  </tr>

  <!-- Greeting -->
  <tr>
    <td style="padding:24px 32px 0;">
      <p style="font-size:15px;color:#2d3748;margin:0;">Dear <strong>{approver.employee_name}</strong>,</p>
      <p style="font-size:13px;color:#718096;margin:8px 0 0;">A Material Gate Pass requires Store HOD approval. Please review the details below and take action.</p>
    </td>
  </tr>

  <!-- Pass Details -->
  <tr>
    <td style="padding:20px 32px;">
      <table width="100%" cellpadding="0" cellspacing="0" style="background:#f8fafc;border-radius:10px;border:1px solid #e2e8f0;overflow:hidden;">
        <tr style="background:#1e3a5f;">
          <td colspan="2" style="padding:10px 16px;font-size:11px;font-weight:700;color:white;letter-spacing:1px;text-transform:uppercase;">Pass Information</td>
        </tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;width:40%;">Pass Number</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;font-weight:700;">{gp.pass_number}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Employee</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.employee.employee_name} ({gp.employee.employee_code})</td></tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Department</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.employee.department}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Direction</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.get_direction_display()}</td></tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Material Type</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.get_type_label()}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Party / Consignee</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.party_name}</td></tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Date</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.pass_date}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Vehicle No.</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.vehicle_number or "-"}</td></tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Reason</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.reason or "-"}</td></tr>
      </table>
    </td>
  </tr>

  <!-- Action Buttons -->
  <tr>
    <td style="padding:0 32px 28px;">
      <p style="font-size:13px;color:#2d3748;font-weight:600;margin:0 0 14px;">Take Action:</p>
      <table cellpadding="0" cellspacing="0" width="100%">
        <tr>
          <td width="48%" align="center">
            <a href="{approve_url}" style="display:block;background:linear-gradient(135deg,#27ae60,#2ecc71);color:white;text-decoration:none;padding:14px 20px;border-radius:10px;font-size:15px;font-weight:700;text-align:center;">&#10003;&nbsp; APPROVE</a>
          </td>
          <td width="4%"></td>
          <td width="48%" align="center">
            <a href="{reject_url}" style="display:block;background:linear-gradient(135deg,#c0392b,#e74c3c);color:white;text-decoration:none;padding:14px 20px;border-radius:10px;font-size:15px;font-weight:700;text-align:center;">&#10007;&nbsp; REJECT</a>
          </td>
        </tr>
      </table>
      <p style="font-size:11px;color:#a0aec0;text-align:center;margin:12px 0 0;">Clicking a button will open a confirmation page where you can add remarks before submitting.</p>
    </td>
  </tr>

  <!-- Footer -->
  <tr>
    <td style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:16px 32px;text-align:center;">
      <p style="font-size:11px;color:#a0aec0;margin:0;">This is an automated email from the ERP Department — Unity Cement.</p>
      <p style="font-size:11px;color:#a0aec0;margin:4px 0 0;">Do not reply to this email. For support, contact the IT/ERP Department.</p>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>'''

        plain_body = (
            f"Dear {approver.employee_name},\n\n"
            f"Material Gate Pass {gp.pass_number} requires Store HOD approval.\n\n"
            f"Employee : {gp.employee.employee_name} ({gp.employee.employee_code})\n"
            f"Direction: {gp.get_direction_display()}\n"
            f"Type     : {gp.get_type_label()}\n"
            f"Party    : {gp.party_name}\n"
            f"Date     : {gp.pass_date}\n\n"
            f"APPROVE: {approve_url}\n"
            f"REJECT : {reject_url}\n\n"
            f"Regards,\nERP Department — Unity Cement"
        )
        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain_body, settings.DEFAULT_FROM_EMAIL, [approver.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def _send_status_mail(gp, action):
    try:
        status_colors = {
            'approved': ('#d1e7dd', '#0f5132', '#27ae60', 'APPROVED &#10003;'),
            'rejected': ('#f8d7da', '#842029', '#e74c3c', 'REJECTED &#10007;'),
            'returned': ('#cff4fc', '#055160', '#0ea5e9', 'RETURNED &#8635;'),
        }
        bg, fg, hdr_color, label = status_colors.get(action, ('#e2e3e5', '#383d41', '#6c757d', action.upper()))
        subject = f'Material Gate Pass {action.upper()} — {gp.pass_number}'

        html_body = f'''
<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:Inter,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:30px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:white;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.08);">

  <!-- Header -->
  <tr>
    <td style="background:linear-gradient(135deg,#1e3a5f,#2d5a8e);padding:28px 32px;text-align:center;">
      <div style="font-size:11px;color:rgba(255,255,255,0.6);letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;">Unity Cement — ERP Department</div>
      <div style="font-size:22px;font-weight:700;color:white;margin-bottom:4px;">Gate Pass Status Update</div>
      <div style="display:inline-block;background:rgba(255,255,255,0.15);border-radius:20px;padding:4px 16px;margin-top:10px;font-size:13px;color:white;font-weight:600;">{gp.pass_number}</div>
    </td>
  </tr>

  <!-- Status Badge -->
  <tr>
    <td style="padding:24px 32px 0;text-align:center;">
      <div style="display:inline-block;background:{bg};color:{fg};border-radius:50px;padding:10px 32px;font-size:16px;font-weight:700;letter-spacing:1px;">{label}</div>
    </td>
  </tr>

  <!-- Greeting -->
  <tr>
    <td style="padding:16px 32px 0;">
      <p style="font-size:15px;color:#2d3748;margin:0;">Dear <strong>{gp.employee.employee_name}</strong>,</p>
      <p style="font-size:13px;color:#718096;margin:8px 0 0;">Your Material Gate Pass <strong>{gp.pass_number}</strong> has been <strong>{action}</strong>.</p>
    </td>
  </tr>

  <!-- Pass Details -->
  <tr>
    <td style="padding:20px 32px;">
      <table width="100%" cellpadding="0" cellspacing="0" style="background:#f8fafc;border-radius:10px;border:1px solid #e2e8f0;overflow:hidden;">
        <tr style="background:#1e3a5f;">
          <td colspan="2" style="padding:10px 16px;font-size:11px;font-weight:700;color:white;letter-spacing:1px;text-transform:uppercase;">Pass Details</td>
        </tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;width:40%;">Pass Number</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;font-weight:700;">{gp.pass_number}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Direction</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.get_direction_display()}</td></tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Material Type</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.get_type_label()}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Party / Consignee</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.party_name}</td></tr>
        <tr style="background:white;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Date</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.pass_date}</td></tr>
        {f'<tr style="background:#f8fafc;"><td style="padding:8px 16px;font-size:12px;font-weight:600;color:#1e3a5f;">Approver Remarks</td><td style="padding:8px 16px;font-size:13px;color:#2d3748;">{gp.approval_remarks}</td></tr>' if gp.approval_remarks else ''}
      </table>
    </td>
  </tr>

  <!-- Footer -->
  <tr>
    <td style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:16px 32px;text-align:center;">
      <p style="font-size:11px;color:#a0aec0;margin:0;">This is an automated email from the ERP Department — Unity Cement.</p>
      <p style="font-size:11px;color:#a0aec0;margin:4px 0 0;">Do not reply to this email. For support, contact the IT/ERP Department.</p>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>'''

        plain_body = (
            f"Dear {gp.employee.employee_name},\n\n"
            f"Your Material Gate Pass ({gp.pass_number}) has been {action}.\n\n"
            f"Direction: {gp.get_direction_display()}\n"
            f"Type     : {gp.get_type_label()}\n"
            f"Party    : {gp.party_name}\n"
            f"Date     : {gp.pass_date}\n"
            + (f"Remarks  : {gp.approval_remarks}\n" if gp.approval_remarks else "") +
            f"\nRegards,\nERP Department — Unity Cement"
        )
        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain_body, settings.DEFAULT_FROM_EMAIL, [gp.employee.email])
        msg.attach_alternative(html_body, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


# ─── Views ──────────────────────────────────────────────────────────────────

@login_required
def pass_list(request):
    if not request.user.perm_mgp_view and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')

    user = request.user
    if user.is_superuser or user.has_role('administrator') or user.has_role('management') or user.has_role('president_plant_head'):
        passes = MaterialGatePass.objects.all()
    elif user.has_role('hr') or user.has_role('security'):
        passes = MaterialGatePass.objects.all()
    elif user.has_role('department_hod'):
        dept_q = Q(employee__department__in=user.get_all_departments())
        extra_dept_q = Q()
        for department in user.get_all_departments():
            extra_dept_q |= Q(employee__additional_departments__contains=f'|{department}|')
        passes = MaterialGatePass.objects.filter(dept_q | extra_dept_q)
    else:
        passes = MaterialGatePass.objects.filter(employee=user)

    status_filter = request.GET.get('status', '')
    dept_filter   = request.GET.get('dept', '')
    search        = request.GET.get('q', '').strip()
    if status_filter:
        passes = passes.filter(status=status_filter)
    FULL_ROLES = ('administrator', 'management', 'president_plant_head')
    if dept_filter and (user.is_superuser or any(user.has_role(role) for role in FULL_ROLES)):
        passes = passes.filter(employee__department=dept_filter)
    if search:
        passes = passes.filter(
            Q(pass_number__icontains=search) |
            Q(employee__employee_name__icontains=search) |
            Q(employee__employee_code__icontains=search) |
            Q(employee__department__icontains=search) |
            Q(party_name__icontains=search) |
            Q(vehicle_number__icontains=search)
        )

    can_approve = _can_mgp_approve(user)
    from django.core.paginator import Paginator
    paginator = Paginator(passes, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'material_pass/pass_list.html', {
        'passes': page_obj, 'page_obj': page_obj,
        'status_filter': status_filter,
        'dept_filter': dept_filter,
        'search': search,
        'department_choices': DEPARTMENT_CHOICES,
        'show_dept_filter': user.is_superuser or any(user.has_role(role) for role in FULL_ROLES),
        'can_approve': can_approve,
    })


@login_required
def pass_create(request):
    if not request.user.perm_mgp_write and not request.user.is_superuser:
        messages.error(request, 'Access Denied.')
        return redirect('dashboard:index')

    form = MaterialGatePassForm(request.POST or None)
    formset = MaterialItemFormSet(request.POST or None)

    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        gp = form.save(commit=False)
        gp.employee = request.user
        gp.status = 'pending'
        gp.save()
        formset.instance = gp
        formset.save()
        _send_approval_mail(gp, request)
        _send_mgp_ack_mail(gp)  # acknowledgement to creator
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_create', 'MGP', f'MGP created: {gp.pass_number} by {request.user.username}')
        messages.success(request, f'Material Gate Pass {gp.pass_number} submitted successfully.')
        return redirect('material_pass:pass_list')

    return render(request, 'material_pass/pass_form.html', {
        'form': form, 'formset': formset, 'title': 'New Material Gate Pass'
    })


@login_required
def pass_detail(request, pk):
    gp = get_object_or_404(MaterialGatePass, pk=pk)
    user = request.user
    can_approve = _can_mgp_approve(user) and gp.status == 'pending'
    can_edit = (
        user.is_superuser or
        user.has_role('administrator') or
        (user.has_role('department_hod') and user.has_department('Store')) or
        (gp.status == 'pending' and gp.employee == user)
    )
    return render(request, 'material_pass/pass_detail.html', {
        'gp': gp, 'can_approve': can_approve, 'can_edit': can_edit
    })


@login_required
def pass_edit(request, pk):
    """Edit MGP — only administrator or Store HOD can edit past/approved records."""
    gp = get_object_or_404(MaterialGatePass, pk=pk)
    user = request.user
    can_edit_past = user.is_superuser or user.has_role('administrator') or (
        user.has_role('department_hod') and user.has_department('Store')
    )
    # Pending records can be edited by creator; approved/past only by admin/store hod
    if gp.status != 'pending' and not can_edit_past:
        messages.error(request, 'Only Administrator or Store HOD can edit approved/past MGP records.')
        return redirect('material_pass:pass_detail', pk=pk)
    if gp.status == 'pending' and gp.employee != user and not can_edit_past:
        messages.error(request, 'Access denied.')
        return redirect('material_pass:pass_detail', pk=pk)

    from .forms import MaterialGatePassForm, MaterialItemFormSet
    form = MaterialGatePassForm(request.POST or None, instance=gp)
    formset = MaterialItemFormSet(request.POST or None, instance=gp)
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        form.save()
        formset.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_edit', 'MGP', f'MGP edited: {gp.pass_number} by {user.username}')
        messages.success(request, f'MGP {gp.pass_number} updated successfully.')
        return redirect('material_pass:pass_detail', pk=pk)
    return render(request, 'material_pass/pass_form.html', {
        'form': form, 'formset': formset,
        'title': f'Edit MGP — {gp.pass_number}',
        'is_edit': True, 'gp': gp,
    })


@login_required
def pass_approve(request, pk):
    gp = get_object_or_404(MaterialGatePass, pk=pk)
    if not _can_mgp_approve(request.user):
        messages.error(request, 'Only Store HOD can approve or reject Material Gate Passes.')
        return redirect('material_pass:pass_list')
    if gp.status != 'pending':
        messages.warning(request, 'This pass has already been actioned.')
        return redirect('material_pass:pass_detail', pk=pk)

    form = ApprovalForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        action  = form.cleaned_data['action']
        remarks = form.cleaned_data['remarks']
        gp.status = action
        gp.approver = request.user
        gp.approval_remarks = remarks
        gp.approved_at = timezone.now()
        gp.save()
        _send_status_mail(gp, action)
        from accounts.models import AuditLog
        log_action = 'mgp_approve' if action == 'approved' else 'mgp_reject'
        AuditLog.log(request, log_action, 'MGP', f'MGP {action}: {gp.pass_number} by {request.user.username}')
        
        notification_type = 'mgp_approved' if action == 'approved' else 'mgp_rejected'
        title = f'MGP {gp.pass_number} {action.title()}'
        description = f'Your Material Gate Pass has been {action}'
        send_workflow_notification(
            module_key='mgp',
            notification_type=notification_type,
            title=title,
            description=description,
            related_id=str(gp.id),
            related_module='MGP',
            requester=gp.employee,
        )
        
        messages.success(request, f'Gate Pass {gp.pass_number} {action}.')
        return redirect('material_pass:pass_list')

    return render(request, 'material_pass/pass_approve.html', {'gp': gp, 'form': form})


def token_action(request, token, action):
    """Email link-based approval."""
    gp = get_object_or_404(MaterialGatePass, approval_token=token)
    if gp.status != 'pending':
        return render(request, 'material_pass/token_done.html', {
            'message': f'This pass has already been {gp.status}.', 'gp': gp
        })
    if request.method == 'POST':
        chosen  = request.POST.get('action_choice', action)
        remarks = request.POST.get('remarks', '')
        gp.status = 'approved' if chosen == 'approve' else 'rejected'
        gp.approval_remarks = remarks
        gp.approved_at = timezone.now()
        gp.save()
        _send_status_mail(gp, gp.status)
        
        notification_type = 'mgp_approved' if gp.status == 'approved' else 'mgp_rejected'
        title = f'MGP {gp.pass_number} {gp.status.title()}'
        description = f'Your Material Gate Pass has been {gp.status}'
        send_workflow_notification(
            module_key='mgp',
            notification_type=notification_type,
            title=title,
            description=description,
            related_id=str(gp.id),
            related_module='MGP',
            requester=gp.employee,
        )
        
        return render(request, 'material_pass/token_done.html', {
            'message': f'Gate Pass {gp.pass_number} has been {gp.status}.', 'gp': gp
        })
    return render(request, 'material_pass/token_action.html', {'gp': gp, 'action': action})


@login_required
def mark_returned(request, pk):
    gp = get_object_or_404(MaterialGatePass, pk=pk)
    if not gp.is_returnable or gp.status != 'approved':
        messages.warning(request, 'Only approved returnable passes can be marked returned.')
        return redirect('material_pass:pass_detail', pk=pk)
    form = ReturnForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        gp.actual_return_date = form.cleaned_data['actual_return_date']
        gp.status = 'returned'
        gp.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_return', 'MGP', f'MGP marked returned: {gp.pass_number} by {request.user.username}')
        messages.success(request, 'Marked as returned.')
        return redirect('material_pass:pass_list')
    return render(request, 'material_pass/mark_returned.html', {'gp': gp, 'form': form})


@login_required
def print_pass(request, pk):
    gp = get_object_or_404(MaterialGatePass, pk=pk)
    return render(request, 'material_pass/challan_print.html', {
        'gp': gp,
        'copies': MGP_PRINT_COPIES,
    })


# ─── Material Request Views ─────────────────────────────────────────────────

@login_required
def request_list(request):
    user = request.user
    if user.is_superuser or any(user.has_role(role) for role in ('administrator', 'management', 'president_plant_head', 'hr', 'security')):
        requests = MaterialRequest.objects.all()
    elif user.has_role('department_hod'):
        dept_q = Q(employee__department__in=user.get_all_departments())
        extra_dept_q = Q()
        for department in user.get_all_departments():
            extra_dept_q |= Q(employee__additional_departments__contains=f'|{department}|')
        requests = MaterialRequest.objects.filter(dept_q | extra_dept_q)
    else:
        requests = MaterialRequest.objects.filter(employee=user)

    # Store should be able to act on requests across the approval workflow.
    if user.has_department('Store') and not user.is_superuser:
        requests = MaterialRequest.objects.filter(
            status__in=[
                'submitted',
                'hod_approved',
                'hod_rejected',
                'approved',
                'rejected',
                'store_approved',
                'store_rejected',
                'converted',
            ]
        )

    status_f = request.GET.get('status', '')
    search = request.GET.get('q', '').strip()
    if status_f:
        requests = requests.filter(status=status_f)
    if search:
        requests = requests.filter(
            Q(request_number__icontains=search) |
            Q(employee__employee_name__icontains=search) |
            Q(employee__employee_code__icontains=search) |
            Q(department__icontains=search) |
            Q(reason__icontains=search) |
            Q(remarks__icontains=search)
        )
    from django.core.paginator import Paginator
    paginator = Paginator(requests, 10)
    page_obj  = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'material_pass/request_list.html', {
        'requests': page_obj, 'page_obj': page_obj,
        'status_filter': status_f,
        'is_store': user.has_department('Store') or user.is_superuser or user.has_role('administrator') or user.has_role('management'),
    })


@login_required
def request_create(request):
    form    = MaterialRequestForm(request.POST or None)
    formset = MaterialRequestItemFormSet(request.POST or None)
    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        mr = form.save(commit=False)
        mr.employee = request.user
        mr.status = 'submitted'
        mr.save()
        formset.instance = mr
        formset.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_create', 'MR', f'Material Gate Pass Request created: {mr.request_number} by {request.user.username}')
        messages.success(request, f'Material Gate Pass Request {mr.request_number} submitted to Store.')
        return redirect('material_pass:request_list')
    return render(request, 'material_pass/request_form.html', {
        'form': form, 'formset': formset, 'title': 'New Material Gate Pass Request'
    })


@login_required
def request_detail(request, pk):
    mr = get_object_or_404(MaterialRequest, pk=pk)
    is_store = request.user.has_department('Store') or request.user.is_superuser or request.user.has_role('administrator') or request.user.has_role('management')
    can_review = is_store and mr.status in ('submitted', 'hod_approved')
    can_convert = is_store and mr.status in ('store_approved', 'approved')
    return render(request, 'material_pass/request_detail.html', {
        'mr': mr, 'can_review': can_review, 'can_convert': can_convert, 'is_store': is_store
    })


@login_required
def request_review(request, pk):
    """Store department approves or rejects the material request."""
    mr = get_object_or_404(MaterialRequest, pk=pk)
    is_store = request.user.has_department('Store') or request.user.is_superuser or request.user.has_role('administrator') or request.user.has_role('management')
    if not is_store:
        messages.error(request, 'Only Store department can review material requests.')
        return redirect('material_pass:request_list')
    if mr.status not in ('submitted', 'hod_approved'):
        messages.warning(request, 'This request has already been actioned.')
        return redirect('material_pass:request_detail', pk=pk)

    form = StoreReviewForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        action = form.cleaned_data['action']
        mr.status = action
        mr.reviewed_by = request.user
        mr.reviewed_at = timezone.now()
        mr.review_remarks = form.cleaned_data['review_remarks']
        mr.save()
        messages.success(request, f'Request {mr.request_number} {action}.')
        return redirect('material_pass:request_detail', pk=pk)
    return render(request, 'material_pass/request_review.html', {'mr': mr, 'form': form})


@login_required
def request_convert(request, pk):
    """Store converts an approved material request into a Material Gate Pass."""
    mr = get_object_or_404(MaterialRequest, pk=pk)
    is_store = request.user.has_department('Store') or request.user.is_superuser or request.user.has_role('administrator') or request.user.has_role('management')
    if not is_store:
        messages.error(request, 'Only Store department can create gate passes from requests.')
        return redirect('material_pass:request_list')
    if mr.status not in ('store_approved', 'approved'):
        messages.warning(request, 'Only store-approved requests can be converted.')
        return redirect('material_pass:request_detail', pk=pk)

    from .models import MaterialItem
    from .forms import MaterialItemFormSet as MGPItemFormSet

    # Pre-fill gate pass form from request data
    initial = {
        'department':   mr.department,
        'is_returnable': mr.is_returnable,
        'reason':       mr.reason,
        'remarks':      mr.remarks,
        'pass_date':    date.today(),
        'expected_return_date': mr.expected_date,
        'direction':    'outgoing',
    }
    form    = MaterialGatePassForm(request.POST or None, initial=initial)
    formset = MaterialItemFormSet(request.POST or None)

    if request.method == 'POST' and form.is_valid() and formset.is_valid():
        gp = form.save(commit=False)
        gp.employee   = request.user
        gp.status     = 'pending'
        gp.request_ref = mr
        gp.save()
        formset.instance = gp
        formset.save()
        mr.status = 'converted'
        mr.save(update_fields=['status'])
        _send_approval_mail(gp, request)
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_create', 'MGP', f'MGP {gp.pass_number} created from MR {mr.request_number} by {request.user.username}')
        messages.success(request, f'Gate Pass {gp.pass_number} created from Request {mr.request_number}.')
        return redirect('material_pass:pass_detail', pk=gp.pk)

    # Pre-populate formset with request items
    item_initial = [{'description': i.description, 'hsn_code': i.hsn_code, 'quantity': i.quantity, 'unit': i.unit} for i in mr.items.all()]
    if request.method == 'GET' and item_initial:
        from django.forms import formset_factory
        formset = MaterialItemFormSet(initial=item_initial, queryset=MaterialItem.objects.none())

    return render(request, 'material_pass/request_convert.html', {
        'mr': mr, 'form': form, 'formset': formset, 'title': f'Create Gate Pass from {mr.request_number}'
    })


@login_required
def daily_report(request):
    report_date   = request.GET.get('date', str(date.today()))
    dept          = request.GET.get('dept', '')
    status_filter = request.GET.get('status', '')
    type_filter   = request.GET.get('type', '')
    view_mode     = request.GET.get('view', 'detail')  # summary | detail
    passes = MaterialGatePass.objects.filter(pass_date=report_date)
    if dept:          passes = passes.filter(employee__department=dept)
    if status_filter: passes = passes.filter(status=status_filter)
    if type_filter != '': passes = passes.filter(is_returnable=bool(int(type_filter)))
    ctx = _report_ctx(passes, report_date, 'Daily', dept, status_filter, type_filter)
    ctx['view_mode'] = view_mode
    ctx['summary'] = _mgp_summary(passes)
    return render(request, 'material_pass/report.html', ctx)


@login_required
def monthly_report(request):
    month  = request.GET.get('month', date.today().strftime('%Y-%m'))
    year, mon = month.split('-')
    dept          = request.GET.get('dept', '')
    status_filter = request.GET.get('status', '')
    type_filter   = request.GET.get('type', '')
    view_mode     = request.GET.get('view', 'detail')  # summary | detail
    passes = MaterialGatePass.objects.filter(pass_date__year=year, pass_date__month=mon)
    if dept:          passes = passes.filter(employee__department=dept)
    if status_filter: passes = passes.filter(status=status_filter)
    if type_filter != '': passes = passes.filter(is_returnable=bool(int(type_filter)))
    ctx = _report_ctx(passes, month, 'Monthly', dept, status_filter, type_filter)
    ctx['view_mode'] = view_mode
    ctx['summary'] = _mgp_summary(passes)
    return render(request, 'material_pass/report.html', ctx)


def _mgp_summary(qs):
    """Build summary aggregates for material gate pass reports."""
    status_map = dict(MaterialGatePass._meta.get_field('status').choices)

    total = qs.count()
    invoice_total = sum((gp.get_grand_total() for gp in qs), 0) if total else 0

    by_status = []
    for row in qs.values('status').annotate(count=Count('id')).order_by('-count'):
        key = row['status'] or ''
        by_status.append({'key': key, 'label': status_map.get(key, key or '—'), 'count': row['count']})

    by_department = []
    for row in qs.values('employee__department').annotate(count=Count('id')).order_by('-count'):
        key = row['employee__department'] or ''
        by_department.append({'label': key or '—', 'count': row['count']})

    by_type = [
        {'label': 'Returnable', 'count': qs.filter(is_returnable=True).count()},
        {'label': 'Non-Returnable', 'count': qs.filter(is_returnable=False).count()},
    ]

    return {'total': total, 'invoice_total': invoice_total, 'by_status': by_status, 'by_department': by_department, 'by_type': by_type}


def _report_ctx(passes, report_date, report_type, dept, status_filter, type_filter):
    return {
        'passes': passes, 'report_date': report_date, 'report_type': report_type,
        'dept_filter': dept, 'department_choices': DEPARTMENT_CHOICES,
        'status_filter': status_filter, 'type_filter': type_filter,
        'total_count':          passes.count(),
        'pending_count':        passes.filter(status='pending').count(),
        'approved_count':       passes.filter(status='approved').count(),
        'returned_count':       passes.filter(status='returned').count(),
        'returnable_count':     passes.filter(is_returnable=True).count(),
        'non_returnable_count': passes.filter(is_returnable=False).count(),
        'grand_total':          sum(gp.get_grand_total() for gp in passes),
    }


@login_required
def export_report(request):
    report_type   = request.GET.get('type', 'daily')
    report_date   = request.GET.get('date', str(date.today()))
    dept          = request.GET.get('dept', '')
    status_filter = request.GET.get('status', '')
    type_filter   = request.GET.get('type_filter', '')

    if report_type == 'monthly':
        year, mon = report_date.split('-')
        passes = MaterialGatePass.objects.filter(pass_date__year=year, pass_date__month=mon)
    else:
        passes = MaterialGatePass.objects.filter(pass_date=report_date)
    if dept:          passes = passes.filter(employee__department=dept)
    if status_filter: passes = passes.filter(status=status_filter)
    if type_filter != '': passes = passes.filter(is_returnable=bool(int(type_filter)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MGP Report'
    headers = ['Pass No', 'Employee', 'Emp Code', 'Department', 'Direction', 'Type',
               'Party', 'Vehicle No', 'Date', 'Items', 'Subtotal', 'CGST', 'SGST',
               'Rounding Off', 'Invoice Total',
               'Approved By', 'Approval Date & Time', 'Status']
    hfill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = 18
    for gp in passes:
        approved_at_str = gp.approved_at.strftime('%d-%m-%Y %H:%M') if gp.approved_at else ''
        ws.append([
            gp.pass_number, gp.employee.employee_name, gp.employee.employee_code,
            gp.employee.department, gp.get_direction_display(), gp.get_type_label(),
            gp.party_name, gp.vehicle_number, str(gp.pass_date),
            gp.items.count(), gp.get_subtotal(), gp.get_cgst_amount(),
            gp.get_sgst_amount(), float(gp.rounding_off), gp.get_grand_total(),
            gp.approver.employee_name if gp.approver else '',
            approved_at_str, gp.get_status_display(),
        ])
    from accounts.report_utils import add_excel_logo_and_note
    add_excel_logo_and_note(ws, len(headers))
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="MGP_Report_{report_date}.xlsx"'
    return response


@login_required
def export_report_pdf(request):
    from reportlab.lib.pagesizes import landscape
    report_type   = request.GET.get('type', 'daily')
    report_date   = request.GET.get('date', str(date.today()))
    dept          = request.GET.get('dept', '')
    status_filter = request.GET.get('status', '')
    type_filter   = request.GET.get('type_filter', '')

    if report_type == 'monthly':
        year, mon = report_date.split('-')
        passes = MaterialGatePass.objects.filter(pass_date__year=year, pass_date__month=mon)
    else:
        passes = MaterialGatePass.objects.filter(pass_date=report_date)
    if dept:          passes = passes.filter(employee__department=dept)
    if status_filter: passes = passes.filter(status=status_filter)
    if type_filter != '': passes = passes.filter(is_returnable=bool(int(type_filter)))

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=12*mm, bottomMargin=10*mm,
                            leftMargin=10*mm, rightMargin=10*mm)
    navy = colors.HexColor('#1e3a5f')
    alt  = colors.HexColor('#f8f9fa')
    bdr  = colors.HexColor('#dee2e6')

    title_sty = ParagraphStyle('t', fontName='Helvetica-Bold', fontSize=14,
                               alignment=TA_CENTER, textColor=navy, spaceAfter=2)
    sub_sty   = ParagraphStyle('s', fontName='Helvetica', fontSize=9,
                               alignment=TA_CENTER, textColor=colors.grey, spaceAfter=6)
    elements  = []
    from accounts.report_utils import build_pdf_header_table
    elements.append(build_pdf_header_table(
        'MATERIAL GATE PASS REPORT',
        '%s Report \u2014 %s  |  Total: %d  |  Generated: %s' % (
            report_type.title(), report_date, passes.count(), timezone.now().strftime('%d %b %Y %H:%M')),
        277
    ))
    elements.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=4*mm))

    headers = ['#', 'Pass No', 'Employee', 'Dept', 'Direction', 'Type',
               'Party', 'Vehicle', 'Date', 'Items', 'Total (₹)', 'Status']
    col_w = [8*mm, 22*mm, 30*mm, 28*mm, 22*mm, 22*mm,
             30*mm, 18*mm, 18*mm, 12*mm, 20*mm, 18*mm]
    data = [headers]
    for i, gp in enumerate(passes, 1):
        data.append([
            str(i), gp.pass_number, gp.employee.employee_name,
            gp.employee.department, gp.get_direction_display(), gp.get_type_label(),
            gp.party_name, gp.vehicle_number or '-', str(gp.pass_date),
            str(gp.items.count()), str(gp.get_grand_total()), gp.get_status_display(),
        ])
    if len(data) == 1:
        data.append(['', 'No records found.'] + [''] * 10)

    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('FONTNAME',       (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',       (0,0), (-1,0),  8),
        ('BACKGROUND',     (0,0), (-1,0),  navy),
        ('TEXTCOLOR',      (0,0), (-1,0),  colors.white),
        ('FONTNAME',       (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE',       (0,1), (-1,-1), 7.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt]),
        ('GRID',           (0,0), (-1,-1), 0.4, bdr),
        ('LINEABOVE',      (0,0), (-1,0),  1.5, navy),
        ('LINEBELOW',      (0,0), (-1,0),  1.5, navy),
        ('ALIGN',          (0,0), (0,-1),  'CENTER'),
        ('ALIGN',          (9,0), (11,-1), 'CENTER'),
        ('VALIGN',         (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING',        (0,0), (-1,-1), 3),
    ]))
    elements.append(tbl)
    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph('This is a computer generated report. — Unity Cement ERP System', sub_sty))

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="MGP_Report_{report_date}.pdf"'
    return response


@login_required
def bulk_action(request):
    if request.method != 'POST':
        return redirect('material_pass:pass_list')
    action = request.POST.get('action')
    ids    = request.POST.getlist('selected_ids')
    if not ids:
        messages.warning(request, 'No passes selected.')
        return redirect('material_pass:pass_list')
    passes = MaterialGatePass.objects.filter(pk__in=ids)
    if action == 'delete':
        if not request.user.is_superuser and not request.user.perm_mgp_write:
            messages.error(request, 'Access Denied.')
            return redirect('material_pass:pass_list')
        count = passes.count()
        passes.delete()
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_bulk_delete', 'MGP', f'{count} MGPs deleted in bulk by {request.user.username}')
        messages.success(request, f'{count} pass(es) deleted.')
    elif action == 'cancel':
        updated = passes.exclude(status__in=['returned', 'rejected']).update(status='rejected')
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_bulk_cancel', 'MGP', f'{updated} MGPs cancelled in bulk by {request.user.username}')
        messages.success(request, f'{updated} pass(es) cancelled.')
    elif action == 'duplicate':
        import uuid as _uuid
        for gp in passes:
            gp.pk              = None
            gp.pass_number     = ''
            gp.status          = 'pending'
            gp.approver        = None
            gp.approved_at     = None
            gp.approval_remarks = ''
            gp.approval_token  = _uuid.uuid4()
            gp.employee        = request.user
            gp.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'mgp_bulk_duplicate', 'MGP', f'{len(ids)} MGPs duplicated in bulk by {request.user.username}')
        messages.success(request, f'{len(ids)} pass(es) duplicated.')
    else:
        messages.error(request, 'Unknown action.')
    return redirect('material_pass:pass_list')


@login_required
def request_bulk_action(request):
    if request.method != 'POST':
        return redirect('material_pass:request_list')
    action = request.POST.get('action')
    ids    = request.POST.getlist('selected_ids')
    if not ids:
        messages.warning(request, 'No requests selected.')
        return redirect('material_pass:request_list')
    requests_qs = MaterialRequest.objects.filter(pk__in=ids)
    if action == 'delete':
        count = requests_qs.count()
        requests_qs.delete()
        from accounts.models import AuditLog
        AuditLog.log(request, 'mr_bulk_delete', 'MR', f'{count} Material Requests deleted in bulk by {request.user.username}')
        messages.success(request, f'{count} request(s) deleted.')
    elif action == 'cancel':
        updated = requests_qs.exclude(status='converted').update(status='rejected')
        from accounts.models import AuditLog
        AuditLog.log(request, 'mr_bulk_cancel', 'MR', f'{updated} Material Requests cancelled in bulk by {request.user.username}')
        messages.success(request, f'{updated} request(s) cancelled.')
    elif action == 'duplicate':
        import uuid as _uuid
        for mr in requests_qs:
            mr.pk          = None
            mr.status      = 'draft'
            mr.reviewed_by = None
            mr.reviewed_at = None
            mr.save()
        from accounts.models import AuditLog
        AuditLog.log(request, 'mr_bulk_duplicate', 'MR', f'{len(ids)} Material Requests duplicated in bulk by {request.user.username}')
        messages.success(request, f'{len(ids)} request(s) duplicated.')
    else:
        messages.error(request, 'Unknown action.')
    return redirect('material_pass:request_list')
