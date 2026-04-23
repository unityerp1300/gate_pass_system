from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count
from django.http import HttpResponse
import csv

from accounts.models import Employee, SystemSetting
from .models import Grievance, GrievanceComment
from .forms import GrievanceForm, GrievanceUpdateForm, GrievanceCommentForm


def _can_manage(user):
    """Management, President/Plant Head, HOD, HR, Admin can manage grievances."""
    manage_roles = ('administrator', 'management', 'president_plant_head', 'department_hod', 'hr')
    return user.is_superuser or any(user.has_role(r) for r in manage_roles)


def _next_grievance_no():
    setting = SystemSetting.get()
    prefix = getattr(setting, 'grv_prefix', 'GRV')
    next_no = getattr(setting, 'grv_next_number', 1)
    # Save incremented number
    SystemSetting.objects.filter(pk=1).update(grv_next_number=next_no + 1)
    fy = timezone.now().year % 100
    return f'{prefix}-{fy:02d}{(fy+1):02d}-{next_no:05d}'


@login_required
def dashboard(request):
    user = request.user
    if _can_manage(user):
        qs = Grievance.objects.all()
    else:
        qs = Grievance.objects.filter(raised_by=user)

    ctx = {
        'total':        qs.count(),
        'open':         qs.filter(status='open').count(),
        'under_review': qs.filter(status='under_review').count(),
        'resolved':     qs.filter(status='resolved').count(),
        'critical':     qs.filter(priority='critical', status__in=['open', 'under_review']).count(),
        'recent':       qs.select_related('raised_by', 'assigned_to')[:10],
        'can_manage':   _can_manage(user),
    }
    return render(request, 'grievance/dashboard.html', ctx)


@login_required
def grievance_list(request):
    user = request.user
    if _can_manage(user):
        qs = Grievance.objects.select_related('raised_by', 'assigned_to')
    else:
        qs = Grievance.objects.filter(raised_by=user).select_related('raised_by', 'assigned_to')

    status   = request.GET.get('status', '')
    category = request.GET.get('category', '')
    priority = request.GET.get('priority', '')
    q        = request.GET.get('q', '')

    if status:   qs = qs.filter(status=status)
    if category: qs = qs.filter(category=category)
    if priority: qs = qs.filter(priority=priority)
    if q:        qs = qs.filter(Q(subject__icontains=q) | Q(grievance_no__icontains=q))

    # Bulk action
    if request.method == 'POST':
        action = request.POST.get('bulk_action')
        ids = request.POST.getlist('selected_ids')
        if ids and action:
            selected = Grievance.objects.filter(pk__in=ids)
            if action == 'delete' and _can_manage(user):
                count = selected.count()
                selected.delete()
                messages.success(request, f'{count} grievance(s) deleted.')
            elif action == 'cancel':
                updated = selected.exclude(status__in=['resolved', 'closed']).update(status='closed')
                messages.success(request, f'{updated} grievance(s) closed.')
            elif action == 'duplicate':
                for g in selected:
                    g.pk = None
                    g.grievance_no = _next_grievance_no()
                    g.status = 'open'
                    g.resolved_at = None
                    g.resolution_note = ''
                    g.save()
                messages.success(request, f'{len(ids)} grievance(s) duplicated.')
            return redirect('grievance:list')

    from django.core.paginator import Paginator
    from .models import CATEGORY_CHOICES, PRIORITY_CHOICES, STATUS_CHOICES
    paginator = Paginator(qs, 15)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    ctx = {
        'grievances':       page_obj,
        'page_obj':         page_obj,
        'can_manage':       _can_manage(user),
        'category_choices': CATEGORY_CHOICES,
        'priority_choices': PRIORITY_CHOICES,
        'status_choices':   STATUS_CHOICES,
        'filters': {'status': status, 'category': category, 'priority': priority, 'q': q},
    }
    return render(request, 'grievance/grievance_list.html', ctx)


@login_required
def grievance_create(request):
    if request.method == 'POST':
        form = GrievanceForm(request.POST, request.FILES)
        if form.is_valid():
            grv = form.save(commit=False)
            grv.raised_by = request.user
            grv.grievance_no = _next_grievance_no()
            grv.save()
            messages.success(request, f'Grievance {grv.grievance_no} raised successfully.')
            return redirect('grievance:detail', pk=grv.pk)
    else:
        form = GrievanceForm()
    return render(request, 'grievance/grievance_form.html', {'form': form, 'title': 'Raise Grievance'})


@login_required
def grievance_detail(request, pk):
    user = request.user
    grv = get_object_or_404(Grievance, pk=pk)

    # Only raiser or managers can view
    if grv.raised_by != user and not _can_manage(user):
        messages.error(request, 'Access denied.')
        return redirect('grievance:list')

    comment_form = GrievanceCommentForm()
    update_form  = GrievanceUpdateForm(instance=grv) if _can_manage(user) else None

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'comment':
            comment_form = GrievanceCommentForm(request.POST)
            if comment_form.is_valid():
                c = comment_form.save(commit=False)
                c.grievance = grv
                c.author = user
                c.save()
                messages.success(request, 'Comment added.')
                return redirect('grievance:detail', pk=pk)
        elif action == 'update' and _can_manage(user):
            update_form = GrievanceUpdateForm(request.POST, instance=grv)
            if update_form.is_valid():
                updated = update_form.save(commit=False)
                if updated.status == 'resolved' and not grv.resolved_at:
                    updated.resolved_at = timezone.now()
                updated.save()
                messages.success(request, 'Grievance updated.')
                return redirect('grievance:detail', pk=pk)

    ctx = {
        'grv':          grv,
        'comments':     grv.comments.select_related('author'),
        'comment_form': comment_form,
        'update_form':  update_form,
        'can_manage':   _can_manage(user),
    }
    return render(request, 'grievance/grievance_detail.html', ctx)


@login_required
def grievance_edit(request, pk):
    grv = get_object_or_404(Grievance, pk=pk)
    user = request.user
    if grv.raised_by != user and not _can_manage(user):
        messages.error(request, 'Access denied.')
        return redirect('grievance:list')
    form = GrievanceForm(request.POST or None, request.FILES or None, instance=grv)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Grievance updated.')
        return redirect('grievance:detail', pk=pk)
    return render(request, 'grievance/grievance_form.html', {'form': form, 'title': 'Edit Grievance', 'grv': grv})


@login_required
def grievance_delete(request, pk):
    if not _can_manage(request.user):
        messages.error(request, 'Access denied.')
        return redirect('grievance:list')
    grv = get_object_or_404(Grievance, pk=pk)
    grv.delete()
    messages.success(request, 'Grievance deleted.')
    return redirect('grievance:list')


@login_required
def grievance_duplicate(request, pk):
    grv = get_object_or_404(Grievance, pk=pk)
    grv.pk = None
    grv.grievance_no = _next_grievance_no()
    grv.status = 'open'
    grv.resolved_at = None
    grv.resolution_note = ''
    grv.save()
    messages.success(request, f'Grievance duplicated as {grv.grievance_no}.')
    return redirect('grievance:detail', pk=grv.pk)


@login_required
def report(request):
    if not _can_manage(request.user):
        messages.error(request, 'Access denied.')
        return redirect('grievance:dashboard')

    qs = Grievance.objects.select_related('raised_by', 'assigned_to')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    status    = request.GET.get('status', '')
    category  = request.GET.get('category', '')
    dept      = request.GET.get('dept', '')

    if date_from: qs = qs.filter(created_at__date__gte=date_from)
    if date_to:   qs = qs.filter(created_at__date__lte=date_to)
    if status:    qs = qs.filter(status=status)
    if category:  qs = qs.filter(category=category)
    if dept:      qs = qs.filter(raised_by__department=dept)

    if request.GET.get('export') == 'csv':
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="grievance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Grievance No', 'Raised By', 'Department', 'Category', 'Priority', 'Subject', 'Status', 'Resolved At', 'Date'])
        for g in qs:
            writer.writerow([
                g.grievance_no, g.raised_by.employee_name,
                g.raised_by.department, g.get_category_display(),
                g.get_priority_display(), g.subject,
                g.get_status_display(),
                g.resolved_at.strftime('%d-%m-%Y %H:%M') if g.resolved_at else '',
                g.created_at.strftime('%d-%m-%Y'),
            ])
        return response

    if request.GET.get('export') == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Grievance Report'
        headers = ['Grievance No', 'Raised By', 'Department', 'Category', 'Priority',
                   'Subject', 'Status', 'Resolved At', 'Date']
        hfill = PatternFill(start_color='7b4bb3', end_color='7b4bb3', fill_type='solid')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = 20
        for g in qs:
            ws.append([
                g.grievance_no, g.raised_by.employee_name,
                g.raised_by.department, g.get_category_display(),
                g.get_priority_display(), g.subject,
                g.get_status_display(),
                g.resolved_at.strftime('%d-%m-%Y %H:%M') if g.resolved_at else '',
                g.created_at.strftime('%d-%m-%Y'),
            ])
        from accounts.report_utils import add_excel_logo_and_note
        add_excel_logo_and_note(ws, len(headers))
        buf = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        buf['Content-Disposition'] = 'attachment; filename="grievance_report.xlsx"'
        from io import BytesIO
        b = BytesIO()
        wb.save(b)
        b.seek(0)
        return HttpResponse(b.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            headers={'Content-Disposition': 'attachment; filename="grievance_report.xlsx"'})

    from accounts.models import DEPARTMENT_CHOICES
    from .models import CATEGORY_CHOICES, STATUS_CHOICES
    ctx = {
        'grievances':       qs,
        'category_choices': CATEGORY_CHOICES,
        'status_choices':   STATUS_CHOICES,
        'dept_choices':     DEPARTMENT_CHOICES,
        'filters': {'date_from': date_from, 'date_to': date_to, 'status': status, 'category': category, 'dept': dept},
        'stats': {
            'total':        qs.count(),
            'open':         qs.filter(status='open').count(),
            'under_review': qs.filter(status='under_review').count(),
            'resolved':     qs.filter(status='resolved').count(),
        },
    }
    return render(request, 'grievance/report.html', ctx)
